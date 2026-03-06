"""
POD Dual-Front Workspace Generator v2
Generates all tracking spreadsheets and folder structure for:
  Front A — Sneaker Culture (primary, app-driven)
  Front B — Generalized Designs (secondary, Etsy organic)

Usage:
    python generate_workspace_v2.py --dir workspace
    python generate_workspace_v2.py --dir workspace --reset
"""

import os
import sys
import json
import argparse
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────
C = {
    "dark_bg":    "1A1A2E", "mid_bg": "16213E", "accent": "0F3460",
    "highlight":  "E94560", "white":  "FFFFFF", "light_gray": "F4F6F9",
    "med_gray":   "CBD5E1", "dark_text": "1E293B",
    "green":  "22C55E", "red":    "EF4444", "amber":  "F59E0B",
    "blue":   "3B82F6", "purple": "8B5CF6",
    # Front identifiers
    "front_a":    "1A1A2E",   # dark navy — sneaker culture
    "front_a_lt": "E8E8F0",   # light navy tint for data rows
    "front_b":    "0F3460",   # deep blue — general
    "front_b_lt": "E8EFF8",   # light blue tint for data rows
    # Standard coding
    "input_blue": "0000FF", "formula_blk": "000000", "link_green": "008000",
}

def style(cell, bold=False, fg=None, bg=None, size=10,
          align="left", wrap=False, italic=False, border=False):
    cell.font = Font(bold=bold, italic=italic,
                     color=fg or C["dark_text"], name="Arial", size=size)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if border:
        thin = Side(style="thin", color=C["med_gray"])
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def section_title(ws, row, col, text, span=10, bg=None):
    cell = ws.cell(row=row, column=col, value=text)
    style(cell, bold=True, fg=C["white"], bg=bg or C["dark_bg"],
          size=12, align="left")
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = 24

def header_row(ws, row, headers, bg=None):
    for i, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=title)
        style(cell, bold=True, fg=C["white"], bg=bg or C["accent"],
              align="center", size=9, border=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 18

def add_dropdown(ws, col_letter, start_row, end_row, options):
    formula = '"' + ','.join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws.add_data_validation(dv)


# ══════════════════════════════════════════════════════════
# 1. FRONT A — DESIGNS TRACKER  (designs_front_a.xlsx)
# ══════════════════════════════════════════════════════════
def create_designs_front_a(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Designs"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "👟  FRONT A — SNEAKER CULTURE DESIGN TRACKER",
                  span=18, bg=C["front_a"])

    headers = [
        ("Design ID", 13), ("Filename", 26), ("Drop ID", 12),
        ("Drop Theme", 18), ("Design Name", 22), ("Slogan / Typography", 32),
        ("Style", 18), ("Date Created", 13), ("Resolution", 13),
        ("Contrast OK?", 13), ("TM Checked?", 13), ("IP Risk", 12),
        ("Brand Ref?", 12), ("Approved?", 12),
        ("Printify Image ID", 18), ("Printify Product ID", 18),
        ("Etsy Listing ID", 18), ("Status", 16),
    ]
    header_row(ws, 2, headers, bg=C["front_a"])

    sample = [
        "DSN-A-0001", "rotation_ready_001.png", "DROP-01", "Rotation Culture",
        "Rotation Ready", "ROTATION READY / WEAR YOUR PAIRS",
        "Bold condensed typography", "2026-03-05", "4500x5400",
        "YES", "YES", "LOW", "NO", "YES", "", "", "", "Pending Upload"
    ]
    for col, val in enumerate(sample, 1):
        cell = ws.cell(row=3, column=col, value=val)
        style(cell, fg=C["dark_text"], bg=C["front_a_lt"], size=9, border=True)

    add_dropdown(ws, "J", 3, 1000, ["YES", "NO", "CHECK"])
    add_dropdown(ws, "K", 3, 1000, ["YES", "NO", "PENDING"])
    add_dropdown(ws, "L", 3, 1000, ["LOW", "MEDIUM", "HIGH", "BLOCKED"])
    add_dropdown(ws, "M", 3, 1000, ["NO", "YES - REMOVE"])
    add_dropdown(ws, "N", 3, 1000, ["YES", "NO", "REVIEW"])
    add_dropdown(ws, "R", 3, 1000,
                 ["Pending Upload", "Uploaded", "Live - Active Drop",
                  "Live - Archived Drop", "Paused", "Rejected"])
    ws.freeze_panes = "A3"

    # Drop Summary sheet
    ws2 = wb.create_sheet("Drop Summary")
    section_title(ws2, 1, 1, "🗓️  DROP SUMMARY", span=6, bg=C["front_a"])
    drop_headers = [("Drop ID", 12), ("Theme", 22), ("# Designs", 12),
                    ("Launch Date", 14), ("Window (hrs)", 14), ("Status", 14)]
    header_row(ws2, 2, drop_headers, bg=C["front_a"])
    for i, drop in enumerate([
        ("DROP-01", "Rotation Culture", 5, "2026-03-05", 72, "Planned"),
        ("DROP-02", "Collector Identity", 5, "", 72, ""),
        ("DROP-03", "Sneakerhead Humour", 5, "", 72, ""),
        ("DROP-04", "Archive Aesthetic", 5, "", 72, ""),
    ], 3):
        for col, val in enumerate(drop, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            style(cell, fg=C["dark_text"],
                  bg=C["front_a_lt"] if i % 2 == 0 else C["white"],
                  size=9, border=True)
    add_dropdown(ws2, "F", 3, 50,
                 ["Planned", "In Production", "Live", "Archived", "Cancelled"])
    ws2.freeze_panes = "A3"

    wb.save(path)
    print("  ✅ designs_front_a.xlsx")


# ══════════════════════════════════════════════════════════
# 2. FRONT B — DESIGNS TRACKER  (designs_front_b.xlsx)
# ══════════════════════════════════════════════════════════
def create_designs_front_b(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Designs"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "🎨  FRONT B — GENERALIZED DESIGN TRACKER",
                  span=17, bg=C["front_b"])

    headers = [
        ("Design ID", 13), ("Filename", 26), ("Niche", 18),
        ("Sub-Niche", 20), ("Phrase / Concept", 32), ("Style", 18),
        ("Date Created", 13), ("Resolution", 13), ("Contrast OK?", 13),
        ("TM Checked?", 13), ("IP Risk", 12), ("Approved?", 12),
        ("Printify Image ID", 18), ("Printify Product ID", 18),
        ("Etsy Listing ID", 18), ("Status", 16), ("Notes", 22),
    ]
    header_row(ws, 2, headers, bg=C["front_b"])

    sample = [
        "DSN-B-0001", "cottage_bloom_001.png", "Cottagecore", "Botanical",
        "Let the wildflowers grow where they will",
        "Minimalist line art", "2026-03-05", "3600x4800",
        "YES", "YES", "LOW", "YES", "", "", "", "Pending Upload", ""
    ]
    for col, val in enumerate(sample, 1):
        cell = ws.cell(row=3, column=col, value=val)
        style(cell, fg=C["dark_text"], bg=C["front_b_lt"], size=9, border=True)

    add_dropdown(ws, "I", 3, 2000, ["YES", "NO", "CHECK"])
    add_dropdown(ws, "J", 3, 2000, ["YES", "NO", "PENDING"])
    add_dropdown(ws, "K", 3, 2000, ["LOW", "MEDIUM", "HIGH", "BLOCKED"])
    add_dropdown(ws, "L", 3, 2000, ["YES", "NO", "REVIEW"])
    add_dropdown(ws, "P", 3, 2000,
                 ["Pending Upload", "Uploaded", "Live on Etsy",
                  "Paused", "Rejected", "Deleted"])
    ws.freeze_panes = "A3"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    section_title(ws2, 1, 1, "📊  FRONT B SUMMARY", span=4, bg=C["front_b"])
    stats = [
        ("Total Designs",   '=COUNTA(Designs!A3:A2000)'),
        ("Approved",        '=COUNTIF(Designs!L3:L2000,"YES")'),
        ("Pending Review",  '=COUNTIF(Designs!L3:L2000,"REVIEW")'),
        ("Live on Etsy",    '=COUNTIF(Designs!P3:P2000,"Live on Etsy")'),
        ("High IP Risk",    '=COUNTIF(Designs!K3:K2000,"HIGH")'),
    ]
    for i, (label, formula) in enumerate(stats, 3):
        ws2.cell(row=i, column=1, value=label)
        style(ws2.cell(row=i, column=1), bold=True,
              fg=C["dark_text"], bg=C["light_gray"], size=10)
        cell = ws2.cell(row=i, column=2, value=formula)
        style(cell, bold=True, fg=C["input_blue"], align="center", size=11)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14

    wb.save(path)
    print("  ✅ designs_front_b.xlsx")


# ══════════════════════════════════════════════════════════
# 3. DROP TRACKER — FRONT A  (drops_front_a.xlsx)
# ══════════════════════════════════════════════════════════
def create_drop_tracker(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Drop Calendar"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "🗓️  FRONT A — DROP CALENDAR & PERFORMANCE",
                  span=14, bg=C["front_a"])

    headers = [
        ("Drop ID", 12), ("Drop Name", 24), ("Theme", 20),
        ("Designs", 10), ("Launch Date", 14), ("End Date", 14),
        ("Window (hrs)", 13), ("App Banner?", 13), ("Community Tab?", 15),
        ("Impressions", 13), ("App Clicks", 12), ("Etsy Redirects", 14),
        ("Orders", 10), ("Revenue ($)", 13),
    ]
    header_row(ws, 2, headers, bg=C["front_a"])

    drops = [
        ("DROP-01", "Rotation Club Collection", "Rotation Culture", 5,
         "", "", 72, "YES", "YES"),
        ("DROP-02", "Collector Identity Series", "Collector Identity", 5,
         "", "", 72, "", ""),
        ("DROP-03", "No Crease Club Drop", "Sneakerhead Humour", 5,
         "", "", 72, "", ""),
        ("DROP-04", "Archive Series Vol.1", "Archive Aesthetic", 5,
         "", "", 72, "", ""),
        ("DROP-05", "Street Culture Drop", "Street Culture", 5,
         "", "", 72, "", ""),
    ]
    for row_i, d in enumerate(drops, 3):
        for col, val in enumerate(d, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            style(cell, fg=C["dark_text"],
                  bg=C["front_a_lt"] if row_i % 2 == 0 else C["white"],
                  size=9, border=True)
        # Revenue formula (orders × avg $50)
        ws.cell(row=row_i, column=14, value=f"=M{row_i}*50")
        ws.cell(row=row_i, column=14).number_format = '$#,##0.00;($#,##0.00);"-"'

    add_dropdown(ws, "H", 3, 50, ["YES", "NO", "SCHEDULED"])
    add_dropdown(ws, "I", 3, 50, ["YES", "NO", "SCHEDULED"])
    ws.freeze_panes = "A3"

    # Per-Drop Performance Detail
    ws2 = wb.create_sheet("Design Performance")
    section_title(ws2, 1, 1, "📊  PER-DESIGN DROP PERFORMANCE", span=10, bg=C["front_a"])
    perf_headers = [
        ("Drop ID", 12), ("Design ID", 13), ("Design Name", 24),
        ("Product Type", 16), ("App Impressions", 16), ("App Clicks", 12),
        ("Etsy Redirects", 14), ("Orders", 10), ("Revenue ($)", 13),
        ("Profit ($)", 13),
    ]
    header_row(ws2, 2, perf_headers, bg=C["front_a"])
    ws2.freeze_panes = "A3"

    wb.save(path)
    print("  ✅ drops_front_a.xlsx")


# ══════════════════════════════════════════════════════════
# 4. APP ANALYTICS TRACKER  (app_analytics.xlsx)
# ══════════════════════════════════════════════════════════
def create_app_analytics(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "App Events"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "📱  FRONT A — APP ANALYTICS TRACKER",
                  span=12, bg=C["front_a"])

    headers = [
        ("Date", 13), ("Drop ID", 12), ("Design Name", 24),
        ("Placement", 20), ("Impressions", 13), ("Clicks", 10),
        ("CTR %", 10), ("Etsy Redirects", 14), ("Redirect Rate %", 15),
        ("Orders (est.)", 13), ("Revenue ($)", 13), ("Notes", 22),
    ]
    header_row(ws, 2, headers, bg=C["front_a"])

    # Sample row
    r = 3
    ws.cell(row=r, column=1, value="2026-03-05")
    ws.cell(row=r, column=2, value="DROP-01")
    ws.cell(row=r, column=3, value="Rotation Ready")
    ws.cell(row=r, column=4, value="community_tab")
    ws.cell(row=r, column=5, value=320)
    ws.cell(row=r, column=6, value=18)
    ws.cell(row=r, column=7, value="=IF(E3>0,F3/E3,0)")
    ws.cell(row=r, column=7).number_format = "0.0%"
    ws.cell(row=r, column=8, value=12)
    ws.cell(row=r, column=9, value="=IF(F3>0,H3/F3,0)")
    ws.cell(row=r, column=9).number_format = "0.0%"
    ws.cell(row=r, column=10, value=2)
    ws.cell(row=r, column=11, value=110.00)
    ws.cell(row=r, column=11).number_format = '$#,##0.00;($#,##0.00);"-"'
    for col in range(1, 13):
        style(ws.cell(row=r, column=col),
              fg=C["dark_text"], bg=C["front_a_lt"], size=9, border=True)

    add_dropdown(ws, "D", 3, 2000,
                 ["community_tab", "sneaker_detail_page",
                  "drop_page", "limited_drop_banner", "profile_page"])
    ws.freeze_panes = "A3"

    # Monthly KPI Summary
    ws2 = wb.create_sheet("Monthly KPIs")
    section_title(ws2, 1, 1, "📈  MONTHLY APP → MERCH KPIs", span=9, bg=C["front_a"])
    kpi_headers = [
        ("Month", 12), ("Total Users (est.)", 18), ("Merch Impressions", 18),
        ("Total Clicks", 13), ("Overall CTR %", 14), ("Etsy Redirects", 14),
        ("Redirect Rate %", 15), ("Orders", 10), ("Revenue ($)", 13),
    ]
    header_row(ws2, 2, kpi_headers, bg=C["front_a"])
    year = datetime.now().year
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    for i, m in enumerate(months, 3):
        ws2.cell(row=i, column=1, value=f"{m} {year}")
        ws2.cell(row=i, column=5,
                 value=f"=IF(C{i}>0,D{i}/C{i},0)").number_format = "0.0%"
        ws2.cell(row=i, column=7,
                 value=f"=IF(D{i}>0,F{i}/D{i},0)").number_format = "0.0%"
        ws2.cell(row=i, column=9, value=0.0).number_format = '$#,##0.00;($#,##0.00);"-"'
        bg = C["front_a_lt"] if i % 2 == 0 else C["white"]
        for col in range(1, 10):
            style(ws2.cell(row=i, column=col), fg=C["dark_text"],
                  bg=bg, size=9, border=True)

    # Rev/1000 users KPI target row
    ws2.cell(row=16, column=1, value="Target: Revenue / 1,000 users")
    ws2.cell(row=16, column=2, value=">$50/month")
    style(ws2.cell(row=16, column=1), bold=True, italic=True,
          fg=C["amber"], size=9)
    style(ws2.cell(row=16, column=2), bold=True,
          fg=C["green"], size=9)

    wb.save(path)
    print("  ✅ app_analytics.xlsx")


# ══════════════════════════════════════════════════════════
# 5. COMBINED SALES TRACKER  (sales.xlsx)
# ══════════════════════════════════════════════════════════
def create_sales_tracker(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "💰  COMBINED SALES TRACKER — BOTH FRONTS", span=20)

    headers = [
        ("Order ID", 14), ("Date", 13), ("Front", 10),
        ("Drop ID", 12), ("Etsy Listing ID", 18), ("Product Title", 30),
        ("Niche / Design", 20), ("Sale Price ($)", 14), ("Shipping Charged", 14),
        ("Printify Cost ($)", 16), ("Printify Ship ($)", 16),
        ("Etsy Txn Fee", 13), ("Etsy Payment Fee", 15),
        ("Etsy Listing Fee", 15), ("Etsy Ads Spend", 14),
        ("Total Costs ($)", 14), ("Net Profit ($)", 14),
        ("Margin %", 12), ("UTM Source", 18), ("Notes", 20),
    ]
    header_row(ws, 2, headers)

    r = 3
    ws.cell(row=r, column=1, value="ORD-0001")
    ws.cell(row=r, column=2, value="2026-03-05")
    ws.cell(row=r, column=3, value="A")
    ws.cell(row=r, column=4, value="DROP-01")
    ws.cell(row=r, column=6, value="Rotation Ready Shirt | Sneakerhead Tee")
    ws.cell(row=r, column=7, value="Sneaker Culture")
    ws.cell(row=r, column=8, value=54.99)
    ws.cell(row=r, column=9, value=0.00)
    ws.cell(row=r, column=10, value=14.50)
    ws.cell(row=r, column=11, value=5.00)
    ws.cell(row=r, column=12, value="=H3*0.065")
    ws.cell(row=r, column=13, value="=H3*0.03+0.25")
    ws.cell(row=r, column=14, value=0.20)
    ws.cell(row=r, column=15, value=0.00)
    ws.cell(row=r, column=16, value="=J3+K3+L3+M3+N3+O3")
    ws.cell(row=r, column=17, value="=H3+I3-P3")
    ws.cell(row=r, column=18, value="=IF((H3+I3)>0,Q3/(H3+I3),0)")
    ws.cell(row=r, column=18).number_format = "0.0%"
    ws.cell(row=r, column=19, value="sneakerapp/community_tab")
    for col in [8,9,10,11,12,13,14,15,16,17]:
        ws.cell(row=r, column=col).number_format = '$#,##0.00;($#,##0.00);"-"'
    for col in range(1, 21):
        bg = C["front_a_lt"] if col <= 4 else C["light_gray"] if col % 2 else C["white"]
        style(ws.cell(row=r, column=col), fg=C["dark_text"], bg=bg, size=9, border=True)

    add_dropdown(ws, "C", 3, 3000, ["A", "B"])
    ws.freeze_panes = "A3"

    # Monthly P&L by Front
    ws2 = wb.create_sheet("Monthly P&L")
    section_title(ws2, 1, 1, "📅  MONTHLY P&L — BY FRONT", span=9)
    year = datetime.now().year
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]

    # Front A block
    ws2.cell(row=2, column=1, value="FRONT A — SNEAKER CULTURE")
    style(ws2.cell(row=2, column=1), bold=True, fg=C["white"],
          bg=C["front_a"], size=11)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    pnl_headers = [("Month",12),("Orders",10),("Gross Revenue",14),
                   ("Costs",12),("Net Profit",14),("Margin%",10),
                   ("Avg Order",12),("Best Design",22),("Drop ID",12)]
    header_row(ws2, 3, pnl_headers, bg=C["front_a"])
    for i, m in enumerate(months, 4):
        ws2.cell(row=i, column=1, value=f"{m} {year}")
        ws2.cell(row=i, column=2,
            value=f'=SUMPRODUCT((MONTH(Orders!B$3:B$3000)={i-3})*(YEAR(Orders!B$3:B$3000)={year})*(Orders!C$3:C$3000="A"))')
        ws2.cell(row=i, column=3,
            value=f'=SUMPRODUCT((MONTH(Orders!B$3:B$3000)={i-3})*(YEAR(Orders!B$3:B$3000)={year})*(Orders!C$3:C$3000="A")*Orders!H$3:H$3000)')
        ws2.cell(row=i, column=4,
            value=f'=SUMPRODUCT((MONTH(Orders!B$3:B$3000)={i-3})*(YEAR(Orders!B$3:B$3000)={year})*(Orders!C$3:C$3000="A")*Orders!P$3:P$3000)')
        ws2.cell(row=i, column=5, value=f"=C{i}-D{i}")
        ws2.cell(row=i, column=6, value=f"=IF(C{i}>0,E{i}/C{i},0)")
        ws2.cell(row=i, column=6).number_format = "0.0%"
        ws2.cell(row=i, column=7, value=f"=IF(B{i}>0,C{i}/B{i},0)")
        for col in [3,4,5,7]:
            ws2.cell(row=i, column=col).number_format = '$#,##0.00;($#,##0.00);"-"'
        bg = C["front_a_lt"] if i % 2 == 0 else C["white"]
        for col in range(1, 10):
            style(ws2.cell(row=i, column=col), fg=C["dark_text"],
                  bg=bg, size=9, border=True)

    # Front B block
    offset = 18
    ws2.cell(row=offset, column=1, value="FRONT B — GENERALIZED DESIGNS")
    style(ws2.cell(row=offset, column=1), bold=True, fg=C["white"],
          bg=C["front_b"], size=11)
    ws2.merge_cells(start_row=offset, start_column=1,
                    end_row=offset, end_column=9)
    header_row(ws2, offset+1, pnl_headers, bg=C["front_b"])
    for i, m in enumerate(months, offset+2):
        ws2.cell(row=i, column=1, value=f"{m} {year}")
        ws2.cell(row=i, column=2,
            value=f'=SUMPRODUCT((MONTH(Orders!B$3:B$3000)={i-offset-1})*(YEAR(Orders!B$3:B$3000)={year})*(Orders!C$3:C$3000="B"))')
        ws2.cell(row=i, column=3,
            value=f'=SUMPRODUCT((MONTH(Orders!B$3:B$3000)={i-offset-1})*(YEAR(Orders!B$3:B$3000)={year})*(Orders!C$3:C$3000="B")*Orders!H$3:H$3000)')
        ws2.cell(row=i, column=4,
            value=f'=SUMPRODUCT((MONTH(Orders!B$3:B$3000)={i-offset-1})*(YEAR(Orders!B$3:B$3000)={year})*(Orders!C$3:C$3000="B")*Orders!P$3:P$3000)')
        ws2.cell(row=i, column=5, value=f"=C{i}-D{i}")
        ws2.cell(row=i, column=6, value=f"=IF(C{i}>0,E{i}/C{i},0)")
        ws2.cell(row=i, column=6).number_format = "0.0%"
        ws2.cell(row=i, column=7, value=f"=IF(B{i}>0,C{i}/B{i},0)")
        for col in [3,4,5,7]:
            ws2.cell(row=i, column=col).number_format = '$#,##0.00;($#,##0.00);"-"'
        bg = C["front_b_lt"] if i % 2 == 0 else C["white"]
        for col in range(1, 10):
            style(ws2.cell(row=i, column=col), fg=C["dark_text"],
                  bg=bg, size=9, border=True)

    ws2.freeze_panes = "A3"
    wb.save(path)
    print("  ✅ sales.xlsx")


# ══════════════════════════════════════════════════════════
# 6. LISTINGS PERFORMANCE TRACKER  (listings.xlsx)
# ══════════════════════════════════════════════════════════
def create_listings_tracker(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "📋  LISTING PERFORMANCE — BOTH FRONTS", span=22)

    headers = [
        ("Front", 8), ("Drop ID", 12), ("Etsy Listing ID", 16),
        ("Design ID", 13), ("Title", 38), ("Niche / Theme", 18),
        ("Product Type", 16), ("Price ($)", 10), ("Cost ($)", 10),
        ("Profit/Sale ($)", 14), ("Listed Date", 13),
        ("Impressions", 13), ("Clicks", 10), ("CTR %", 10),
        ("Sales", 10), ("CVR %", 10), ("Revenue ($)", 13),
        ("Ad Spend ($)", 13), ("ROAS", 10), ("Status", 14), ("Action Flag", 20),
    ]
    header_row(ws, 2, headers)

    r = 3
    static = {
        1: "A", 2: "DROP-01", 3: "", 4: "DSN-A-0001",
        5: "Rotation Ready Shirt | Sneakerhead T Shirt | Sneaker Collector Tee",
        6: "Sneaker Culture", 7: "Oversized Tee",
        8: 54.99, 9: 14.50, 11: "2026-03-05",
        12: 380, 13: 22, 15: 3, 17: 164.97, 18: 2.50
    }
    for col, val in static.items():
        cell = ws.cell(row=r, column=col, value=val)
        bg = C["front_a_lt"] if col <= 2 else (
             C["light_gray"] if col % 2 else C["white"])
        style(cell, fg=C["input_blue"] if isinstance(val, float) else C["dark_text"],
              bg=bg, size=9, border=True)
    ws.cell(row=r, column=10, value="=H3-I3")
    ws.cell(row=r, column=14, value="=IF(L3>0,M3/L3,0)").number_format = "0.0%"
    ws.cell(row=r, column=16, value="=IF(M3>0,O3/M3,0)").number_format = "0.0%"
    ws.cell(row=r, column=19, value="=IF(R3>0,Q3/R3,0)").number_format = "0.0x"
    ws.cell(row=r, column=21,
            value='=IF(L3<200,"⚠️ Low Impressions",IF(N3<0.005,"📸 Fix Thumbnail",IF(P3<0.008,"📝 Fix Description",IF(S3>0,IF(S3<2,"❌ Pause Ads","✅ Scale"),"➕ Add Ads"))))')

    add_dropdown(ws, "A", 3, 2000, ["A", "B"])
    add_dropdown(ws, "G", 3, 2000,
                 ["Oversized Tee", "Heavyweight Hoodie", "Crewneck",
                  "Dad Hat", "T-Shirt", "Mug", "Tote Bag", "Other"])
    add_dropdown(ws, "T", 3, 2000,
                 ["Live - Active", "Live - Archived", "Paused", "Draft", "Deleted"])
    ws.freeze_panes = "A3"
    wb.save(path)
    print("  ✅ listings.xlsx")


# ══════════════════════════════════════════════════════════
# 7. TRADEMARK LOG  (trademark_log.xlsx) — shared both fronts
# ══════════════════════════════════════════════════════════
def create_trademark_log(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clearance Log"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "⚖️  TRADEMARK & COPYRIGHT CLEARANCE LOG — BOTH FRONTS",
                  span=13)

    headers = [
        ("Check ID", 12), ("Front", 8), ("Design ID", 13),
        ("Phrase / Element", 35), ("Check Date", 13),
        ("USPTO Result", 16), ("Google Check", 14),
        ("Etsy Search Check", 16), ("Substring Match", 15),
        ("Brand Ref?", 12), ("Risk Level", 12),
        ("Decision", 12), ("Notes", 30),
    ]
    header_row(ws, 2, headers)

    sample = ["TM-0001", "A", "DSN-A-0001",
              "ROTATION READY", "2026-03-05", "No match found",
              "No match found", "< 100 results", "None", "NO", "LOW",
              "APPROVED", "Original phrase — sneaker culture context"]
    for col, val in enumerate(sample, 1):
        cell = ws.cell(row=3, column=col, value=val)
        bg = C["front_a_lt"] if col <= 2 else (
             C["light_gray"] if col % 2 else C["white"])
        style(cell, fg=C["dark_text"], bg=bg, size=9, border=True)

    add_dropdown(ws, "B", 3, 3000, ["A", "B"])
    add_dropdown(ws, "F", 3, 3000,
                 ["No match found","Match - LIVE","Match - DEAD",
                  "Match - Class specific","ERROR - check manually"])
    add_dropdown(ws, "I", 3, 3000, ["None","Partial match","Full match"])
    add_dropdown(ws, "J", 3, 3000, ["NO","YES - BLOCKED"])
    add_dropdown(ws, "K", 3, 3000, ["LOW","MEDIUM","HIGH","BLOCKED"])
    add_dropdown(ws, "L", 3, 3000,
                 ["APPROVED","REJECTED","MANUAL REVIEW","PENDING"])
    ws.freeze_panes = "A3"
    wb.save(path)
    print("  ✅ trademark_log.xlsx")


# ══════════════════════════════════════════════════════════
# 8. NICHE TRACKER — FRONT B  (niches_front_b.xlsx)
# ══════════════════════════════════════════════════════════
def create_niche_tracker(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Niches"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "🔍  FRONT B — NICHE RESEARCH TRACKER",
                  span=16, bg=C["front_b"])

    headers = [
        ("Niche ID", 12), ("Niche", 22), ("Sub-Niche", 24),
        ("Etsy Results", 14), ("Top Seller Reviews", 18),
        ("Google Trend", 14), ("FB Group Size", 15), ("Reddit Size", 15),
        ("Avg Price ($)", 13), ("Est. Profit ($)", 15),
        ("Competition", 14), ("Demand", 12), ("Gift Potential", 14),
        ("Evergreen?", 12), ("Priority Score", 14), ("Decision", 12),
    ]
    header_row(ws, 2, headers, bg=C["front_b"])

    samples = [
        ["NCH-001","Occupation","ICU Nurse",8200,450,"Stable","125000","82k",
         29.99,11.50,"Low","High","Yes","Yes"],
        ["NCH-002","Pet","Dachshund",4100,220,"Growing","380000","55k",
         27.99,10.80,"Low","High","Yes","Yes"],
        ["NCH-003","Hobby","Disc Golf",1800,85,"Growing","42000","320k",
         26.99,10.20,"Very Low","Medium","Yes","Yes"],
    ]
    comp_map = {"Very Low":5,"Low":4,"Medium":3,"High":2,"Very High":1}
    demand_map = {"Very High":5,"High":4,"Medium":3,"Low":2,"Very Low":1}
    for row_i, s in enumerate(samples, 3):
        for col, val in enumerate(s, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            bg = C["front_b_lt"] if row_i % 2 == 0 else C["white"]
            style(cell, fg=C["input_blue"] if isinstance(val, float) else C["dark_text"],
                  bg=bg, size=9, border=True)
        score = (demand_map.get(s[11],3) *
                 (2 if s[12]=="Yes" else 1) *
                 comp_map.get(s[10],3))
        ws.cell(row=row_i, column=15, value=score)
        style(ws.cell(row=row_i, column=15), bold=True, fg=C["white"],
              bg=C["green"] if score>=16 else (C["amber"] if score>=8 else C["red"]),
              align="center", size=10)
        ws.cell(row=row_i, column=16, value="PURSUE" if score>=8 else "SKIP")
        style(ws.cell(row=row_i, column=16), bold=True, fg=C["white"],
              bg=C["green"] if score>=8 else C["red"], align="center", size=9)

    add_dropdown(ws, "F", 3, 500,
                 ["Growing","Stable","Declining","Seasonal","Unknown"])
    add_dropdown(ws, "K", 3, 500,
                 ["Very Low","Low","Medium","High","Very High"])
    add_dropdown(ws, "L", 3, 500,
                 ["Very High","High","Medium","Low","Very Low"])
    add_dropdown(ws, "M", 3, 500, ["Yes","No","Partial"])
    add_dropdown(ws, "N", 3, 500, ["Yes","No","Seasonal"])
    ws.freeze_panes = "A3"

    # Phrase Bank
    ws2 = wb.create_sheet("Phrase Bank")
    section_title(ws2, 1, 1, "💬  SAFE PHRASE BANK — AUTO-POPULATED", span=8,
                  bg=C["front_b"])
    p_headers = [("Phrase ID",12),("Niche",16),("Sub-Niche",20),("Phrase",45),
                 ("TM Status",14),("Risk",10),("Design ID",18),("Date Added",13)]
    header_row(ws2, 2, p_headers, bg=C["front_b"])
    add_dropdown(ws2, "E", 3, 3000, ["SAFE","FLAGGED","BLOCKED","PENDING"])
    add_dropdown(ws2, "F", 3, 3000, ["LOW","MEDIUM","HIGH","BLOCKED"])
    ws2.freeze_panes = "A3"

    wb.save(path)
    print("  ✅ niches_front_b.xlsx")


# ══════════════════════════════════════════════════════════
# 9. FINANCIALS  (financials.xlsx) — dual-front P&L
# ══════════════════════════════════════════════════════════
def create_financials(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.row_dimensions[1].height = 30
    section_title(ws, 1, 1, "🧾  EXPENSES TRACKER — BOTH FRONTS", span=9)

    headers = [
        ("Expense ID",9),("Date",13),("Front",10),("Category",20),
        ("Description",35),("Amount ($)",13),
        ("Tax Deductible?",16),("Receipt?",12),("Notes",25),
    ]
    header_row(ws, 2, headers)

    expenses = [
        ["EXP-001","2026-03-05","A","Listing Fees","Front A — 10 listings",2.00,"YES","NO",""],
        ["EXP-002","2026-03-05","B","Listing Fees","Front B — 40 listings",8.00,"YES","NO",""],
        ["EXP-003","2026-03-05","Both","Subscriptions","Printify Premium",29.00,"YES","NO",""],
        ["EXP-004","2026-03-05","A","Advertising","Etsy Ads — Front A",20.00,"YES","NO",""],
        ["EXP-005","2026-03-05","B","Advertising","Etsy Ads — Front B",30.00,"YES","NO",""],
        ["EXP-006","2026-03-05","A","Samples","Front A product samples",45.00,"YES","YES","Quality check"],
        ["EXP-007","2026-03-05","Both","Software","Canva Pro",15.00,"YES","NO",""],
        ["EXP-008","2026-03-05","B","Software","Ideogram Pro",8.00,"YES","NO",""],
    ]
    for row_i, exp in enumerate(expenses, 3):
        for col, val in enumerate(exp, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            bg = C["front_a_lt"] if exp[2]=="A" else (
                 C["front_b_lt"] if exp[2]=="B" else C["light_gray"])
            style(cell, fg=C["input_blue"] if isinstance(val, float) else C["dark_text"],
                  bg=bg, size=9, border=True)

    add_dropdown(ws, "C", 3, 1000, ["A","B","Both"])
    add_dropdown(ws, "D", 3, 1000,
                 ["Listing Fees","Advertising","Samples","Subscriptions",
                  "Software/Tools","Shipping","Equipment","Education","Other"])
    add_dropdown(ws, "G", 3, 1000, ["YES","NO","PARTIAL"])
    add_dropdown(ws, "H", 3, 1000, ["YES","NO"])
    ws.freeze_panes = "A3"

    # P&L Summary
    ws2 = wb.create_sheet("P&L Summary")
    section_title(ws2, 1, 1, "📈  ANNUAL P&L — BOTH FRONTS COMBINED", span=8)
    pnl_headers = [("Month",12),("Front A Revenue",16),("Front B Revenue",16),
                   ("Total Revenue",14),("Total Costs",14),
                   ("Front A Profit",16),("Front B Profit",16),("Net Profit",13)]
    header_row(ws2, 2, pnl_headers)
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    year = datetime.now().year
    for i, m in enumerate(months, 3):
        ws2.cell(row=i, column=1, value=f"{m} {year}")
        for col in range(2, 9):
            ws2.cell(row=i, column=col, value=0.00)
            ws2.cell(row=i, column=col).number_format = '$#,##0.00;($#,##0.00);"-"'
        ws2.cell(row=i, column=4, value=f"=B{i}+C{i}")
        ws2.cell(row=i, column=8, value=f"=F{i}+G{i}")
        bg = C["light_gray"] if i%2==0 else C["white"]
        for col in range(1, 9):
            style(ws2.cell(row=i, column=col), fg=C["dark_text"],
                  bg=bg, size=9, border=True)

    # Totals
    r = 15
    ws2.cell(row=r, column=1, value="FULL YEAR")
    for col in range(2, 9):
        ws2.cell(row=r, column=col,
                 value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}14)")
        ws2.cell(row=r, column=col).number_format = '$#,##0.00;($#,##0.00);"-"'
        style(ws2.cell(row=r, column=col), bold=True, fg=C["white"],
              bg=C["highlight"], align="center", size=10, border=True)
    style(ws2.cell(row=r, column=1), bold=True, fg=C["white"],
          bg=C["highlight"], align="center", size=10, border=True)

    # Tax Prep
    ws3 = wb.create_sheet("Tax Prep")
    section_title(ws3, 1, 1,
                  "🧮  TAX PREP — Set aside 25-30% of every payout", span=4)
    tax_rows = [
        ("Total Revenue (Year)", "='P&L Summary'!D15"),
        ("Total Deductible Expenses",
         "=SUMIF(Expenses!G3:G1000,\"YES\",Expenses!F3:F1000)"),
        ("Taxable Income Est.", "=B3-B4"),
        ("Tax Set-Aside (28%)", "=B5*0.28"),
        ("Actual Tax Paid", 0.00),
        ("Difference", "=B6-B7"),
    ]
    for i, (label, val) in enumerate(tax_rows, 3):
        ws3.cell(row=i, column=1, value=label)
        ws3.cell(row=i, column=2, value=val)
        style(ws3.cell(row=i, column=1), bold=True,
              fg=C["dark_text"], bg=C["light_gray"], size=10)
        style(ws3.cell(row=i, column=2), bold=True,
              fg=C["input_blue"], align="center", size=11)
        ws3.cell(row=i, column=2).number_format = '$#,##0.00;($#,##0.00);"-"'
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 16

    wb.save(path)
    print("  ✅ financials.xlsx")


# ══════════════════════════════════════════════════════════
# UPDATE WORKBOOKS SCRIPT
# ══════════════════════════════════════════════════════════
UPDATE_SCRIPT = '''"""
update_workbooks.py  —  Dual-Front Edition
Auto-populates spreadsheets from pipeline JSON output.

Usage:
    python update_workbooks.py --log logs/front_a_batch_20260305.json --front A
    python update_workbooks.py --log logs/front_b_batch_20260305.json --front B
    python update_workbooks.py --log logs/batch.json --front A --dry-run
"""

import json
import argparse
import os
from datetime import date
from openpyxl import load_workbook

WORKSPACE = os.path.dirname(os.path.abspath(__file__))
SP = os.path.join(WORKSPACE, "spreadsheets")

DESIGNS_A  = os.path.join(SP, "designs_front_a.xlsx")
DESIGNS_B  = os.path.join(SP, "designs_front_b.xlsx")
TM_LOG     = os.path.join(SP, "trademark_log.xlsx")
NICHES_B   = os.path.join(SP, "niches_front_b.xlsx")


def next_id(ws, prefix, col=1, start_row=3):
    last = 0
    for row in ws.iter_rows(min_row=start_row, max_col=col, values_only=True):
        val = row[0]
        if val and str(val).startswith(prefix):
            try:
                num = int(str(val).split("-")[-1])
                last = max(last, num)
            except (IndexError, ValueError):
                pass
    return f"{prefix}{str(last + 1).zfill(4)}"


def append_design_a(ws, record):
    dsn_id = next_id(ws, "DSN-A-")
    ws.append([
        dsn_id, record.get("filename",""), record.get("drop_id",""),
        record.get("drop_theme",""), record.get("design_name",""),
        record.get("slogan",""), record.get("style",""),
        str(date.today()), record.get("resolution",""),
        "YES" if record.get("contrast_ok") else "NO",
        "YES" if record.get("tm_checked") else "PENDING",
        record.get("ip_risk","LOW"),
        "YES" if record.get("brand_ref") else "NO",
        "YES" if record.get("approved") else "REVIEW",
        record.get("printify_image_id",""), record.get("printify_product_id",""),
        record.get("etsy_listing_id",""), record.get("status","Pending Upload"),
    ])
    return dsn_id


def append_design_b(ws, record):
    dsn_id = next_id(ws, "DSN-B-")
    ws.append([
        dsn_id, record.get("filename",""), record.get("niche",""),
        record.get("sub_niche",""), record.get("phrase",""),
        record.get("style",""), str(date.today()),
        record.get("resolution",""),
        "YES" if record.get("contrast_ok") else "NO",
        "YES" if record.get("tm_checked") else "PENDING",
        record.get("ip_risk","LOW"),
        "YES" if record.get("approved") else "REVIEW",
        record.get("printify_image_id",""), record.get("printify_product_id",""),
        record.get("etsy_listing_id",""),
        record.get("status","Pending Upload"), record.get("notes",""),
    ])
    return dsn_id


def append_tm_log(ws, record, dsn_id, front):
    tm_id = next_id(ws, "TM-")
    phrase = record.get("slogan") or record.get("phrase","")
    ws.append([
        tm_id, front, dsn_id, phrase, str(date.today()),
        record.get("uspto_result","No match found"),
        record.get("google_check",""), record.get("etsy_check",""),
        record.get("substring_match","None"),
        "YES" if record.get("brand_ref") else "NO",
        record.get("ip_risk","LOW"),
        "APPROVED" if record.get("approved") else "MANUAL REVIEW",
        record.get("tm_notes",""),
    ])


def append_phrase_b(ws, record, dsn_id):
    phr_id = next_id(ws, "PHR-")
    ws.append([
        phr_id, record.get("niche",""), record.get("sub_niche",""),
        record.get("phrase",""),
        "SAFE" if record.get("approved") else "FLAGGED",
        record.get("ip_risk","LOW"), dsn_id, str(date.today()),
    ])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--log", required=True)
    parser.add_argument("--front", required=True, choices=["A","B"])
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    with open(args.log) as f:
        data = json.load(f)
    records = data if isinstance(data, list) else [data]

    tm_wb    = load_workbook(TM_LOG)
    tm_ws    = tm_wb["Clearance Log"]

    if args.front == "A":
        dsn_wb = load_workbook(DESIGNS_A)
        dsn_ws = dsn_wb["Designs"]
    else:
        dsn_wb   = load_workbook(DESIGNS_B)
        dsn_ws   = dsn_wb["Designs"]
        niche_wb = load_workbook(NICHES_B)
        phrase_ws = niche_wb["Phrase Bank"]

    added = 0
    for record in records:
        if not record.get("filename"):
            print(f"  ⚠️  Skipping — no filename: {record}")
            continue

        if args.front == "A":
            dsn_id = append_design_a(dsn_ws, record)
        else:
            dsn_id = append_design_b(dsn_ws, record)
            if record.get("phrase"):
                append_phrase_b(phrase_ws, record, dsn_id)

        append_tm_log(tm_ws, record, dsn_id, args.front)
        added += 1
        icon = "✅" if record.get("approved") else "⚠️ "
        name = record.get("design_name") or record.get("phrase","?")
        print(f"  {icon} [{args.front}] {dsn_id} — {name}")

    if not args.dry_run:
        dsn_wb.save(DESIGNS_A if args.front == "A" else DESIGNS_B)
        tm_wb.save(TM_LOG)
        if args.front == "B":
            niche_wb.save(NICHES_B)
        print(f"\\n  ✅ {added} records written to spreadsheets.")
    else:
        print(f"\\n  Dry run — {added} records would be written.")


if __name__ == "__main__":
    main()
'''

# ══════════════════════════════════════════════════════════
# EXAMPLE JSON — BOTH FRONTS
# ══════════════════════════════════════════════════════════
EXAMPLE_JSON_A = json.dumps([
    {
        "filename": "rotation_ready_001.png",
        "drop_id": "DROP-01",
        "drop_theme": "Rotation Culture",
        "design_name": "Rotation Ready",
        "slogan": "ROTATION READY / WEAR YOUR PAIRS",
        "style": "Bold condensed typography",
        "resolution": "4500x5400",
        "contrast_ok": True,
        "tm_checked": True,
        "ip_risk": "LOW",
        "brand_ref": False,
        "approved": True,
        "status": "Pending Upload",
        "uspto_result": "No match found",
        "google_check": "No match found",
        "etsy_check": "< 100 results",
        "substring_match": "None",
        "tm_notes": "Original streetwear phrase — no TM risk"
    }
], indent=2)

EXAMPLE_JSON_B = json.dumps([
    {
        "filename": "cottage_bloom_001.png",
        "niche": "Cottagecore",
        "sub_niche": "Botanical",
        "phrase": "Let the wildflowers grow where they will",
        "style": "Minimalist line art",
        "resolution": "3600x4800",
        "contrast_ok": True,
        "tm_checked": True,
        "ip_risk": "LOW",
        "approved": True,
        "status": "Pending Upload",
        "uspto_result": "No match found",
        "google_check": "No match found",
        "etsy_check": "< 200 results",
        "substring_match": "None",
        "tm_notes": "Original phrase — no TM risk"
    }
], indent=2)

README = f"""# POD Dual-Front Workspace
## Front A — Sneaker Culture (Primary) | Front B — Generalized Designs (Secondary)
Generated: {date.today().isoformat()}

## Folder Structure

```
workspace/
├── front_a_sneaker/
│   ├── designs/        ← Drop PNG files for sneaker culture designs
│   ├── approved/
│   ├── rejected/
│   └── drops/          ← Drop metadata JSON files
├── front_b_general/
│   ├── designs/        ← PNG files for generalized niche designs
│   ├── approved/
│   └── rejected/
├── spreadsheets/
│   ├── designs_front_a.xlsx     ← Front A design registry (auto-populated)
│   ├── designs_front_b.xlsx     ← Front B design registry (auto-populated)
│   ├── sales.xlsx               ← Combined orders, split P&L by front
│   ├── listings.xlsx            ← All listings with front tag + action flags
│   ├── trademark_log.xlsx       ← Shared IP clearance log (auto-populated)
│   ├── drops_front_a.xlsx       ← Drop calendar + 72hr window tracker
│   ├── app_analytics.xlsx       ← App CTR, redirects, revenue per 1k users
│   ├── niches_front_b.xlsx      ← Niche scoring + phrase bank
│   └── financials.xlsx          ← Expenses, P&L, tax prep
├── logs/
├── update_workbooks.py
├── pipeline_output_example_front_a.json
├── pipeline_output_example_front_b.json
└── WORKSPACE_README.md
```

## Quick Start

```bash
# 1. Generate workspace
python generate_workspace_v2.py --dir workspace

# 2. After a Front A design batch
python update_workbooks.py --log logs/front_a_batch_YYYYMMDD.json --front A

# 3. After a Front B design batch
python update_workbooks.py --log logs/front_b_batch_YYYYMMDD.json --front B

# 4. Rebuild from scratch
python generate_workspace_v2.py --dir workspace --reset
```

## Spreadsheet Guide

| File | Front | Auto-populated? | Update frequency |
|---|---|---|---|
| designs_front_a.xlsx | A | ✅ Yes | After each drop batch |
| designs_front_b.xlsx | B | ✅ Yes | After each niche batch |
| sales.xlsx | Both | Manual | After each Etsy payout |
| listings.xlsx | Both | Manual | Weekly from Etsy Stats |
| trademark_log.xlsx | Both | ✅ Yes | After each batch |
| drops_front_a.xlsx | A | Manual | Before each drop launch |
| app_analytics.xlsx | A | Manual | Weekly from app dashboard |
| niches_front_b.xlsx | B | ✅ Phrase Bank | After each batch |
| financials.xlsx | Both | Manual | Monthly |

## Colour Code
- 🟣 Dark navy rows = Front A (Sneaker Culture)
- 🔵 Dark blue rows = Front B (Generalized)
- Blue text = manual inputs
- Black text = formulas (do not overwrite)
- 🟢 Green = approved / scale / good
- 🟡 Amber = review needed
- 🔴 Red = blocked / problem
"""


# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Generate dual-front POD workspace")
    parser.add_argument("--dir", default="workspace")
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()

    base = os.path.abspath(args.dir)
    if args.reset and os.path.exists(base):
        import shutil; shutil.rmtree(base)
        print(f"  🗑️  Reset: removed {base}")

    folders = [
        os.path.join(base, "front_a_sneaker", "designs"),
        os.path.join(base, "front_a_sneaker", "approved"),
        os.path.join(base, "front_a_sneaker", "rejected"),
        os.path.join(base, "front_a_sneaker", "drops"),
        os.path.join(base, "front_b_general", "designs"),
        os.path.join(base, "front_b_general", "approved"),
        os.path.join(base, "front_b_general", "rejected"),
        os.path.join(base, "spreadsheets"),
        os.path.join(base, "logs"),
    ]
    for f in folders:
        os.makedirs(f, exist_ok=True)

    print(f"\n🚀  Generating dual-front workspace at: {base}\n")
    print("  Creating spreadsheets...")
    sp = os.path.join(base, "spreadsheets")

    create_designs_front_a(os.path.join(sp, "designs_front_a.xlsx"))
    create_designs_front_b(os.path.join(sp, "designs_front_b.xlsx"))
    create_drop_tracker(os.path.join(sp, "drops_front_a.xlsx"))
    create_app_analytics(os.path.join(sp, "app_analytics.xlsx"))
    create_sales_tracker(os.path.join(sp, "sales.xlsx"))
    create_listings_tracker(os.path.join(sp, "listings.xlsx"))
    create_trademark_log(os.path.join(sp, "trademark_log.xlsx"))
    create_niche_tracker(os.path.join(sp, "niches_front_b.xlsx"))
    create_financials(os.path.join(sp, "financials.xlsx"))

    print("\n  Creating scripts & config...")
    with open(os.path.join(base, "update_workbooks.py"), "w", encoding="utf-8") as f:
        f.write(UPDATE_SCRIPT)
    print("  ✅ update_workbooks.py")

    with open(os.path.join(base, "pipeline_output_example_front_a.json"), "w", encoding="utf-8") as f:
        f.write(EXAMPLE_JSON_A)
    print("  ✅ pipeline_output_example_front_a.json")

    with open(os.path.join(base, "pipeline_output_example_front_b.json"), "w", encoding="utf-8") as f:
        f.write(EXAMPLE_JSON_B)
    print("  ✅ pipeline_output_example_front_b.json")

    with open(os.path.join(base, "WORKSPACE_README.md"), "w", encoding="utf-8") as f:
        f.write(README)
    print("  ✅ WORKSPACE_README.md")

    print(f"""
✅  Dual-front workspace ready!

    {base}/
    ├── front_a_sneaker/         ← Drop your sneaker culture PNGs here
    ├── front_b_general/         ← Drop your general niche PNGs here
    ├── spreadsheets/
    │   ├── designs_front_a.xlsx   👟 Front A design registry
    │   ├── designs_front_b.xlsx   🎨 Front B design registry
    │   ├── drops_front_a.xlsx     🗓️  Drop calendar + performance
    │   ├── app_analytics.xlsx     📱 App CTR + merch analytics
    │   ├── sales.xlsx             💰 Combined orders + P&L by front
    │   ├── listings.xlsx          📋 All listings + action flags
    │   ├── trademark_log.xlsx     ⚖️  Shared IP clearance log
    │   ├── niches_front_b.xlsx    🔍 Niche scoring + phrase bank
    │   └── financials.xlsx        🧾 Expenses + P&L + tax prep
    ├── logs/
    ├── update_workbooks.py
    └── pipeline_output_example_front_[a|b].json

Next:
  Front A: python update_workbooks.py --log logs/front_a_batch.json --front A
  Front B: python update_workbooks.py --log logs/front_b_batch.json --front B
""")


if __name__ == "__main__":
    main()
