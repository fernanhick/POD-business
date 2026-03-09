"""
export_pins.py - Export generated pins into folders for manual Pinterest posting.

Creates a folder per pin with:
  - The pin image (renamed for easy identification)
  - pin_info.txt with copy-paste-ready Title, Description, Board, Link, Alt Text

Usage:
    python export_pins.py                    # Export all draft pins
    python export_pins.py --design rotation_society_001.png  # Export pins for one design
    python export_pins.py --clear            # Clear export folder first
"""

import argparse
import os
import shutil
import sqlite3
import sys
import textwrap
from pathlib import Path

WORKSPACE = Path(__file__).resolve().parent.parent
DB_PATH = WORKSPACE / "pinterest" / "pinterest.db"
EXPORT_DIR = WORKSPACE / "pinterest" / "export"


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# Board name mapping (env key -> human-readable board name)
BOARD_NAMES = {
    "PINTEREST_BOARD_SNEAKER_CULTURE": "Sneaker Collection Goals",
    "PINTEREST_BOARD_OUTFIT_IDEAS": "Kicks & Fits",
    "PINTEREST_BOARD_ROOM_DECOR": "Sneaker Cave Decor",
    "PINTEREST_BOARD_GIFTS": "Gifts for Sneakerheads",
    "PINTEREST_BOARD_STREETWEAR": "Streetwear Culture",
}

# Template -> board env key mapping
TEMPLATE_BOARDS = {
    "lifestyle_sneaker_shelf": "PINTEREST_BOARD_SNEAKER_CULTURE",
    "lifestyle_rotation": "PINTEREST_BOARD_SNEAKER_CULTURE",
    "lifestyle_streetwear": "PINTEREST_BOARD_OUTFIT_IDEAS",
    "lifestyle_room": "PINTEREST_BOARD_ROOM_DECOR",
    "quote_collector": "PINTEREST_BOARD_SNEAKER_CULTURE",
    "quote_hustle": "PINTEREST_BOARD_STREETWEAR",
    "quote_style": "PINTEREST_BOARD_OUTFIT_IDEAS",
    "quote_deadstock": "PINTEREST_BOARD_SNEAKER_CULTURE",
    "list_top5_sneaker": "PINTEREST_BOARD_SNEAKER_CULTURE",
    "list_gift_guide": "PINTEREST_BOARD_GIFTS",
    "list_style_essentials": "PINTEREST_BOARD_OUTFIT_IDEAS",
    "list_room_upgrade": "PINTEREST_BOARD_ROOM_DECOR",
    "mood_dark_aesthetic": "PINTEREST_BOARD_STREETWEAR",
    "mood_street_vibes": "PINTEREST_BOARD_STREETWEAR",
    "mood_cozy_room": "PINTEREST_BOARD_ROOM_DECOR",
    "mood_minimal": "PINTEREST_BOARD_OUTFIT_IDEAS",
    "product_tee": "PINTEREST_BOARD_STREETWEAR",
    "product_hoodie": "PINTEREST_BOARD_STREETWEAR",
    "product_poster": "PINTEREST_BOARD_ROOM_DECOR",
    "product_gift": "PINTEREST_BOARD_GIFTS",
}

ETSY_SHOP_URL = "https://www.etsy.com/shop/RotationClub"


def get_board_name(template_id: str) -> str:
    env_key = TEMPLATE_BOARDS.get(template_id, "")
    return BOARD_NAMES.get(env_key, "Sneaker Collection Goals")


def get_etsy_link(design_filename: str) -> str:
    """Return Etsy listing URL if available, otherwise shop URL."""
    # Try to find listing ID from spreadsheet
    try:
        from openpyxl import load_workbook
        sp = WORKSPACE / "spreadsheets" / "designs_front_a.xlsx"
        if sp.exists():
            wb = load_workbook(sp, read_only=True)
            ws = wb["Designs"]
            headers = {
                (ws.cell(row=2, column=c).value or "").strip(): c
                for c in range(1, ws.max_column + 1)
                if ws.cell(row=2, column=c).value
            }
            fn_col = headers.get("Filename", 2)
            etsy_col = headers.get("Etsy Listing ID")
            if etsy_col:
                for row in range(3, ws.max_row + 1):
                    fn = ws.cell(row=row, column=fn_col).value
                    if fn and fn == design_filename:
                        listing_id = ws.cell(row=row, column=etsy_col).value
                        if listing_id:
                            wb.close()
                            return f"https://www.etsy.com/listing/{listing_id}"
            wb.close()
    except Exception:
        pass
    return ETSY_SHOP_URL


def export_pins(design_filter: str | None = None, latest_only: bool = True):
    conn = get_conn()

    query = "SELECT * FROM pins WHERE status='draft' ORDER BY created_at ASC"
    pins = conn.execute(query).fetchall()
    conn.close()

    if design_filter:
        pins = [p for p in pins if p["design_filename"] == design_filter]

    if not pins:
        print("No draft pins found.")
        return

    # Group by design
    designs = {}
    for pin in pins:
        df = pin["design_filename"]
        if df not in designs:
            designs[df] = []
        designs[df].append(pin)

    # Keep only latest 20 per design (one full template set) to avoid duplicates
    if latest_only:
        for df in designs:
            designs[df] = designs[df][-20:]

    total = 0
    for design_filename, design_pins in designs.items():
        design_stem = Path(design_filename).stem
        design_dir = EXPORT_DIR / design_stem
        design_dir.mkdir(parents=True, exist_ok=True)

        for i, pin in enumerate(design_pins, 1):
            pin = dict(pin)
            template_id = pin["template_id"]
            folder_name = f"pin_{i:02d}_{template_id}"
            pin_dir = design_dir / folder_name
            pin_dir.mkdir(parents=True, exist_ok=True)

            # Copy pin image
            src_image = Path(pin["image_path"])
            if src_image.exists():
                dst_image = pin_dir / f"{design_stem}_{template_id}.png"
                shutil.copy2(src_image, dst_image)

            # Build pin info
            board = get_board_name(template_id)
            link = get_etsy_link(design_filename)
            title = pin["title"]
            description = pin["description"]
            keywords = pin.get("keywords", "")
            alt_text = f"{design_stem.replace('_', ' ').title()} - Sneaker culture graphic tee by RotationClub"

            # Write pin_info.txt
            info_text = textwrap.dedent(f"""\
                ═══════════════════════════════════════════
                 PIN {i:02d} — {template_id}
                ═══════════════════════════════════════════

                BOARD:
                {board}

                TITLE:
                {title}

                DESCRIPTION:
                {description}

                DESTINATION LINK:
                {link}

                ALT TEXT:
                {alt_text}

                ───────────────────────────────────────────
                KEYWORDS (for reference):
                {keywords}

                PIN TYPE: {pin["pin_type"]}
                DESIGN: {design_filename}
                ═══════════════════════════════════════════
            """)

            (pin_dir / "pin_info.txt").write_text(info_text, encoding="utf-8")
            total += 1

        print(f"  {design_stem}: {len(design_pins)} pins exported to {design_dir}")

    print(f"\nTotal: {total} pins exported to {EXPORT_DIR}")
    print(f"\nFor each pin folder:")
    print(f"  1. Upload the .png image to Pinterest")
    print(f"  2. Copy Title from pin_info.txt")
    print(f"  3. Copy Description from pin_info.txt")
    print(f"  4. Set the Board listed in pin_info.txt")
    print(f"  5. Paste the Destination Link")
    print(f"  6. Add Alt Text")
    print(f"  7. Publish or Schedule")


def main():
    parser = argparse.ArgumentParser(description="Export pins for manual Pinterest posting")
    parser.add_argument("--design", help="Filter by design filename")
    parser.add_argument("--clear", action="store_true", help="Clear export folder before exporting")
    args = parser.parse_args()

    if args.clear and EXPORT_DIR.exists():
        shutil.rmtree(EXPORT_DIR)
        print(f"Cleared {EXPORT_DIR}")

    export_pins(design_filter=args.design)


if __name__ == "__main__":
    main()
