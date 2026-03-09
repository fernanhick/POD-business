"""
printify_mockups.py  --  Pull All Mockup Images from Printify Products

Fetches all available mockup angles/styles for a Printify product and
downloads them locally. These images can then be uploaded as additional
Etsy listing photos to improve search visibility (Etsy recommends 5+ images).

The Printify API returns mockup images in the product's "images" array.
Each image has a variant_id association and a src URL. This script pulls
all unique mockup URLs and saves them.

Usage:
    # Fetch mockups for a specific product ID
    python printify_mockups.py --product-id 12345abc

    # Fetch mockups for all products in the spreadsheet that have a Printify Product ID
    python printify_mockups.py --front A

    # Fetch + upload to Etsy listing (requires Etsy API to be connected)
    python printify_mockups.py --front A --upload-to-etsy

Setup:
    Requires PRINTIFY_TOKEN and PRINTIFY_SHOP_ID in workspace/.env
    For --upload-to-etsy: requires ETSY_ACCESS_TOKEN and ETSY_API_KEY

Flow:
    1. GET /v1/shops/{shop_id}/products/{product_id}.json
       -> product["images"] contains all mockup URLs with variant/placement info
    2. Download each unique image URL
    3. Save to workspace/mockup_output/{design_name}/
    4. (Optional) Upload to Etsy listing via PUT /v3/application/listings/{id}/images

TODO: Implement when ready to integrate
    - [ ] Fetch product mockup images from Printify API
    - [ ] Deduplicate by image URL (Printify returns same URL for multiple variants)
    - [ ] Download and save locally
    - [ ] Optional: upload additional images to Etsy listing via Etsy API
    - [ ] Optional: integrate into printify_upload.py publish flow
    - [ ] Optional: let user select which mockup angles to include from webapp UI

API References:
    Printify: GET /v1/shops/{shop_id}/products/{product_id}.json
              -> response.images[].src (mockup image URLs)
              -> response.images[].variant_ids (which variants this mockup shows)
              -> response.images[].is_default (primary listing image)

    Etsy:     POST /v3/application/shops/{shop_id}/listings/{listing_id}/images
              -> multipart form upload with "image" file field
"""

import argparse
import os
import sys
import time
import requests
from pathlib import Path
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

TOKEN = os.environ.get("PRINTIFY_TOKEN", "")
SHOP_ID = os.environ.get("PRINTIFY_SHOP_ID", "")
HEADERS = {"Authorization": f"Bearer {TOKEN}"}
BASE = "https://api.printify.com/v1"

WORKSPACE = Path(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = WORKSPACE / "mockup_output"


def fetch_product_mockups(product_id: str) -> list[dict]:
    """Fetch all mockup images for a Printify product.

    Returns list of {url, variant_ids, is_default, position} dicts.
    """
    resp = requests.get(
        f"{BASE}/shops/{SHOP_ID}/products/{product_id}.json",
        headers=HEADERS,
    )
    resp.raise_for_status()
    product = resp.json()

    images = product.get("images", [])
    seen_urls = set()
    mockups = []

    for img in images:
        url = img.get("src", "")
        if not url or url in seen_urls:
            continue
        seen_urls.add(url)
        mockups.append({
            "url": url,
            "variant_ids": img.get("variant_ids", []),
            "is_default": img.get("is_default", False),
            "position": img.get("position", "front"),
        })

    return mockups


def download_mockups(
    product_id: str,
    design_name: str | None = None,
    out_dir: Path | None = None,
) -> list[Path]:
    """Download all mockup images for a product.

    Returns list of saved file paths.
    """
    mockups = fetch_product_mockups(product_id)
    if not mockups:
        print(f"  No mockup images found for product {product_id}")
        return []

    folder_name = design_name or product_id
    out = out_dir or OUTPUT_DIR / folder_name
    out.mkdir(parents=True, exist_ok=True)

    saved: list[Path] = []
    for i, mockup in enumerate(mockups):
        url = mockup["url"]
        ext = ".png" if ".png" in url.lower() else ".jpg"
        suffix = "default" if mockup["is_default"] else f"angle_{i}"
        filename = f"{folder_name}_{suffix}{ext}"
        filepath = out / filename

        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            filepath.write_bytes(resp.content)
            saved.append(filepath)
            print(f"  [{i+1}/{len(mockups)}] Saved: {filename}")
        except Exception as e:
            print(f"  [{i+1}/{len(mockups)}] Failed: {e}")

        time.sleep(0.3)

    return saved


def upload_mockups_to_etsy(listing_id: str, image_paths: list[Path]) -> int:
    """Upload mockup images to an Etsy listing.

    Requires ETSY_ACCESS_TOKEN and ETSY_API_KEY in environment.
    Returns count of successfully uploaded images.
    """
    etsy_token = os.environ.get("ETSY_ACCESS_TOKEN", "")
    etsy_key = os.environ.get("ETSY_API_KEY", "")
    etsy_shop = os.environ.get("ETSY_NUMERIC_SHOP_ID", "")

    if not all([etsy_token, etsy_key, etsy_shop]):
        print("  Etsy API not configured — skipping upload")
        return 0

    uploaded = 0
    for path in image_paths:
        try:
            with open(path, "rb") as f:
                resp = requests.post(
                    f"https://openapi.etsy.com/v3/application/shops/{etsy_shop}/listings/{listing_id}/images",
                    headers={
                        "Authorization": f"Bearer {etsy_token}",
                        "x-api-key": etsy_key,
                    },
                    files={"image": (path.name, f, "image/png")},
                )
                resp.raise_for_status()
                uploaded += 1
                print(f"  Uploaded to Etsy: {path.name}")
        except Exception as e:
            print(f"  Etsy upload failed for {path.name}: {e}")
        time.sleep(0.5)

    return uploaded


def process_front(front: str, upload_to_etsy: bool = False):
    """Process all products for a front that have Printify Product IDs."""
    sp_map = {
        "A": WORKSPACE / "spreadsheets" / "designs_front_a.xlsx",
        "B": WORKSPACE / "spreadsheets" / "designs_front_b.xlsx",
    }
    spreadsheet = sp_map.get(front)
    if not spreadsheet or not spreadsheet.exists():
        print(f"  Spreadsheet not found for front {front}")
        return

    wb = load_workbook(spreadsheet)
    ws = wb["Designs"]
    headers = {
        (ws.cell(row=2, column=c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=2, column=c).value
    }

    filename_col = headers.get("Filename", 2)
    product_col = headers.get("Printify Product ID")
    etsy_col = headers.get("Etsy Listing ID")

    if not product_col:
        print("  'Printify Product ID' column not found in spreadsheet")
        wb.close()
        return

    for row in range(3, ws.max_row + 1):
        filename = ws.cell(row=row, column=filename_col).value
        product_id = ws.cell(row=row, column=product_col).value
        etsy_id = ws.cell(row=row, column=etsy_col).value if etsy_col else None

        if not filename or not product_id:
            continue

        design_name = Path(str(filename)).stem
        print(f"\n  === {design_name} (Printify: {product_id}) ===")

        saved = download_mockups(str(product_id), design_name)

        if upload_to_etsy and etsy_id and saved:
            print(f"  Uploading {len(saved)} images to Etsy listing {etsy_id}...")
            upload_mockups_to_etsy(str(etsy_id), saved)

    wb.close()


def main():
    parser = argparse.ArgumentParser(
        description="Pull mockup images from Printify and optionally upload to Etsy"
    )
    parser.add_argument("--product-id", help="Specific Printify product ID")
    parser.add_argument("--front", choices=["A", "B"], help="Process all products for a front")
    parser.add_argument("--upload-to-etsy", action="store_true",
                        help="Upload downloaded mockups to Etsy listings")
    parser.add_argument("--out", help="Output directory override")

    args = parser.parse_args()

    if not TOKEN or not SHOP_ID:
        print("  Error: PRINTIFY_TOKEN and PRINTIFY_SHOP_ID must be set in .env")
        sys.exit(1)

    if args.product_id:
        out = Path(args.out) if args.out else None
        saved = download_mockups(args.product_id, out_dir=out)
        print(f"\n  Downloaded {len(saved)} mockup images")

    elif args.front:
        process_front(args.front, upload_to_etsy=args.upload_to_etsy)
        print("\n  Done!")

    else:
        parser.error("Provide --product-id or --front")


if __name__ == "__main__":
    main()
