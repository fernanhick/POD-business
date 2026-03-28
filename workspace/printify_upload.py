"""
printify_upload.py  --  Bulk Upload to Printify & Publish to Etsy

Uploads approved PNGs to Printify, creates products, and optionally
publishes them to your connected Etsy store. Writes back Printify
product IDs and image IDs to the design tracker spreadsheets.

Usage:
    python printify_upload.py --front A
    python printify_upload.py --front B
    python printify_upload.py --front A --draft          # upload but don't publish (for drops)
    python printify_upload.py --front A --dry-run        # preview what would be uploaded
    python printify_upload.py sync-ids --front A         # poll Printify for Etsy listing IDs

Setup:
    1. Get your Printify API token: Settings > Connections > Personal Access Token
    2. Get your shop ID from the Printify dashboard URL: /shop/{SHOP_ID}/
    3. Set environment variables or edit the config below:
       export PRINTIFY_TOKEN="your_token_here"
       export PRINTIFY_SHOP_ID="your_shop_id_here"
"""

import base64
import glob
import os
import sys
import time
import json
import re
import argparse
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv

# ── Load .env from workspace root ─────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# ── Configuration ─────────────────────────────────────────────────
TOKEN   = os.environ.get("PRINTIFY_TOKEN", "")
SHOP_ID = os.environ.get("PRINTIFY_SHOP_ID", "")
HEADERS = {"Authorization": f"Bearer {TOKEN}"}
BASE    = "https://api.printify.com/v1"

WORKSPACE = os.path.dirname(os.path.abspath(__file__))
SP = os.path.join(WORKSPACE, "spreadsheets")

LISTING_LIMITS = {
    "title_max": 140,
    "description_max": 2800,
    "tag_max_count": 13,
    "tag_max_length": 20,
}

# ── Per-front product configuration ──────────────────────────────
# Update blueprint_id, provider_id, variant_ids after querying the
# Printify catalog API (see docs at bottom of this file).
# ── Shared product configs (same garment for both fronts) ────────
_TSHIRT_PRODUCT = {
    "blueprint_id": 12,       # Unisex Jersey Short Sleeve Tee (Bella+Canvas 3001)
    "provider_id":  39,       # SwiftPOD
    "variants": [
        # Black: S–4XL
        {"id": 18100, "size": "S"},
        {"id": 18101, "size": "M"},
        {"id": 18102, "size": "L"},
        {"id": 18103, "size": "XL"},
        {"id": 18104, "size": "2XL"},
        {"id": 18105, "size": "3XL"},
        {"id": 18106, "size": "4XL"},
        # White: S–4XL
        {"id": 18540, "size": "S"},
        {"id": 18541, "size": "M"},
        {"id": 18542, "size": "L"},
        {"id": 18543, "size": "XL"},
        {"id": 18544, "size": "2XL"},
        {"id": 18545, "size": "3XL"},
        {"id": 18546, "size": "4XL"},
    ],
    "base_cost_cents": 1129,
    "oversize_costs": {
        "2XL": 1382, "3XL": 1612, "4XL": 1863,
    },
    "shipping_cents": 429,
}

_HOODIE_PRODUCT = {
    "blueprint_id": 77,       # Unisex Heavy Blend Hooded Sweatshirt (Gildan 18500)
    "provider_id":  39,       # SwiftPOD
    "variants": [
        # Black: S–5XL
        {"id": 32918, "size": "S"},
        {"id": 32919, "size": "M"},
        {"id": 32920, "size": "L"},
        {"id": 32921, "size": "XL"},
        {"id": 32922, "size": "2XL"},
        {"id": 32923, "size": "3XL"},
        {"id": 32924, "size": "4XL"},
        {"id": 32925, "size": "5XL"},
        # White: S–5XL
        {"id": 32910, "size": "S"},
        {"id": 32911, "size": "M"},
        {"id": 32912, "size": "L"},
        {"id": 32913, "size": "XL"},
        {"id": 32914, "size": "2XL"},
        {"id": 32915, "size": "3XL"},
        {"id": 32916, "size": "4XL"},
        {"id": 32917, "size": "5XL"},
    ],
    "base_cost_cents": 2224,
    "oversize_costs": {
        "2XL": 2437, "3XL": 2558, "4XL": 2618, "5XL": 2618,
    },
    "shipping_cents": 769,
}

FRONT_CONFIG = {
    "A": {
        "designs_dir":  os.path.join(WORKSPACE, "front_a_sneaker", "approved"),
        "spreadsheet":  os.path.join(SP, "designs_front_a.xlsx"),
        "sheet_name":   "Designs",
        "products": {
            "tshirt": {
                **_TSHIRT_PRODUCT,
                "tags": [
                    "sneakerhead shirt",     # 17
                    "sneaker collector",      # 17
                    "streetwear tee",         # 14
                    "sneaker culture tee",    # 19
                    "rotationclub tee",       # 16
                    "sneaker lover gift",     # 18
                    "kicks and culture",      # 17
                    "sneakerhead gift",       # 16
                    "sneaker dad gift",       # 16
                    "hypebeast tee",          # 13
                    "sneaker gift idea",      # 17
                    "streetwear for men",     # 18
                    "graphic tee for him",    # 19
                ],
                "title_template": (
                    "{name} Sneakerhead Shirt | Sneaker Culture Tee "
                    "| Streetwear Graphic Tee | Gift for Sneaker Collector"
                ),
                "description_template": "",
            },
            "hoodie": {
                **_HOODIE_PRODUCT,
                "tags": [
                    "sneakerhead hoodie",     # 18
                    "sneaker collector",       # 17
                    "streetwear hoodie",       # 17
                    "sneaker culture",         # 15
                    "rotationclub hoodie",     # 19
                    "sneaker lover gift",      # 18
                    "kicks and culture",       # 17
                    "sneakerhead gift",        # 16
                    "sneaker dad gift",        # 16
                    "hypebeast hoodie",        # 16
                    "sneaker gift idea",       # 17
                    "streetwear for men",      # 18
                    "graphic hoodie",          # 14
                ],
                "title_template": (
                    "{name} Sneakerhead Hoodie | Sneaker Culture Hoodie "
                    "| Streetwear Graphic Hoodie | Gift for Sneaker Collector"
                ),
                "description_template": "",
            },
        },
    },
    "B": {
        "designs_dir":  os.path.join(WORKSPACE, "front_b_general", "approved"),
        "spreadsheet":  os.path.join(SP, "designs_front_b.xlsx"),
        "sheet_name":   "Designs",
        "products": {
            "tshirt": {
                **_TSHIRT_PRODUCT,
                "tags": [
                    "workout shirt",           # 13
                    "gym shirt",               # 9
                    "fitness gift",            # 12
                    "funny gym shirt",         # 15
                    "weightlifting tee",       # 17
                    "WOD shirt",               # 9
                    "lifting shirt",           # 13
                    "functional fitness",      # 19
                    "gym humor",               # 9
                    "barbell shirt",           # 13
                    "exercise shirt",          # 14
                    "gym lover gift",          # 14
                    "fitness tshirt",          # 14
                ],
                "title_template": (
                    "{name} Workout Shirt | Funny Gym Tee "
                    "| WOD Fitness Gift | Unisex Graphic Tee"
                ),
                "description_template": (
                    "A graphic tee for anyone who lives for the WOD. "
                    "Whether you're chasing PRs or surviving burpees, "
                    "this shirt says it all.\n\n"
                    "DETAILS:\n"
                    "* Soft, lightweight ring-spun cotton (Bella+Canvas 3001)\n"
                    "* DTG printed -- vibrant, wash-resistant graphics\n"
                    "* Unisex fit -- true to size\n"
                    "* Available in Black and White, S through 5XL\n\n"
                    "Great gift for gym lovers, lifters, and anyone "
                    "who knows what AMRAP means."
                ),
            },
            "hoodie": {
                **_HOODIE_PRODUCT,
                "tags": [
                    "workout hoodie",          # 14
                    "gym hoodie",              # 10
                    "fitness gift",            # 12
                    "funny gym hoodie",        # 16
                    "weightlifting hood",      # 18
                    "WOD hoodie",              # 10
                    "lifting hoodie",          # 14
                    "functional fitness",      # 19
                    "gym humor",               # 9
                    "barbell hoodie",          # 14
                    "exercise hoodie",         # 15
                    "gym lover gift",          # 14
                    "fitness hoodie",          # 14
                ],
                "title_template": (
                    "{name} Workout Hoodie | Funny Gym Hoodie "
                    "| WOD Fitness Gift | Unisex Graphic Hoodie"
                ),
                "description_template": (
                    "A graphic hoodie for anyone who lives for the WOD. "
                    "Whether you're chasing PRs or surviving burpees, "
                    "this hoodie says it all.\n\n"
                    "DETAILS:\n"
                    "* Premium heavyweight fleece blend (Gildan 18500)\n"
                    "* DTG printed -- vibrant, wash-resistant graphics\n"
                    "* Oversized fit -- size up for drop-shoulder "
                    "silhouette\n"
                    "* Available in Black and White, S through 5XL\n\n"
                    "Great gift for gym lovers, lifters, and anyone "
                    "who knows what AMRAP means."
                ),
            },
        },
    },
    "C": {
        "designs_dir":  os.path.join(WORKSPACE, "front_custom", "approved"),
        "spreadsheet":  None,
        "sheet_name":   None,
        "products": {
            "tshirt": {
                **_TSHIRT_PRODUCT,
                "tags": [
                    "graphic tee",
                    "custom design tee",
                    "unisex graphic shirt",
                    "streetwear tee",
                    "trendy graphic tee",
                    "cool shirt for him",
                    "cool shirt for her",
                    "unique graphic tee",
                    "art tee shirt",
                    "pop culture tee",
                    "graphic tee gift",
                    "funny graphic tee",
                    "statement tee",
                ],
                "title_template": (
                    "{name} Graphic Tee | Unisex Graphic T-Shirt "
                    "| Cool Graphic Tee | Unique Design Shirt"
                ),
                "description_template": (
                    "A bold graphic tee featuring a unique custom design. "
                    "Stand out from the crowd with this statement piece.\n\n"
                    "DETAILS:\n"
                    "* Premium soft cotton (Bella+Canvas 3001)\n"
                    "* DTG printed -- vibrant, wash-resistant graphics\n"
                    "* True-to-size unisex fit\n"
                    "* Available in Black and White, S through 4XL"
                ),
            },
            "hoodie": {
                **_HOODIE_PRODUCT,
                "tags": [
                    "graphic hoodie",
                    "custom design hoodie",
                    "unisex hoodie",
                    "streetwear hoodie",
                    "trendy hoodie",
                    "cool hoodie for him",
                    "cool hoodie for her",
                    "unique graphic hoodie",
                    "art hoodie",
                    "pop culture hoodie",
                    "graphic hoodie gift",
                    "funny graphic hoodie",
                    "statement hoodie",
                ],
                "title_template": (
                    "{name} Graphic Hoodie | Unisex Graphic Hoodie "
                    "| Cool Graphic Hoodie | Unique Design Hoodie"
                ),
                "description_template": (
                    "A bold graphic hoodie featuring a unique custom design. "
                    "Stand out from the crowd with this statement piece.\n\n"
                    "DETAILS:\n"
                    "* Premium heavyweight fleece blend (Gildan 18500)\n"
                    "* DTG printed -- vibrant, wash-resistant graphics\n"
                    "* Oversized fit -- size up for drop-shoulder silhouette\n"
                    "* Available in Black and White, S through 5XL"
                ),
            },
        },
    },
}


def _normalize_spaces(value):
    return re.sub(r"\s+", " ", value or "").strip()


def _load_design_metadata_index():
    metadata = {}
    for path in sorted(glob.glob(os.path.join(WORKSPACE, "logs", "front_*.json"))):
        try:
            payload = json.loads(open(path, "r", encoding="utf-8").read())
        except Exception:
            continue
        rows = payload if isinstance(payload, list) else [payload]
        for row in rows:
            if not isinstance(row, dict):
                continue
            filename = row.get("filename")
            if not filename:
                continue
            metadata[filename] = {
                "style": row.get("style"),
                "color_palette": row.get("color_palette"),
                "generation_model": row.get("generation_model"),
            }
    return metadata


def _variant_label_from_palette(palette):
    if not palette:
        return None
    descriptor = palette.split(",", 1)[1].strip() if "," in palette else palette.strip()
    descriptor = re.sub(r"\b(aesthetic|finish|warmth|look)\b", "", descriptor, flags=re.IGNORECASE)
    descriptor = _normalize_spaces(descriptor)
    if not descriptor:
        return None
    words = descriptor.split()
    return " ".join(words[:3]).title()


def _variant_label_from_style(style):
    if not style:
        return None
    normalized = style.replace("+", " ")
    normalized = re.sub(r"\b(text|graphic|letters|lettering|typography|layout)\b", "", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"[^a-zA-Z0-9 ]+", " ", normalized)
    normalized = _normalize_spaces(normalized)
    if not normalized:
        return None
    words = normalized.split()
    return " ".join(words[:3]).title()


def _number_to_edition_word(value):
    words = {
        0: "Zero",
        1: "One",
        2: "Two",
        3: "Three",
        4: "Four",
        5: "Five",
        6: "Six",
        7: "Seven",
        8: "Eight",
        9: "Nine",
        10: "Ten",
        11: "Eleven",
        12: "Twelve",
        13: "Thirteen",
        14: "Fourteen",
        15: "Fifteen",
        16: "Sixteen",
        17: "Seventeen",
        18: "Eighteen",
        19: "Nineteen",
        20: "Twenty",
    }
    return words.get(value, str(value))


def _has_duplicate_base_name(base_stem, directory):
    if not directory or not os.path.isdir(directory):
        return False
    pattern = re.compile(rf"^{re.escape(base_stem)}_\d+\.png$", re.IGNORECASE)
    matches = [name for name in os.listdir(directory) if pattern.match(name)]
    return len(matches) > 1


def display_name_from_filename(filename, directory=None):
    file_name = os.path.basename(filename)
    stem = os.path.splitext(file_name)[0]
    base_stem = re.sub(r"_\d+$", "", stem)
    base_name = base_stem.replace("_", " ").title()

    if not _has_duplicate_base_name(base_stem, directory):
        return base_name

    metadata = _load_design_metadata_index().get(file_name, {})
    variant_label = _variant_label_from_palette(metadata.get("color_palette"))
    if not variant_label:
        variant_label = _variant_label_from_style(metadata.get("style"))
    if not variant_label:
        suffix_match = re.search(r"_(\d+)$", stem)
        if suffix_match:
            variant_label = f"Edition {_number_to_edition_word(int(suffix_match.group(1)))}"

    if variant_label:
        return f"{base_name} {variant_label}"
    return base_name


def _clip_text(value, max_length):
    value = (value or "").strip()
    if len(value) <= max_length:
        return value

    clipped = value[: max_length + 1].rsplit(" ", 1)[0].rstrip(" ,;:-|/")
    if clipped and len(clipped) >= int(max_length * 0.6):
        return clipped
    return value[:max_length].rstrip(" ,;:-|/")


def _normalize_tag(tag):
    normalized = _normalize_spaces(tag).lower()
    normalized = normalized.replace("_", " ")
    return normalized


def normalize_listing_content(title, description, tags):
    warnings = []

    normalized_title = _clip_text(_normalize_spaces(title), LISTING_LIMITS["title_max"])
    if len(normalized_title) < len((title or "").strip()):
        warnings.append(
            f"title trimmed to {LISTING_LIMITS['title_max']} chars"
        )

    description = (description or "").strip()
    normalized_description = _clip_text(description, LISTING_LIMITS["description_max"])
    if len(normalized_description) < len(description):
        warnings.append(
            f"description trimmed to {LISTING_LIMITS['description_max']} chars"
        )

    normalized_tags = []
    seen = set()
    for tag in tags or []:
        candidate = _clip_text(_normalize_tag(tag), LISTING_LIMITS["tag_max_length"])
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        normalized_tags.append(candidate)

    if len(normalized_tags) > LISTING_LIMITS["tag_max_count"]:
        warnings.append(
            f"tags trimmed to {LISTING_LIMITS['tag_max_count']} entries"
        )
    normalized_tags = normalized_tags[: LISTING_LIMITS["tag_max_count"]]

    return normalized_title, normalized_description, normalized_tags, warnings


def build_front_a_description(design_name, product_type):
    if product_type == "hoodie":
        intro = (
            f"{design_name} is a sneakerhead hoodie and sneaker culture graphic hoodie made for sneaker collectors, sneakerheads, and streetwear fans."
        )
        body = (
            "This streetwear graphic hoodie works as an everyday hoodie for sneaker lovers, hypebeasts, and collectors who want sneaker culture clothing that fits the rotation. "
            "Designed for layered cold-weather fits, casual streetwear outfits, and the deadstock shelf alike."
        )
        details = (
            "DETAILS:\n"
            "* Premium heavyweight fleece blend (Gildan 18500)\n"
            "* DTG printed for a clean, durable graphic finish\n"
            "* Relaxed streetwear fit; size up for a roomier drop-shoulder look\n"
            "* Available in Black and White, sizes S through 5XL"
        )
        close = (
            "Great gift for a sneaker collector, sneaker lover, hypebeast, or anyone shopping for a graphic hoodie, sneakerhead hoodie, or streetwear hoodie."
        )
    else:
        intro = (
            f"{design_name} is a sneakerhead shirt and sneaker culture graphic tee made for sneaker collectors, sneakerheads, and streetwear fans."
        )
        body = (
            "This streetwear graphic tee works as an everyday shirt for sneaker lovers and collectors who want sneaker culture clothing that fits casual outfits, weekend rotation days, and daily wear. "
            "Designed for the daily rotation and the deadstock shelf alike."
        )
        details = (
            "DETAILS:\n"
            "* Soft, lightweight ring-spun cotton (Bella+Canvas 3001)\n"
            "* DTG printed for a clean, durable graphic finish\n"
            "* Unisex fit; true to size with an easy everyday feel\n"
            "* Available in Black and White, sizes S through 4XL"
        )
        close = (
            "Great gift for a sneaker collector, sneaker lover, or anyone shopping for a graphic tee, sneakerhead shirt, or streetwear graphic tee."
        )

    return f"{intro}\n\n{body}\n\n{details}\n\n{close}"


def build_listing_copy(front, product_type, design_name, cfg):
    raw_title = cfg["title_template"].format(name=design_name)
    if front == "A":
        raw_description = build_front_a_description(design_name, product_type)
    else:
        raw_description = cfg["description_template"]

    raw_tags = build_product_tags(cfg, design_name=design_name, product_type=product_type)
    return normalize_listing_content(raw_title, raw_description, raw_tags)


def check_config():
    if not TOKEN:
        print("  Error: PRINTIFY_TOKEN not set.")
        print("  Set it via: export PRINTIFY_TOKEN='your_token_here'")
        print("  Or edit the TOKEN variable at the top of this script.")
        sys.exit(1)
    if not SHOP_ID:
        print("  Error: PRINTIFY_SHOP_ID not set.")
        print("  Set it via: export PRINTIFY_SHOP_ID='your_shop_id_here'")
        sys.exit(1)


def upload_image(filepath):
    """Upload a design PNG to Printify via base64 JSON. Returns the image ID."""
    with open(filepath, "rb") as f:
        contents = base64.b64encode(f.read()).decode("utf-8")
    resp = requests.post(
        f"{BASE}/uploads/images.json",
        headers={**HEADERS, "Content-Type": "application/json"},
        json={"file_name": os.path.basename(filepath), "contents": contents},
    )
    resp.raise_for_status()
    return resp.json()["id"]


def calc_price(cost_cents, shipping_cents):
    """40% profit margin including shipping: price = (cost + ship) / 0.60, rounded to .99"""
    raw = (cost_cents + shipping_cents) / 0.60
    dollars = int(raw / 100) + 1  # round up to next dollar
    return dollars * 100 - 1      # e.g. $23.99 = 2399


def build_product_tags(cfg, design_name=None, product_type=None):
    """Build tags: dynamic from design name first, then base tags, capped at 13 (Etsy max)."""
    base_tags = list(cfg.get("tags", []))
    dynamic = []
    if design_name:
        name_lower = design_name.lower().strip()
        dynamic.append(name_lower)
        first_word = name_lower.split()[0] if name_lower.split() else None
        if first_word:
            suffix = "hoodie" if product_type == "hoodie" else "tee"
            dynamic.append(f"{first_word} {suffix}")
    # Deduplicate: dynamic tags first, then fill with base tags
    seen = set()
    merged = []
    for tag in dynamic + base_tags:
        if tag not in seen:
            seen.add(tag)
            merged.append(tag)
    return [t[:20] for t in merged[:13]]  # Etsy max 20 chars per tag


def ensure_product_tags(product_id, title, description, tags, attempts=3, delay=0.8):
    """Reapply tags if Printify returns an empty tag list after product creation."""
    normalized = [tag for tag in (tags or []) if tag]
    if not normalized:
        return

    for attempt in range(1, attempts + 1):
        product = get_product(str(product_id))
        current_tags = [tag for tag in product.get("tags", []) if tag]
        if current_tags:
            return

        update_product(str(product_id), title, description, normalized)

        product = get_product(str(product_id))
        current_tags = [tag for tag in product.get("tags", []) if tag]
        if current_tags:
            return

        if attempt < attempts:
            time.sleep(delay)

    raise RuntimeError(
        f"Printify product {product_id} still has no tags after {attempts} update attempt(s)"
    )


def create_product(image_id, title, description, cfg, design_name=None):
    """Create a product on Printify with the uploaded image. Returns product ID."""
    product_type = "hoodie" if "hoodie" in title.lower() else "tshirt"
    tags = build_product_tags(cfg, design_name=design_name, product_type=product_type)
    title, description, tags, _ = normalize_listing_content(title, description, tags)

    # Build variants with calculated per-size pricing
    base_cost = cfg["base_cost_cents"]
    shipping = cfg["shipping_cents"]
    oversize = cfg.get("oversize_costs", {})

    variants = []
    variant_ids = []
    for v in cfg["variants"]:
        cost = oversize.get(v["size"], base_cost)
        price = calc_price(cost, shipping)
        variants.append({"id": v["id"], "price": price, "is_enabled": True})
        variant_ids.append(v["id"])

    payload = {
        "title": title,
        "description": description,
        "tags": tags,
        "blueprint_id": cfg["blueprint_id"],
        "print_provider_id": cfg["provider_id"],
        "variants": variants,
        "print_areas": [{
            "variant_ids": variant_ids,
            "placeholders": [{
                "position": "front",
                "images": [{
                    "id": image_id,
                    "x": 0.5, "y": 0.5,
                    "scale": 1.0, "angle": 0,
                }],
            }],
        }],
    }
    resp = requests.post(
        f"{BASE}/shops/{SHOP_ID}/products.json",
        headers={**HEADERS, "Content-Type": "application/json"},
        json=payload,
    )
    resp.raise_for_status()
    product_id = resp.json()["id"]
    ensure_product_tags(product_id, title, description, tags)
    return product_id


def publish_product(product_id):
    """Publish a Printify product to the connected Etsy store."""
    resp = requests.post(
        f"{BASE}/shops/{SHOP_ID}/products/{product_id}/publish.json",
        headers={**HEADERS, "Content-Type": "application/json"},
        json={
            "title": True, "description": True, "images": True,
            "variants": True, "tags": True, "keyFeatures": True,
            "shipping_template": True,
        },
    )
    resp.raise_for_status()


def get_product(product_id):
    """Fetch product details from Printify (used to get Etsy listing ID)."""
    resp = requests.get(
        f"{BASE}/shops/{SHOP_ID}/products/{product_id}.json",
        headers=HEADERS,
    )
    resp.raise_for_status()
    return resp.json()


def update_spreadsheet_ids(spreadsheet_path, sheet_name, filename,
                           image_id=None, product_id=None, etsy_id=None,
                           status=None):
    """
    Find a design row by filename and update its Printify/Etsy IDs.
    Columns searched: B (Filename), columns updated vary by front.
    """
    wb = load_workbook(spreadsheet_path)
    ws = wb[sheet_name]

    # Find the filename column (B=2) and ID columns
    # Front A: O=Printify Image ID, P=Printify Product ID, Q=Etsy Listing ID, R=Status
    # Front B: M=Printify Image ID, N=Printify Product ID, O=Etsy Listing ID, P=Status
    header_row = 2  # row 1 is the title banner
    headers = {ws.cell(row=header_row, column=c).value: c
               for c in range(1, ws.max_column + 1)}

    filename_col = headers.get("Filename", 2)
    img_col = headers.get("Printify Image ID")
    prod_col = headers.get("Printify Product ID")
    etsy_col = headers.get("Etsy Listing ID")
    status_col = headers.get("Status")

    updated = False
    for row in range(3, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=filename_col).value
        if cell_val and cell_val.strip() == filename.strip():
            if image_id and img_col:
                ws.cell(row=row, column=img_col, value=image_id)
            if product_id and prod_col:
                ws.cell(row=row, column=prod_col, value=product_id)
            if etsy_id and etsy_col:
                ws.cell(row=row, column=etsy_col, value=etsy_id)
            if status and status_col:
                ws.cell(row=row, column=status_col, value=status)
            updated = True
            break

    if updated:
        wb.save(spreadsheet_path)
    else:
        print(f"    Warning: '{filename}' not found in spreadsheet")
    wb.close()
    return updated


def run_upload(front, product_type="tshirt", draft=False, dry_run=False):
    """Main upload loop for a front."""
    cfg = FRONT_CONFIG[front]
    pcfg = cfg["products"][product_type]
    designs_dir = cfg["designs_dir"]

    if not os.path.isdir(designs_dir):
        print(f"  Designs folder not found: {designs_dir}")
        print(f"  Place approved PNGs there first.")
        sys.exit(1)

    files = sorted(f for f in os.listdir(designs_dir) if f.lower().endswith(".png"))
    if not files:
        print(f"  No PNG files in {designs_dir}")
        return

    action = "publish to Etsy" if not draft else "create as draft (no publish)"
    print(f"  Front {front} ({product_type}): {len(files)} designs to upload ({action})\n")

    results = []
    for filename in files:
        filepath = os.path.join(designs_dir, filename)
        base_name = display_name_from_filename(filename, designs_dir)
        title, desc, tags, warnings = build_listing_copy(front, product_type, base_name, pcfg)

        if dry_run:
            print(f"  [DRY] {filename} -> '{title[:60]}...'")
            if warnings:
                print(f"        Warnings: {', '.join(warnings)}")
            continue

        try:
            # Step 1: Upload image
            image_id = upload_image(filepath)
            print(f"  [1/3] Uploaded image: {filename} (id={image_id})")

            # Step 2: Create product
            product_id = create_product(image_id, title, desc, pcfg, design_name=base_name)
            print(f"  [2/3] Created product: {product_id}")
            if warnings:
                print(f"        Auto-normalized: {', '.join(warnings)}")

            # Step 3: Publish (unless draft mode for drops)
            if not draft:
                publish_product(product_id)
                print(f"  [3/3] Published to Etsy")
                status = "Published"
            else:
                print(f"  [3/3] Kept as draft (drop mode)")
                status = "Draft on Printify"

            # Step 4: Write IDs back to spreadsheet
            update_spreadsheet_ids(
                cfg["spreadsheet"], cfg["sheet_name"], filename,
                image_id=image_id, product_id=product_id, status=status,
            )

            results.append({
                "filename": filename, "image_id": image_id,
                "product_id": product_id, "status": status,
            })
            print(f"  [OK]  {title[:60]}...\n")

        except requests.HTTPError as e:
            print(f"  [ERR] {filename}: HTTP {e.response.status_code} "
                  f"-- {e.response.text[:200]}\n")
        except Exception as e:
            print(f"  [ERR] {filename}: {e}\n")

        time.sleep(1.5)  # ~40 uploads/min, well under Printify rate limit

    if not dry_run:
        # Save upload log
        log_path = os.path.join(WORKSPACE, "logs",
                                f"upload_{front}_{int(time.time())}.json")
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2)
        print(f"  Upload log: {log_path}")
        print(f"  Done: {len(results)}/{len(files)} uploaded successfully")


def sync_etsy_ids(front):
    """
    Poll Printify for Etsy listing IDs on products that have been published
    but don't yet have an Etsy ID in the spreadsheet.
    """
    cfg = FRONT_CONFIG[front]
    wb = load_workbook(cfg["spreadsheet"])
    ws = wb[cfg["sheet_name"]]

    header_row = 2
    headers = {ws.cell(row=header_row, column=c).value: c
               for c in range(1, ws.max_column + 1)}

    prod_col = headers.get("Printify Product ID")
    etsy_col = headers.get("Etsy Listing ID")

    if not prod_col or not etsy_col:
        print("  Could not find ID columns in spreadsheet")
        wb.close()
        return

    synced = 0
    for row in range(3, ws.max_row + 1):
        product_id = ws.cell(row=row, column=prod_col).value
        etsy_id = ws.cell(row=row, column=etsy_col).value

        if product_id and not etsy_id:
            try:
                product = get_product(str(product_id))
                external = product.get("external", {})
                eid = external.get("id")
                if eid:
                    ws.cell(row=row, column=etsy_col, value=str(eid))
                    fname = ws.cell(row=row, column=2).value or "?"
                    print(f"  [SYNC] {fname}: Etsy ID = {eid}")
                    synced += 1
                time.sleep(0.5)
            except Exception as e:
                print(f"  [ERR] Product {product_id}: {e}")

    if synced:
        wb.save(cfg["spreadsheet"])
    wb.close()
    print(f"  Synced {synced} Etsy listing IDs")


def update_product(product_id, title, description, tags):
    """Update an existing Printify product's title, description, and tags."""
    title, description, tags, _ = normalize_listing_content(title, description, tags)
    payload = {
        "title": title,
        "description": description,
        "tags": tags,
    }
    resp = requests.put(
        f"{BASE}/shops/{SHOP_ID}/products/{product_id}.json",
        headers={**HEADERS, "Content-Type": "application/json"},
        json=payload,
    )
    resp.raise_for_status()
    return resp.json()


def run_update(front, product_type="tshirt", republish=False, dry_run=False):
    """Update all existing products for a front with current tags/title/description."""
    cfg = FRONT_CONFIG[front]
    pcfg = cfg["products"][product_type]
    spreadsheet = cfg["spreadsheet"]
    sheet_name = cfg["sheet_name"]

    wb = load_workbook(spreadsheet)
    ws = wb[sheet_name]

    header_row = 2
    headers = {ws.cell(row=header_row, column=c).value: c
               for c in range(1, ws.max_column + 1)}

    filename_col = headers.get("Filename", 2)
    prod_col = headers.get("Printify Product ID")

    if not prod_col:
        print("  Could not find 'Printify Product ID' column in spreadsheet")
        wb.close()
        return

    # Collect rows with product IDs
    rows = []
    for row in range(3, ws.max_row + 1):
        product_id = ws.cell(row=row, column=prod_col).value
        filename = ws.cell(row=row, column=filename_col).value
        if product_id and filename:
            rows.append((row, str(product_id).strip(), filename.strip()))
    wb.close()

    if not rows:
        print(f"  No products found in spreadsheet for Front {front}")
        return

    action = "update + republish" if republish else "update on Printify"
    print(f"  Front {front} ({product_type}): {len(rows)} products to {action}\n")

    updated = 0
    for row_num, product_id, filename in rows:
        base_name = display_name_from_filename(filename, cfg["designs_dir"])
        title, desc, tags, warnings = build_listing_copy(front, product_type, base_name, pcfg)

        if dry_run:
            print(f"  [DRY] {filename} (id={product_id})")
            print(f"        Title: {title[:70]}...")
            print(f"        Tags:  {', '.join(tags[:5])}... ({len(tags)} total)\n")
            if warnings:
                print(f"        Warnings: {', '.join(warnings)}\n")
            continue

        try:
            update_product(product_id, title, desc, tags)
            print(f"  [UPD] {filename} (id={product_id}) -- tags: {len(tags)}, title+desc updated")
            if warnings:
                print(f"        Auto-normalized: {', '.join(warnings)}")

            if republish:
                publish_product(product_id)
                print(f"        Re-published to Etsy")

            updated += 1
        except requests.HTTPError as e:
            print(f"  [ERR] {filename}: HTTP {e.response.status_code} "
                  f"-- {e.response.text[:200]}")
        except Exception as e:
            print(f"  [ERR] {filename}: {e}")

        time.sleep(1.0)

    print(f"\n  Done: {updated}/{len(rows)} products updated")


def main():
    parser = argparse.ArgumentParser(
        description="Bulk upload designs to Printify and publish to Etsy"
    )
    sub = parser.add_subparsers(dest="command")

    # ── upload command (default) ──
    up = sub.add_parser("upload", help="Upload and publish designs")
    up.add_argument("--front", required=True, choices=["A", "B"],
                    help="A = Sneaker Culture, B = Generalized")
    up.add_argument("--product", default="tshirt", choices=["tshirt", "hoodie"],
                    help="Product type: tshirt or hoodie (default: tshirt)")
    up.add_argument("--draft", action="store_true",
                    help="Create on Printify but don't publish to Etsy (for drops)")
    up.add_argument("--dry-run", action="store_true",
                    help="Preview what would be uploaded without doing it")

    # ── update command ──
    upd = sub.add_parser("update",
                         help="Update existing products with current tags/title/description")
    upd.add_argument("--front", required=True, choices=["A", "B"],
                     help="A = Sneaker Culture, B = Generalized")
    upd.add_argument("--product", default="tshirt", choices=["tshirt", "hoodie"],
                     help="Product type: tshirt or hoodie (default: tshirt)")
    upd.add_argument("--republish", action="store_true",
                     help="Re-publish to Etsy after updating (pushes changes live)")
    upd.add_argument("--dry-run", action="store_true",
                     help="Preview what would be updated without doing it")

    # ── sync-ids command ──
    sync = sub.add_parser("sync-ids",
                          help="Poll Printify for Etsy listing IDs")
    sync.add_argument("--front", required=True, choices=["A", "B"])

    # ── catalog command (helper) ──
    cat = sub.add_parser("catalog",
                         help="Browse Printify catalog to find blueprint/variant IDs")
    cat.add_argument("--blueprints", action="store_true",
                     help="List all available blueprints")
    cat.add_argument("--providers", type=int, metavar="BLUEPRINT_ID",
                     help="List print providers for a blueprint")
    cat.add_argument("--variants", nargs=2, type=int,
                     metavar=("BLUEPRINT_ID", "PROVIDER_ID"),
                     help="List variants for a blueprint + provider")

    args = parser.parse_args()

    if args.command == "upload":
        if not args.dry_run:
            check_config()
        run_upload(args.front, product_type=args.product, draft=args.draft, dry_run=args.dry_run)

    elif args.command == "update":
        if not args.dry_run:
            check_config()
        run_update(args.front, product_type=args.product,
                   republish=args.republish, dry_run=args.dry_run)

    elif args.command == "sync-ids":
        check_config()
        sync_etsy_ids(args.front)

    elif args.command == "catalog":
        check_config()
        if args.blueprints:
            resp = requests.get(f"{BASE}/catalog/blueprints.json",
                                headers=HEADERS)
            resp.raise_for_status()
            for bp in resp.json():
                print(f"  [{bp['id']}] {bp['title']}")
        elif args.providers:
            resp = requests.get(
                f"{BASE}/catalog/blueprints/{args.providers}/print_providers.json",
                headers=HEADERS)
            resp.raise_for_status()
            for pp in resp.json():
                print(f"  [{pp['id']}] {pp['title']}")
        elif args.variants:
            bid, pid = args.variants
            resp = requests.get(
                f"{BASE}/catalog/blueprints/{bid}/print_providers/{pid}/variants.json",
                headers=HEADERS)
            resp.raise_for_status()
            for v in resp.json().get("variants", []):
                print(f"  [{v['id']}] {v['title']} "
                      f"(size={v.get('options', {}).get('size', '?')}, "
                      f"color={v.get('options', {}).get('color', '?')})")
        else:
            cat.print_help()

    else:
        # Default to upload if --front is provided without subcommand
        parser.print_help()


if __name__ == "__main__":
    main()
