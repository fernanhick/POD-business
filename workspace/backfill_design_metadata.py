from __future__ import annotations

import argparse
import json
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

WORKSPACE_DIR = Path(__file__).resolve().parent
LOGS_DIR = WORKSPACE_DIR / "logs"
SPREADSHEETS_DIR = WORKSPACE_DIR / "spreadsheets"
A_FILE = SPREADSHEETS_DIR / "designs_front_a.xlsx"
B_FILE = SPREADSHEETS_DIR / "designs_front_b.xlsx"
APP_DB = WORKSPACE_DIR.parent / "webapp" / "backend" / "app" / "app_state.db"

PALETTE_OPTIONS = [
    "black and cream, vintage wash",
    "off-white and charcoal, clean matte finish",
    "forest green and ecru, military aesthetic",
    "washed black and bone white, faded streetwear",
    "navy blue and gold, luxury streetwear",
    "burgundy and cream, vintage sport aesthetic",
    "rust orange and sand, earth tone warmth",
    "pure black and white, high contrast monochrome",
    "olive drab and tan, utilitarian workwear",
    "slate grey and neon green, tech streetwear",
    "terracotta and ivory, warm bohemian",
    "deep teal and warm cream, coastal vintage",
    "dusty rose and charcoal, soft modern",
    "mustard yellow and dark brown, retro 70s",
    "lavender and slate, muted pastel",
    "red and black, bold graphic",
]

VISUAL_MODE_TO_MODEL = {
    "text_only": "ideogram",
    "graphic_text": "hf",
    "graphic_only": "hf",
    "text_only_openai": "openai",
    "graphic_openai": "openai_graphic",
    "text_gpt_image": "gpt_image",
    "graphic_gpt_image": "gpt_image_graphic",
}


def header_map(ws) -> dict[str, int]:
    return {
        (cell.value or "").strip(): idx
        for idx, cell in enumerate(ws[2], start=1)
        if isinstance(cell.value, str)
    }


def ensure_header_column(ws, name: str) -> int:
    headers = header_map(ws)
    if name in headers:
        return headers[name]
    col = ws.max_column + 1
    ws.cell(row=2, column=col, value=name)
    return col


def merge_meta(target: dict[str, dict[str, str]], filename: str, colorway: str | None, model: str | None) -> None:
    if not filename:
        return
    row = target.setdefault(filename, {})
    if colorway and not row.get("colorway"):
        row["colorway"] = colorway
    if model and not row.get("model"):
        row["model"] = model


def load_from_logs() -> dict[str, dict[str, str]]:
    by_file: dict[str, dict[str, str]] = {}
    for log_file in sorted(LOGS_DIR.glob("*.json")):
        try:
            data = json.loads(log_file.read_text(encoding="utf-8"))
        except Exception:
            continue

        records = data if isinstance(data, list) else [data]
        for record in records:
            if not isinstance(record, dict):
                continue
            filename = str(record.get("filename") or "").strip()
            colorway = str(record.get("color_palette") or "").strip() or None
            model = str(record.get("generation_model") or "").strip() or None
            merge_meta(by_file, filename, colorway, model)
    return by_file


def infer_from_job_payload(payload: dict[str, Any]) -> tuple[str | None, str | None]:
    model: str | None = None
    colorway: str | None = None

    visual_mode = payload.get("visualMode")
    if isinstance(visual_mode, str):
        model = VISUAL_MODE_TO_MODEL.get(visual_mode)

    palette = payload.get("palette")
    if isinstance(palette, int):
        colorway = PALETTE_OPTIONS[palette % len(PALETTE_OPTIONS)]
    elif isinstance(palette, str) and palette.isdigit():
        idx = int(palette)
        colorway = PALETTE_OPTIONS[idx % len(PALETTE_OPTIONS)]

    return colorway, model


def load_from_jobs_db() -> dict[str, dict[str, str]]:
    by_file: dict[str, dict[str, str]] = {}
    if not APP_DB.exists():
        return by_file

    conn = sqlite3.connect(APP_DB)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT payload_json, generated_files FROM jobs WHERE generated_files IS NOT NULL AND generated_files != ''"
        ).fetchall()
    except Exception:
        conn.close()
        return by_file

    for row in rows:
        try:
            payload = json.loads(row["payload_json"] or "{}")
        except Exception:
            payload = {}

        colorway, model = infer_from_job_payload(payload)

        try:
            generated_files = json.loads(row["generated_files"] or "[]")
        except Exception:
            generated_files = []

        if not isinstance(generated_files, list):
            continue

        for name in generated_files:
            filename = str(name or "").strip()
            merge_meta(by_file, filename, colorway, model)

    conn.close()
    return by_file


def apply_to_sheet(file_path: Path, meta: dict[str, dict[str, str]], force: bool, dry_run: bool) -> tuple[int, int]:
    if not file_path.exists():
        return 0, 0

    wb = load_workbook(file_path)
    ws = wb["Designs"]

    headers = header_map(ws)
    filename_col = headers.get("Filename", 2)
    colorway_col = ensure_header_column(ws, "Colorway")
    model_col = ensure_header_column(ws, "Generation Model")

    updated = 0
    considered = 0

    for row in range(3, ws.max_row + 1):
        filename = ws.cell(row=row, column=filename_col).value
        if not filename:
            continue
        key = str(filename).strip()
        info = meta.get(key)
        if not info:
            continue

        considered += 1
        current_colorway = str(ws.cell(row=row, column=colorway_col).value or "").strip()
        current_model = str(ws.cell(row=row, column=model_col).value or "").strip()

        new_colorway = info.get("colorway", "").strip()
        new_model = info.get("model", "").strip()

        changed = False
        if new_colorway and (force or not current_colorway):
            ws.cell(row=row, column=colorway_col, value=new_colorway)
            changed = True
        if new_model and (force or not current_model):
            ws.cell(row=row, column=model_col, value=new_model)
            changed = True

        if changed:
            updated += 1

    if not dry_run:
        wb.save(file_path)
    wb.close()
    return considered, updated


def main() -> None:
    parser = argparse.ArgumentParser(description="Backfill Colorway and Generation Model for existing design rows")
    parser.add_argument("--force", action="store_true", help="Overwrite non-empty Colorway/Generation Model cells")
    parser.add_argument("--dry-run", action="store_true", help="Preview only; do not write files")
    args = parser.parse_args()

    meta = load_from_logs()
    from_jobs = load_from_jobs_db()

    for filename, info in from_jobs.items():
        merge_meta(meta, filename, info.get("colorway"), info.get("model"))

    a_considered, a_updated = apply_to_sheet(A_FILE, meta, force=args.force, dry_run=args.dry_run)
    b_considered, b_updated = apply_to_sheet(B_FILE, meta, force=args.force, dry_run=args.dry_run)

    print(f"Matched rows: A={a_considered}, B={b_considered}")
    print(f"Updated rows: A={a_updated}, B={b_updated}")
    if args.dry_run:
        print("Dry run only; no files were changed.")


if __name__ == "__main__":
    main()
