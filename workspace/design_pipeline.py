"""
design_pipeline.py  --  End-to-End Design Pipeline Orchestrator

Chains all pipeline stages: prompt generation -> image rendering ->
quality inspection -> trademark screening -> pipeline JSON output.
Feeds directly into update_workbooks.py for spreadsheet tracking.

Usage:
    # Generate prompts + TM check for Front A (no image generation)
    python design_pipeline.py batch --front A --drop DROP-01

    # Generate prompts for Front B from a phrase list
    python design_pipeline.py batch --front B --phrases phrases.csv --niche Cottagecore --sub-niche Botanical

    # Full pipeline: generate prompts + render images via Ideogram + inspect + TM check
    python design_pipeline.py batch --front A --drop DROP-01 --render ideogram

    # Inspect an existing folder of designs and run TM checks
    python design_pipeline.py process --front A --folder front_a_sneaker/designs --drop DROP-01

    # After pipeline completes, apply to spreadsheets:
    python update_workbooks.py --log logs/<output_file>.json --front A

Setup:
    For image generation, set API keys as environment variables:
        export IDEOGRAM_API_KEY="your_key"
        export LEONARDO_API_KEY="your_key"
"""

import os
import sys
import csv
import json
import time
import base64
import argparse
from datetime import date

import requests
import numpy as np
from PIL import Image, ImageDraw
from dotenv import load_dotenv

# ── Load .env from workspace root ─────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# ── Import sibling modules ────────────────────────────────────────
WORKSPACE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, WORKSPACE)

from inspect_designs import inspect_design, batch_inspect
from trademark_check import screen_phrase

# ── API Keys (from environment) ───────────────────────────────────
IDEOGRAM_API_KEY = os.environ.get("IDEOGRAM_API_KEY", "")
LEONARDO_API_KEY = os.environ.get("LEONARDO_API_KEY", "")
HF_API_TOKEN     = os.environ.get("HF_API_TOKEN", "")
OPENAI_API_KEY   = os.environ.get("OPENAI_API_KEY", "")

# ── Hugging Face Inference Config ─────────────────────────────────
HF_ROUTER = "https://router.huggingface.co/hf-inference/models"
HF_MODEL  = "stabilityai/stable-diffusion-xl-base-1.0"


# ── Background Removal ──────────────────────────────────────────────
# Edge-connected flood fill: removes ONLY white pixels reachable from
# the image border.  White text / graphics inside the design are safe
# because they are surrounded by non-white pixels (not connected to
# the border through white).

def remove_background(filepath, tolerance=20):
    """Remove the background from a rendered design PNG (legacy white-based).

    Uses flood fill from border seed points on a downscaled copy (fast),
    then refines at full resolution so only genuinely near-white pixels
    are made transparent.  Interior white elements are never touched.
    Kept for backward compatibility — new renders use chroma magenta instead.
    """
    img = Image.open(filepath).convert("RGBA")
    w, h = img.size

    # --- Downscale for fast flood fill (~800px short side) -----------
    scale = max(1, min(w, h) // 800)
    sw, sh = w // scale, h // scale
    small = img.resize((sw, sh), Image.NEAREST).convert("RGB")

    # Flood fill from corners + edge midpoints with a magenta marker.
    # PIL's floodfill spreads through adjacent pixels whose channel
    # values are within ±tolerance of the seed pixel.
    marker = (255, 0, 255)
    seeds = [
        (0, 0), (sw - 1, 0), (0, sh - 1), (sw - 1, sh - 1),
        (sw // 2, 0), (sw // 2, sh - 1),
        (0, sh // 2), (sw - 1, sh // 2),
    ]
    for seed in seeds:
        try:
            ImageDraw.floodfill(small, seed, marker, thresh=tolerance)
        except Exception:
            pass

    # --- Build boolean mask from flood-filled region ----------------
    small_arr = np.array(small)
    bg_small = (
        (small_arr[:, :, 0] == 255)
        & (small_arr[:, :, 1] == 0)
        & (small_arr[:, :, 2] == 255)
    )

    # --- Upscale mask to original resolution ------------------------
    mask_img = Image.fromarray((bg_small * 255).astype(np.uint8))
    mask_full = np.array(mask_img.resize((w, h), Image.NEAREST)) > 128

    # --- Refine: only apply where the full-res pixel is near-white --
    arr = np.array(img)
    white_thresh = max(0, 255 - tolerance * 2)
    is_white = np.all(arr[:, :, :3] > white_thresh, axis=2)
    final_bg = mask_full & is_white

    # --- Apply transparency -----------------------------------------
    arr[final_bg, 3] = 0

    result = Image.fromarray(arr)
    result.save(filepath, "PNG")

    removed_pct = final_bg.sum() / (w * h) * 100
    print(f"  [BG-RM] Removed background ({removed_pct:.0f}% transparent): "
          f"{os.path.basename(filepath)}")


def _detect_bg_color(img_rgb):
    """Detect the dominant background color by sampling many edge pixels.

    AI renderers often add black borders, so corners alone are unreliable.
    Samples 100 points along all 4 edges, filters out near-black and
    near-white, then returns the most common color (median).
    """
    w, h = img_rgb.size
    samples = []
    # Sample along all 4 edges at regular intervals
    for i in range(100):
        x = int(i / 99 * (w - 1))
        y = int(i / 99 * (h - 1))
        samples.append(img_rgb.getpixel((x, 0)))         # top
        samples.append(img_rgb.getpixel((x, h - 1)))     # bottom
        samples.append(img_rgb.getpixel((0, y)))          # left
        samples.append(img_rgb.getpixel((w - 1, y)))      # right

    # Filter out near-black (< 40 all channels) and near-white (> 220 all)
    filtered = [(r, g, b) for r, g, b in samples
                if not (r < 40 and g < 40 and b < 40)
                and not (r > 220 and g > 220 and b > 220)]

    if not filtered:
        # Fallback: use unfiltered
        filtered = samples

    # Return median color
    rs = sorted(c[0] for c in filtered)
    gs = sorted(c[1] for c in filtered)
    bs = sorted(c[2] for c in filtered)
    mid = len(filtered) // 2
    return rs[mid], gs[mid], bs[mid]


def remove_chroma_bg(filepath, tolerance=80):
    """Remove chroma-key background (magenta #FF00FF) from rendered designs.

    Detects the actual background color by sampling edge pixels (handles
    black borders that AI renderers add), then removes ALL pixels
    matching that color anywhere in the image. Safe for chroma colors
    that never appear in the design itself.
    """
    img = Image.open(filepath).convert("RGBA")
    w, h = img.size

    # --- Detect actual background color from edges ------------------
    bg_r, bg_g, bg_b = _detect_bg_color(img.convert("RGB"))

    # --- Remove ALL pixels near the background color ----------------
    arr = np.array(img)
    color_dist = (
        (arr[:, :, 0].astype(int) - bg_r) ** 2
        + (arr[:, :, 1].astype(int) - bg_g) ** 2
        + (arr[:, :, 2].astype(int) - bg_b) ** 2
    )
    final_bg = color_dist < (tolerance * tolerance * 3)

    # --- Also remove near-black border pixels AI renderers add ------
    is_near_black = (
        (arr[:, :, 0] < 25) & (arr[:, :, 1] < 25) & (arr[:, :, 2] < 25)
    )
    # Only remove black at edges (rows/cols within 3% of border)
    border_x = max(1, int(w * 0.03))
    border_y = max(1, int(h * 0.03))
    edge_mask = np.zeros((h, w), dtype=bool)
    edge_mask[:border_y, :] = True       # top
    edge_mask[-border_y:, :] = True      # bottom
    edge_mask[:, :border_x] = True       # left
    edge_mask[:, -border_x:] = True      # right
    final_bg = final_bg | (is_near_black & edge_mask)

    # --- Apply transparency -----------------------------------------
    arr[final_bg, 3] = 0

    result = Image.fromarray(arr)
    result.save(filepath, "PNG")

    removed_pct = final_bg.sum() / (w * h) * 100
    print(f"  [BG-RM] Removed chroma bg ({removed_pct:.0f}% transparent): "
          f"{os.path.basename(filepath)}")
    return removed_pct


# ── Front A: Sneaker Culture Design Themes ────────────────────────
DESIGN_THEMES = {
    "sneaker_rotation_culture": [
        "rotation_ready", "todays_rotation", "rotation_club",
        "rotation_standard", "rotation_elite",
    ],
    "sneaker_collector_identity": [
        "collector_status", "collectors_club", "pair_society",
        "collectors_uniform", "archive_series",
    ],
    "sneakerhead_humor": [
        "no_crease_club", "one_more_pair", "deadstock_energy",
        "wear_your_pairs", "sneaker_daily",
    ],
    "sneaker_archival_aesthetic": [
        "sneaker_archive", "street_rotation", "sneaker_rotation_society",
        "sneaker_culture", "rotation_society",
    ],
}

PALETTE_OPTIONS = [
    # Original 4
    "black and cream, vintage wash",
    "off-white and charcoal, clean matte finish",
    "forest green and ecru, military aesthetic",
    "washed black and bone white, faded streetwear",
    # Streetwear extended
    "navy blue and gold, luxury streetwear",
    "burgundy and cream, vintage sport aesthetic",
    "rust orange and sand, earth tone warmth",
    "pure black and white, high contrast monochrome",
    "olive drab and tan, utilitarian workwear",
    "slate grey and neon green, tech streetwear",
    # Broader / Front B friendly
    "terracotta and ivory, warm bohemian",
    "deep teal and warm cream, coastal vintage",
    "dusty rose and charcoal, soft modern",
    "mustard yellow and dark brown, retro 70s",
    "lavender and slate, muted pastel",
    "red and black, bold graphic",
]

# ── Visual Styles (mixed into batches for variety) ────────────────
# Each style defines an Ideogram prompt strategy. Some are text-only,
# some combine text with graphic elements. The pipeline cycles through
# these so each batch has a natural mix.

VISUAL_STYLES = [
    # ── Text-only styles ──────────────────────────────────────────
    {
        "id": "bold_distressed",
        "label": "Bold clean typography",
        "prompt": (
            'Minimalist streetwear typography design for t-shirt print, '
            'bold condensed font, clean solid ink fills, '
            'text: "{slogan}", {palette} color palette, '
            'centered layout, vintage streetwear aesthetic, '
            'design fills most of the canvas with tight margins, no large empty areas, '
            'smooth color fields, no grain, no speckle, no noise, no distressed texture, '
            'isolated on solid bright magenta (#FF00FF) background, high contrast, clean edges'
        ),
    },
    {
        "id": "stacked_block",
        "label": "Stacked block letters",
        "prompt": (
            'Bold stacked block letter typography design, '
            'heavy weight condensed sans-serif, '
            'text: "{slogan}", {palette} color palette, '
            'solid matte fills, urban street poster aesthetic, '
            'design fills most of the canvas with tight margins, no large empty areas, '
            'smooth color fields, no grain, no speckle, no noise, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        ),
    },
    {
        "id": "vintage_badge",
        "label": "Vintage badge/stamp",
        "prompt": (
            'Vintage circular badge stamp design, '
            'clean bold border ring with text: "{slogan}", '
            '{palette} color palette, retro streetwear patch aesthetic, '
            'centered composition, fills most of the canvas with tight margins, '
            'smooth color fields, no grain, no speckle, no noise, '
            'isolated on solid bright magenta (#FF00FF) background, '
            't-shirt print ready, high contrast'
        ),
    },
    # ── Graphic + text styles ─────────────────────────────────────
    {
        "id": "sneaker_sole",
        "label": "Sneaker sole tread + text",
        "prompt": (
            'Streetwear graphic design combining abstract sneaker sole '
            'tread pattern with bold text: "{slogan}", '
            '{palette} color palette, clean vector-like fills, '
            'urban street culture aesthetic, no brand logos, '
            'fills most of the canvas with tight margins, no large empty areas, '
            'smooth color fields, no grain, no speckle, no noise, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        ),
    },
    {
        "id": "urban_skyline",
        "label": "Urban skyline silhouette + text",
        "prompt": (
            'Minimalist urban city skyline silhouette with bold '
            'streetwear text: "{slogan}", {palette} color palette, '
            'clean matte finish, street culture aesthetic, '
            'centered composition, isolated on solid bright magenta (#FF00FF) background, '
            'fills most of the canvas with tight margins, no large empty areas, '
            'smooth color fields, no grain, no speckle, no noise, '
            'no brand logos, t-shirt print ready'
        ),
    },
    {
        "id": "sneaker_outline",
        "label": "Abstract sneaker outline + text",
        "prompt": (
            'Abstract minimalist sneaker outline illustration with '
            'bold condensed text: "{slogan}", {palette} color palette, '
            'line art style, streetwear graphic tee design, '
            'no specific brand shoe, no logos, clean matte finish, '
            'fills most of the canvas with tight margins, no large empty areas, '
            'smooth color fields, no grain, no speckle, no noise, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        ),
    },
    {
        "id": "shelf_archive",
        "label": "Collector shelf graphic + text",
        "prompt": (
            'Minimalist line art illustration of a sneaker collection shelf '
            'with bold streetwear text: "{slogan}", {palette} color palette, '
            'collector culture aesthetic, no brand logos, no specific shoes, '
            'clean matte finish, fills most of the canvas with tight margins, '
            'smooth color fields, no grain, no speckle, no noise, '
            'isolated on solid bright magenta (#FF00FF) background, '
            't-shirt print ready'
        ),
    },
    {
        "id": "drip_paint",
        "label": "Paint drip graphic + text",
        "prompt": (
            'Bold streetwear typography with paint drip and splatter effects, '
            'text: "{slogan}", {palette} color palette, '
            'graffiti street art influenced, urban culture aesthetic, '
            'high contrast, no brand logos, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        ),
    },
]

FRONT_A_NEGATIVE = (
    "nike, adidas, jordan, supreme, off-white, yeezy, brand logo, swoosh, "
    "three stripes, jumpman, realistic sneaker photo, photorealistic, "
    "blurry, watermark, cluttered background, nsfw, misspelling, "
    "magenta elements in design, pink text, magenta graphics"
)

# ── Filename collision avoidance ─────────────────────────────────
# Track filenames within a batch to prevent duplicates, and check
# all relevant folders (designs, approved, rejected) + spreadsheet
# to avoid overwriting or reusing names from previous runs.

_batch_filenames = set()
_known_filenames = set()   # populated once per pipeline run from disk + spreadsheet


def _load_known_filenames(output_dir):
    """Scan designs/approved/rejected folders and spreadsheet for all used filenames."""
    global _known_filenames
    _known_filenames = set()

    if not output_dir:
        return

    parent = os.path.dirname(os.path.abspath(output_dir))
    for sibling in ("designs", "approved", "rejected"):
        sibling_dir = os.path.join(parent, sibling)
        if os.path.isdir(sibling_dir):
            for f in os.listdir(sibling_dir):
                if f.lower().endswith(".png"):
                    _known_filenames.add(f)

    # Also check the spreadsheet for filenames that may no longer be on disk
    front_code = None
    if "front_a" in output_dir.replace("\\", "/").lower():
        front_code = "A"
    elif "front_b" in output_dir.replace("\\", "/").lower():
        front_code = "B"

    if front_code:
        sheet_map = {
            "A": os.path.join(WORKSPACE, "spreadsheets", "designs_front_a.xlsx"),
            "B": os.path.join(WORKSPACE, "spreadsheets", "designs_front_b.xlsx"),
        }
        sheet_path = sheet_map.get(front_code)
        if sheet_path and os.path.isfile(sheet_path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(sheet_path, read_only=True)
                ws = wb["Designs"]
                # Find Filename column in header row 2
                fname_col = None
                for idx, cell in enumerate(ws[2], start=1):
                    if isinstance(cell.value, str) and cell.value.strip() == "Filename":
                        fname_col = idx
                        break
                if fname_col:
                    for row in ws.iter_rows(min_row=3, max_col=fname_col, min_col=fname_col):
                        val = row[0].value
                        if isinstance(val, str) and val.strip():
                            _known_filenames.add(val.strip())
                wb.close()
            except Exception:
                pass  # spreadsheet read failure is non-fatal

    if _known_filenames:
        print(f"  [NAMES] Loaded {len(_known_filenames)} existing filenames to avoid collisions")


def _safe_filename(base_name, output_dir=None):
    """
    Generate a unique filename like base_name_001.png.
    Increments the suffix if the name is already used in this batch,
    exists on disk (designs/approved/rejected), or in the spreadsheet.
    """
    global _batch_filenames
    suffix = 1
    while True:
        candidate = f"{base_name}_{str(suffix).zfill(3)}.png"
        conflict = candidate in _batch_filenames or candidate in _known_filenames
        if not conflict:
            _batch_filenames.add(candidate)
            return candidate
        suffix += 1


# ── Prompt Builders ───────────────────────────────────────────────

def build_sneaker_prompt(design_name, theme, palette, drop_id="DROP-01",
                         output_dir=None):
    """
    Build a complete pipeline record for a Front A sneaker culture design.
    Cycles through VISUAL_STYLES so each batch has a natural mix of
    text-only and graphic+text designs.
    """
    import random
    slogan = design_name.replace("_", " ").upper()
    style = random.choice(VISUAL_STYLES)

    return {
        # Spreadsheet fields (required by update_workbooks.py --front A)
        "filename":    _safe_filename(design_name, output_dir),
        "drop_id":     drop_id,
        "drop_theme":  theme.replace("_", " ").title(),
        "design_name": design_name.replace("_", " ").title(),
        "slogan":      slogan,
        "style":       style["label"],
        "resolution":  "4500x5400",
        "contrast_ok": None,
        "tm_checked":  False,
        "ip_risk":     "LOW",
        "brand_ref":   False,
        "approved":    None,
        "status":      "Pending Upload",
        "uspto_result":    "",
        "google_check":    "",
        "etsy_check":      "",
        "substring_match": "None",
        "tm_notes":        "",
        # Ideogram renders the complete design (text + optional graphic)
        "image_prompt": style["prompt"].format(slogan=slogan, palette=palette),
        "negative_prompt": FRONT_A_NEGATIVE,
        "color_palette": palette,
        # No Pillow overlay needed when using Ideogram
        "_text_overlay": slogan,
    }


VISUAL_STYLES_B = [
    # ── Illustration + text (Ideogram renders both) ───────────────
    {
        "id": "illustrated_quote",
        "label": "Illustrated quote design",
        "prompt": (
            '{niche} aesthetic t-shirt design, {sub_niche} themed '
            'illustration with hand-lettered text: "{phrase}", '
            'warm earth tones, cozy inviting style, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready, '
            'high contrast, clean edges, no brand logos'
        ),
    },
    {
        "id": "botanical_frame",
        "label": "Botanical frame with text",
        "prompt": (
            'Delicate botanical wreath frame illustration with '
            'text: "{phrase}" in the center, '
            '{niche} aesthetic, {sub_niche} elements, '
            'warm muted earth tones, minimalist line art, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        ),
    },
    {
        "id": "icon_and_text",
        "label": "Niche icon + bold text",
        "prompt": (
            'Simple {sub_niche} icon illustration above bold text: '
            '"{phrase}", {niche} aesthetic, '
            'clean minimal design, warm tones, '
            'centered composition, isolated on solid bright magenta (#FF00FF) background, '
            't-shirt print ready, high contrast'
        ),
    },
    # ── Text-only styles ──────────────────────────────────────────
    {
        "id": "handlettered",
        "label": "Hand-lettered typography",
        "prompt": (
            'Hand-lettered typography design, text: "{phrase}", '
            '{niche} aesthetic style, warm earthy color palette, '
            'organic flowing letterforms, slight texture, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        ),
    },
    {
        "id": "retro_sign",
        "label": "Retro sign typography",
        "prompt": (
            'Vintage retro sign typography design, text: "{phrase}", '
            '{niche} aesthetic, {sub_niche} theme, '
            'aged worn texture, muted warm tones, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        ),
    },
    # ── Illustration-only (Pillow adds text after) ────────────────
    {
        "id": "illustration_only",
        "label": "Illustration only (text added later)",
        "prompt": (
            '{sub_niche} themed illustration, {niche} aesthetic, '
            'warm earth tones, minimalist style, '
            'clean centered composition with space for text, '
            'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready, '
            'high contrast, no text, no letters, no words'
        ),
        "text_overlay": True,
    },
]

FRONT_B_NEGATIVE = (
    "brand logo, watermark, blurry, low quality, realistic human face, "
    "nsfw, copyrighted character, disney, marvel, pokemon, nintendo, "
    "busy background, misspelling, "
    "magenta elements in design, pink text, magenta graphics"
)


def build_general_prompt(phrase, niche, sub_niche, style=None,
                         output_dir=None):
    """
    Build a complete pipeline record for a Front B generalized design.
    Randomly picks a visual style — some have text rendered by AI,
    some are illustration-only with text added by Pillow.
    """
    import random
    safe_name = phrase.lower().replace(" ", "_")[:40]
    vis = random.choice(VISUAL_STYLES_B) if style is None else VISUAL_STYLES_B[0]

    needs_overlay = vis.get("text_overlay", False)

    return {
        # Spreadsheet fields (required by update_workbooks.py --front B)
        "filename":    _safe_filename(safe_name, output_dir),
        "niche":       niche,
        "sub_niche":   sub_niche,
        "phrase":      phrase,
        "style":       vis["label"],
        "resolution":  "3600x4500",
        "contrast_ok": None,
        "tm_checked":  False,
        "ip_risk":     "LOW",
        "approved":    None,
        "status":      "Pending Upload",
        "notes":       "",
        "uspto_result":    "",
        "google_check":    "",
        "etsy_check":      "",
        "substring_match": "None",
        "tm_notes":        "",
        "image_prompt": vis["prompt"].format(
            phrase=phrase, niche=niche, sub_niche=sub_niche,
        ),
        "negative_prompt": FRONT_B_NEGATIVE,
        # Only needs Pillow text overlay if style is illustration-only
        "_text_overlay": phrase if needs_overlay else None,
    }


# ── Image Generation ──────────────────────────────────────────────

def render_ideogram(record, output_dir):
    """
    Call Ideogram API to render a design. Saves PNG and returns filepath.
    Ideogram excels at text rendering — prompt includes the actual slogan,
    so no Pillow text overlay is needed afterward.
    Uses V_2_TURBO ($0.05/image) by default.
    """
    if not IDEOGRAM_API_KEY:
        print("  [SKIP] IDEOGRAM_API_KEY not set")
        return None

    # Build Ideogram-specific prompt WITH text (Ideogram handles text well)
    slogan = record.get("slogan") or record.get("phrase", "")
    palette = record.get("color_palette", "black and cream, vintage wash")
    style = record.get("style", "bold condensed streetwear typography")

    ideogram_prompt = (
        f"Minimalist streetwear typography design for t-shirt print, "
        f'{style}, text: "{slogan}", '
        f"{palette} color palette, clean matte finish, centered layout, "
        f"vintage streetwear aesthetic, isolated on solid bright magenta (#FF00FF) background, "
        f"high contrast, clean edges, t-shirt print ready, "
        f"design fills most of the canvas with tight margins, no large empty areas, "
        f"smooth color fields, no grain, no speckle, no noise"
    )

    hint = record.get("_prompt_hint", "")
    if hint:
        ideogram_prompt += f" {hint}"

    try:
        response = requests.post(
            "https://api.ideogram.ai/generate",
            headers={
                "Api-Key": IDEOGRAM_API_KEY,
                "Content-Type": "application/json",
            },
            json={
                "image_request": {
                    "prompt": ideogram_prompt,
                    "aspect_ratio": "ASPECT_3_4",
                    "model": "V_2_TURBO",
                    "style_type": "DESIGN",
                    "negative_prompt": (
                        "nike, adidas, jordan, supreme, brand logo, swoosh, "
                        "realistic sneaker, photorealistic, blurry, watermark, "
                        "cluttered background, nsfw, "
                        "magenta elements in design, pink text, magenta graphics"
                    ),
                }
            },
            timeout=90,
        )
        response.raise_for_status()
        image_url = response.json()["data"][0]["url"]

        # Download the image
        img_resp = requests.get(image_url, timeout=30)
        img_resp.raise_for_status()

        filepath = os.path.join(output_dir, record["filename"])
        with open(filepath, "wb") as f:
            f.write(img_resp.content)

        # Mark that text overlay is NOT needed (Ideogram rendered it)
        record["_text_overlay"] = None

        print(f"  [RENDER] Ideogram V2 Turbo: {record['filename']} (~$0.05)")
        return filepath

    except Exception as e:
        print(f"  [ERR] Ideogram render failed for {record['filename']}: {e}")
        return None


def render_leonardo(record, output_dir):
    """Call Leonardo.ai API to render a design. Saves PNG and returns filepath."""
    if not LEONARDO_API_KEY:
        print("  [SKIP] LEONARDO_API_KEY not set")
        return None

    try:
        # Step 1: Create generation
        gen_response = requests.post(
            "https://cloud.leonardo.ai/api/rest/v1/generations",
            headers={
                "Authorization": f"Bearer {LEONARDO_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "prompt": record["image_prompt"],
                "negative_prompt": record.get("negative_prompt", ""),
                "modelId": "b24e16ff-06e3-43eb-8d33-4416c2d75876",  # Phoenix
                "width": 832,
                "height": 1040,
                "num_images": 1,
                "guidance_scale": 7,
                "presetStyle": "ILLUSTRATION",
            },
            timeout=30,
        )
        gen_response.raise_for_status()
        generation_id = gen_response.json()["sdGenerationJob"]["generationId"]

        # Step 2: Poll for completion
        for _ in range(40):
            time.sleep(3)
            poll = requests.get(
                f"https://cloud.leonardo.ai/api/rest/v1/generations/{generation_id}",
                headers={"Authorization": f"Bearer {LEONARDO_API_KEY}"},
                timeout=15,
            )
            data = poll.json().get("generations_by_pk", {})
            if data.get("status") == "COMPLETE":
                images = data.get("generated_images", [])
                if images:
                    image_url = images[0]["url"]
                    img_resp = requests.get(image_url, timeout=30)
                    img_resp.raise_for_status()
                    filepath = os.path.join(output_dir, record["filename"])
                    with open(filepath, "wb") as f:
                        f.write(img_resp.content)
                    print(f"  [RENDER] Leonardo: {record['filename']}")
                    return filepath

        print(f"  [ERR] Leonardo timed out for {record['filename']}")
        return None

    except Exception as e:
        print(f"  [ERR] Leonardo render failed for {record['filename']}: {e}")
        return None


def render_huggingface(record, output_dir):
    """
    Call Hugging Face free Inference API (SDXL) to render a design.
    Generates graphic/texture only — text is added by Pillow in a later stage.
    Saves PNG and returns filepath.
    Returns "FALLBACK" (string) on 402 Payment Required so stage_render
    can try Ideogram instead.
    """
    if not HF_API_TOKEN:
        print("  [SKIP] HF_API_TOKEN not set")
        return "FALLBACK"

    try:
        url = f"{HF_ROUTER}/{HF_MODEL}"
        headers = {"Authorization": f"Bearer {HF_API_TOKEN}"}
        hf_prompt = record["image_prompt"]
        hint = record.get("_prompt_hint", "")
        if hint:
            hf_prompt += f" {hint}"
        payload = {
            "inputs": hf_prompt,
            "parameters": {
                "negative_prompt": record.get("negative_prompt", ""),
                "width": 832,
                "height": 1216,
                "guidance_scale": 7.5,
                "num_inference_steps": 30,
            },
        }

        resp = requests.post(url, headers=headers, json=payload, timeout=120)

        if resp.status_code == 402:
            print(f"  [WARN] HF credits exhausted (402) for {record['filename']} -- falling back to Ideogram")
            return "FALLBACK"

        if resp.status_code == 503:
            # Model loading — wait and retry once
            wait = resp.json().get("estimated_time", 30)
            print(f"  [WAIT] Model loading, retrying in {int(wait)}s...")
            time.sleep(min(wait, 60))
            resp = requests.post(url, headers=headers, json=payload, timeout=120)
            if resp.status_code == 402:
                print(f"  [WARN] HF credits exhausted (402) -- falling back to Ideogram")
                return "FALLBACK"

        resp.raise_for_status()

        if "image" not in resp.headers.get("content-type", ""):
            print(f"  [ERR] HF returned non-image: {resp.text[:200]}")
            return None

        filepath = os.path.join(output_dir, record["filename"])
        with open(filepath, "wb") as f:
            f.write(resp.content)

        print(f"  [RENDER] HF/SDXL: {record['filename']}")
        return filepath

    except Exception as e:
        print(f"  [ERR] HF render failed for {record['filename']}: {e}")
        return None


def _openai_generate(prompt, record, output_dir, quality="standard"):
    """Shared DALL-E 3 API call. Returns filepath or None."""
    if not OPENAI_API_KEY:
        print("  [SKIP] OPENAI_API_KEY not set")
        return None
    try:
        response = requests.post(
            "https://api.openai.com/v1/images/generations",
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": "dall-e-3",
                "prompt": prompt,
                "n": 1,
                "size": "1024x1792",
                "quality": quality,
            },
            timeout=120,
        )
        response.raise_for_status()
        image_url = response.json()["data"][0]["url"]
        img_resp = requests.get(image_url, timeout=60)
        img_resp.raise_for_status()
        filepath = os.path.join(output_dir, record["filename"])
        with open(filepath, "wb") as f:
            f.write(img_resp.content)
        return filepath
    except Exception as e:
        print(f"  [ERR] OpenAI render failed for {record['filename']}: {e}")
        return None


def render_openai(record, output_dir, hd=False):
    """
    DALL-E 3: render a complete design with text (typography + graphic).
    Produces a flat 2D artwork on pure white — NOT a t-shirt mockup.
    Sets _text_overlay = None since DALL-E renders the text.
    """
    slogan = record.get("slogan") or record.get("phrase", "")
    palette = record.get("color_palette", "black and cream, vintage wash")
    style = record.get("style", "bold condensed streetwear typography")
    quality = "hd" if hd else "standard"

    prompt = (
        f'Flat 2D streetwear typography design, {style}, '
        f'large bold text reading exactly "{slogan}", '
        f'{palette} color palette, clean matte finish, '
        f'centered balanced composition, portrait orientation. '
        f'Design fills most of the canvas with tight margins (about 85-92% coverage). '
        f'Avoid large empty/blank background areas. '
        f'The design is MINIMAL — only the words "{slogan}" appear as text. '
        f'Do NOT add any other text, numbers, dates, labels, captions, slogans, '
        f'taglines, signatures, or decorative words anywhere in the image. '
        f'Keep it simple and clean — one phrase, bold and prominent. '
        f'Use smooth solid colors and crisp edges. No grain, no speckle, no noise, no distressed texture. '
        f'No t-shirt, no mockup, no garment, no person — just the standalone artwork. '
        f'No brand logos. No photorealistic elements. '
        f'Background is solid bright magenta (#FF00FF), '
        f'completely flat and uniform. No magenta/pink in the design itself.'
    )

    hint = record.get("_prompt_hint", "")
    if hint:
        prompt += f"\nAdditional style direction: {hint}"

    filepath = _openai_generate(prompt, record, output_dir, quality)
    if filepath:
        record["_text_overlay"] = None
        record["_chroma_key"] = True
        cost = "~$0.08" if hd else "~$0.04"
        print(f"  [RENDER] OpenAI DALL-E 3 ({quality}): {record['filename']} ({cost})")
    return filepath


def render_openai_graphic(record, output_dir, hd=False):
    """
    DALL-E 3: render a graphic/illustration only (no text).
    Produces a flat 2D artwork on pure white — NOT a t-shirt mockup.
    Keeps _text_overlay so Pillow can add text later.
    """
    slogan = record.get("slogan") or record.get("phrase", "")
    palette = record.get("color_palette", "black and cream, vintage wash")
    quality = "hd" if hd else "standard"

    prompt = (
        f'Flat 2D streetwear illustration, emblem or badge design, '
        f'{palette} color palette, urban vintage aesthetic, '
        f'centered balanced composition, portrait orientation, '
        f'fills most of the canvas with tight margins (about 85-92% coverage), '
        f'avoid large empty/blank background areas, '
        f'suitable for screen printing. '
        f'Use smooth solid colors and crisp edges. No grain, no speckle, no noise, no distressed texture. '
        f'IMPORTANT: This is a purely visual design with ZERO text. '
        f'No words, no letters, no numbers, no dates, no labels, no captions anywhere. '
        f'No t-shirt, no mockup, no garment, no person — just the standalone artwork. '
        f'No brand logos. No photorealistic elements. '
        f'Background is solid bright magenta (#FF00FF), '
        f'completely flat and uniform. No magenta/pink in the design itself.'
    )

    hint = record.get("_prompt_hint", "")
    if hint:
        prompt += f"\nAdditional style direction: {hint}"

    filepath = _openai_generate(prompt, record, output_dir, quality)
    if filepath:
        record["_chroma_key"] = True
        if slogan and record.get("_text_overlay") is None:
            record["_text_overlay"] = slogan
        cost = "~$0.08" if hd else "~$0.04"
        print(f"  [RENDER] OpenAI DALL-E 3 graphic ({quality}): {record['filename']} ({cost})")
    return filepath


def _gpt_image_generate(prompt, record, output_dir, quality="medium"):
    """Shared GPT Image 1 API call. Returns filepath or None.
    Uses b64_json response format with native transparent backgrounds."""
    if not OPENAI_API_KEY:
        print("  [SKIP] OPENAI_API_KEY not set")
        return None
    try:
        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        }

        payload = {
            "model": "gpt-image-1",
            "prompt": prompt,
            "n": 1,
            "size": "1024x1536",
            "quality": quality,
            "background": "transparent",
            "output_format": "png",
        }
        response = requests.post(
            "https://api.openai.com/v1/images/generations",
            headers=headers,
            json=payload,
            timeout=180,
        )

        if response.status_code == 400:
            detail = ""
            try:
                detail = response.json().get("error", {}).get("message", "")
            except Exception:
                detail = (response.text or "").strip()
            if detail:
                print(f"  [WARN] GPT Image 1 strict payload rejected for {record['filename']}: {detail}")

            # Retry with a more broadly supported payload.
            fallback_payload = {
                "model": "gpt-image-1",
                "prompt": prompt,
                "n": 1,
                "size": "1024x1024",
                "quality": quality,
            }
            response = requests.post(
                "https://api.openai.com/v1/images/generations",
                headers=headers,
                json=fallback_payload,
                timeout=180,
            )

        response.raise_for_status()
        b64_data = response.json()["data"][0]["b64_json"]
        filepath = os.path.join(output_dir, record["filename"])
        with open(filepath, "wb") as f:
            f.write(base64.b64decode(b64_data))
        return filepath
    except Exception as e:
        detail = ""
        try:
            if "response" in locals() and response is not None:
                if response.text:
                    detail = response.text.strip()
        except Exception:
            detail = ""
        if detail:
            print(f"  [ERR] GPT Image 1 render failed for {record['filename']}: {e} | API: {detail}")
        else:
            print(f"  [ERR] GPT Image 1 render failed for {record['filename']}: {e}")
        return None


def render_gpt_image(record, output_dir, quality="medium"):
    """
    GPT Image 1: render a complete design with text (typography + graphic).
    Native transparent background — no chroma-key needed.
    ~98% text rendering accuracy.
    """
    slogan = record.get("slogan") or record.get("phrase", "")
    palette = record.get("color_palette", "black and cream, vintage wash")
    style = record.get("style", "bold condensed streetwear typography")

    prompt = (
        f'Flat 2D streetwear typography design, {style}, '
        f'large bold text reading exactly "{slogan}", '
        f'{palette} color palette, clean matte finish, '
        f'centered balanced composition, portrait orientation. '
        f'Design fills most of the canvas with tight margins (about 85-92% coverage). '
        f'Avoid large empty/blank background areas. '
        f'The design is MINIMAL — only the words "{slogan}" appear as text. '
        f'Do NOT add any other text, numbers, dates, labels, captions, slogans, '
        f'taglines, signatures, or decorative words anywhere in the image. '
        f'Keep it simple and clean — one phrase, bold and prominent. '
        f'Use smooth solid colors and crisp edges. No grain, no speckle, no noise, no distressed texture. '
        f'No t-shirt, no mockup, no garment, no person — just the standalone artwork. '
        f'No brand logos. No photorealistic elements.'
    )

    hint = record.get("_prompt_hint", "")
    if hint:
        prompt += f"\nAdditional style direction: {hint}"

    filepath = _gpt_image_generate(prompt, record, output_dir, quality)
    if filepath:
        record["_text_overlay"] = None
        record["_transparent_bg"] = True
        cost_map = {"low": "~$0.02", "medium": "~$0.06", "high": "~$0.26"}
        cost = cost_map.get(quality, "~$0.06")
        print(f"  [RENDER] GPT Image 1 ({quality}): {record['filename']} ({cost})")
    return filepath


def render_gpt_image_graphic(record, output_dir, quality="medium"):
    """
    GPT Image 1: render a graphic/illustration only (no text).
    Native transparent background — no chroma-key needed.
    Keeps _text_overlay so Pillow can add text later.
    """
    slogan = record.get("slogan") or record.get("phrase", "")
    palette = record.get("color_palette", "black and cream, vintage wash")

    prompt = (
        f'Flat 2D streetwear illustration, emblem or badge design, '
        f'{palette} color palette, urban vintage aesthetic, '
        f'centered balanced composition, portrait orientation, '
        f'fills most of the canvas with tight margins (about 85-92% coverage), '
        f'avoid large empty/blank background areas, '
        f'suitable for screen printing. '
        f'Use smooth solid colors and crisp edges. No grain, no speckle, no noise, no distressed texture. '
        f'IMPORTANT: This is a purely visual design with ZERO text. '
        f'No words, no letters, no numbers, no dates, no labels, no captions anywhere. '
        f'No t-shirt, no mockup, no garment, no person — just the standalone artwork. '
        f'No brand logos. No photorealistic elements.'
    )

    hint = record.get("_prompt_hint", "")
    if hint:
        prompt += f"\nAdditional style direction: {hint}"

    filepath = _gpt_image_generate(prompt, record, output_dir, quality)
    if filepath:
        record["_transparent_bg"] = True
        if slogan and record.get("_text_overlay") is None:
            record["_text_overlay"] = slogan
        cost_map = {"low": "~$0.02", "medium": "~$0.06", "high": "~$0.26"}
        cost = cost_map.get(quality, "~$0.06")
        print(f"  [RENDER] GPT Image 1 graphic ({quality}): {record['filename']} ({cost})")
    return filepath


_esrgan_model = None


def _load_esrgan():
    """Lazy-load Real-ESRGAN x2 model via spandrel. Returns model or None."""
    global _esrgan_model
    if _esrgan_model is not None:
        return _esrgan_model
    try:
        import torch
        import spandrel
        from huggingface_hub import hf_hub_download

        model_path = hf_hub_download("ai-forever/Real-ESRGAN", "RealESRGAN_x2.pth")
        _esrgan_model = spandrel.ModelLoader().load_from_file(model_path).eval()
        _esrgan_model = _esrgan_model.to(torch.device("cpu"))
        print("  [UPSCALE] Real-ESRGAN x2 model loaded (CPU)")
        return _esrgan_model
    except Exception as e:
        print(f"  [WARN] Real-ESRGAN not available, using LANCZOS: {e}")
        return None


def _esrgan_upscale(img_rgb, scale=2, tile_size=384, overlap=16):
    """Upscale an RGB PIL image using Real-ESRGAN with tiled processing."""
    import torch

    model = _load_esrgan()
    if model is None:
        return None

    arr = np.array(img_rgb).astype(np.float32) / 255.0
    h, w, c = arr.shape
    out_h, out_w = h * scale, w * scale
    output = np.zeros((out_h, out_w, c), dtype=np.float32)
    count = np.zeros((out_h, out_w, 1), dtype=np.float32)

    for y in range(0, h, tile_size - overlap):
        for x in range(0, w, tile_size - overlap):
            y2 = min(y + tile_size, h)
            x2 = min(x + tile_size, w)
            y1 = max(0, y2 - tile_size)
            x1 = max(0, x2 - tile_size)
            tile = arr[y1:y2, x1:x2]
            tensor = torch.from_numpy(tile).permute(2, 0, 1).unsqueeze(0)
            with torch.no_grad():
                sr = model(tensor)
            sr_tile = sr.squeeze(0).permute(1, 2, 0).clamp(0, 1).numpy()
            oy1, ox1 = y1 * scale, x1 * scale
            oy2, ox2 = oy1 + sr_tile.shape[0], ox1 + sr_tile.shape[1]
            output[oy1:oy2, ox1:ox2] += sr_tile
            count[oy1:oy2, ox1:ox2] += 1

    output = output / np.maximum(count, 1)
    return Image.fromarray((output * 255).astype(np.uint8))


def stage_trim_and_fill(records, designs_dir, target_w=4500, target_h=5400,
                        padding_pct=0.05, ai_upscale=True):
    """
    After bg removal, trim transparent edges and resize content to fill
    the target canvas. Uses Real-ESRGAN x2 for the heavy upscale (AI
    sharpness), then LANCZOS for final exact sizing.
    """
    trimmed = 0
    for record in records:
        filepath = record.get("_rendered_path")
        if not filepath or not os.path.isfile(filepath):
            filepath = os.path.join(designs_dir, record["filename"])
        if not os.path.isfile(filepath):
            continue

        try:
            img = Image.open(filepath).convert("RGBA")
            bbox = img.getbbox()
            if bbox is None:
                continue

            cropped = img.crop(bbox)
            cw, ch = cropped.size

            pad_x = int(target_w * padding_pct)
            pad_y = int(target_h * padding_pct)
            avail_w = target_w - 2 * pad_x
            avail_h = target_h - 2 * pad_y

            needed_scale = min(avail_w / cw, avail_h / ch)

            # --- AI upscale (Real-ESRGAN x2) if scaling > 1.5x ----------
            if ai_upscale and needed_scale > 1.5:
                # Separate alpha channel, upscale RGB, resize alpha
                alpha = cropped.split()[3]
                rgb = cropped.convert("RGB")

                sr_rgb = _esrgan_upscale(rgb)
                if sr_rgb is not None:
                    sr_w, sr_h = sr_rgb.size
                    sr_alpha = alpha.resize((sr_w, sr_h), Image.LANCZOS)
                    cropped = sr_rgb.convert("RGBA")
                    cropped.putalpha(sr_alpha)
                    cw, ch = sr_w, sr_h
                    needed_scale = min(avail_w / cw, avail_h / ch)
                    print(f"  [UPSCALE] AI x2: {record['filename']} "
                          f"({cw}x{ch})")

            # --- Final resize to exact target with LANCZOS ---------------
            new_w = int(cw * needed_scale)
            new_h = int(ch * needed_scale)
            resized = cropped.resize((new_w, new_h), Image.LANCZOS)

            canvas = Image.new("RGBA", (target_w, target_h), (0, 0, 0, 0))
            paste_x = (target_w - new_w) // 2
            paste_y = (target_h - new_h) // 2
            canvas.paste(resized, (paste_x, paste_y))

            canvas.save(filepath, "PNG")
            record["resolution"] = f"{target_w}x{target_h}"
            trimmed += 1

        except Exception as e:
            print(f"  [ERR] Trim & fill failed for {record['filename']}: {e}")

    if trimmed:
        print(f"\n  Stage: Trim & Fill — {trimmed} design(s) resized to {target_w}x{target_h}")
    return records


def stage_ideogram_text(records, designs_dir):
    """
    Stage 2b (Ideogram): Add text to HuggingFace-generated graphics
    using Ideogram's remix API. Ideogram re-renders the design with
    the text integrated — much higher quality than Pillow overlay.
    Overwrites the original graphic with the text-composited version.
    """
    if not IDEOGRAM_API_KEY:
        print("  [SKIP] IDEOGRAM_API_KEY not set — falling back to Pillow")
        return records

    composited = 0
    for record in records:
        text = record.get("_text_overlay", "")
        if not text:
            continue

        filepath = record.get("_rendered_path")
        if not filepath or not os.path.isfile(filepath):
            filepath = os.path.join(designs_dir, record["filename"])
        if not os.path.isfile(filepath):
            continue

        palette = record.get("color_palette", "black and cream, vintage wash")
        style_label = record.get("style", "bold streetwear typography")

        remix_prompt = (
            f'Streetwear typography design for t-shirt print, '
            f'{style_label}, add bold prominent text: "{text}", '
            f'{palette} color palette, centered layout, '
            f'vintage streetwear aesthetic, high contrast, clean edges, '
            f'isolated on solid bright magenta (#FF00FF) background, t-shirt print ready'
        )

        try:
            with open(filepath, "rb") as img_file:
                response = requests.post(
                    "https://api.ideogram.ai/remix",
                    headers={"Api-Key": IDEOGRAM_API_KEY},
                    files={
                        "image_file": (
                            record["filename"], img_file, "image/png"
                        ),
                    },
                    data={
                        "image_request": json.dumps({
                            "prompt": remix_prompt,
                            "model": "V_2_TURBO",
                            "style_type": "DESIGN",
                            "image_weight": 50,
                            "negative_prompt": FRONT_A_NEGATIVE,
                        }),
                    },
                    timeout=90,
                )
            response.raise_for_status()
            image_url = response.json()["data"][0]["url"]

            img_resp = requests.get(image_url, timeout=30)
            img_resp.raise_for_status()

            with open(filepath, "wb") as f:
                f.write(img_resp.content)

            record["_text_overlay"] = None
            composited += 1
            print(f"  [REMIX] Ideogram text added: {record['filename']} (~$0.05)")

        except Exception as e:
            print(f"  [ERR] Ideogram remix failed for {record['filename']}: {e}")
            print(f"         Falling back to Pillow for this design.")

    print(f"  [STAGE 2b] Ideogram text composited onto "
          f"{composited}/{len(records)} designs")
    return records


def stage_text_overlay(records, designs_dir, font_path=None, font_size=None):
    """
    Stage 2b: Composite phrase text onto AI-generated graphic backgrounds.
    This is the key step — AI generates the visual, Pillow adds clean text.
    Overwrites the original render with the text-composited version.
    """
    from inspect_designs import add_text_overlay
    composited = 0

    for record in records:
        text = record.get("_text_overlay", "")
        if not text:
            continue

        filepath = record.get("_rendered_path")
        if not filepath or not os.path.isfile(filepath):
            filepath = os.path.join(designs_dir, record["filename"])
        if not os.path.isfile(filepath):
            continue

        try:
            add_text_overlay(
                filepath, text, filepath,
                font_path=font_path,
                font_size=font_size,
                text_color=None,  # auto-detect based on background brightness
            )
            composited += 1
        except Exception as e:
            print(f"  [ERR] Text overlay failed for {record['filename']}: {e}")

    print(f"  [STAGE 2b] Text composited onto {composited}/{len(records)} designs")
    return records


def stage_remove_bg(records, designs_dir):
    """Stage 2c: Remove background from rendered designs.
    All renderers request magenta (#FF00FF) chroma-key backgrounds.
    Tries chroma removal first; if it removes < 5% of pixels,
    falls back to legacy white-bg flood fill."""
    removed = 0
    for record in records:
        if record.get("_transparent_bg"):
            print(f"  [BG-RM] Skipped (native transparent bg): {record['filename']}")
            removed += 1
            continue
        filepath = record.get("_rendered_path") or \
                   os.path.join(designs_dir, record["filename"])
        if os.path.isfile(filepath):
            try:
                pct = remove_chroma_bg(filepath)
                if pct < 5:
                    # Chroma didn't find much — try legacy white removal
                    print(f"  [BG-RM] Chroma only {pct:.0f}%, trying white removal...")
                    remove_background(filepath)
                removed += 1
            except Exception as e:
                print(f"  [ERR] Background removal failed for "
                      f"{record['filename']}: {e}")
    print(f"  [STAGE 2c] Background removed from "
          f"{removed}/{len(records)} designs")
    return records


# ── Pipeline Stages ───────────────────────────────────────────────

def stage_generate_prompts_a(drop_id, themes=None, palette_index=0,
                             output_dir=None):
    """Stage 1: Generate prompt records for Front A."""
    global _batch_filenames
    _batch_filenames = set()
    _load_known_filenames(output_dir)
    palette = PALETTE_OPTIONS[palette_index % len(PALETTE_OPTIONS)]
    use_themes = themes or DESIGN_THEMES
    records = []
    for theme, names in use_themes.items():
        for name in names:
            record = build_sneaker_prompt(name, theme, palette, drop_id, output_dir)
            record["palette_index"] = palette_index % len(PALETTE_OPTIONS)
            records.append(record)
    print(f"  [STAGE 1] Generated {len(records)} Front A prompt records")
    return records


def stage_generate_prompts_b(phrases_csv=None, phrases=None,
                             niche="General", sub_niche="General",
                             style="minimalist line art", output_dir=None):
    """Stage 1: Generate prompt records for Front B from phrases."""
    global _batch_filenames
    _batch_filenames = set()
    _load_known_filenames(output_dir)
    phrase_list = list(phrases or [])

    if phrases_csv and os.path.isfile(phrases_csv):
        with open(phrases_csv, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0].strip():
                    val = row[0].strip()
                    if val.lower() not in ("phrase", "phrases", "text"):
                        phrase_list.append(val)

    records = []
    for phrase in phrase_list:
        records.append(build_general_prompt(
            phrase, niche, sub_niche, style, output_dir))
    print(f"  [STAGE 1] Generated {len(records)} Front B prompt records")
    return records


def stage_render(records, renderer, output_dir):
    """Stage 2: Render images via API, auto-upscale to print resolution.
    If a renderer returns "FALLBACK", tries render_ideogram as fallback."""
    from inspect_designs import upscale_if_needed
    os.makedirs(output_dir, exist_ok=True)
    rendered = 0
    fallback_count = 0
    for record in records:
        record["_render_attempted"] = True
        filepath = renderer(record, output_dir)
        if filepath == "FALLBACK":
            filepath = render_ideogram(record, output_dir)
            if filepath:
                fallback_count += 1
        if filepath and filepath != "FALLBACK":
            upscale_if_needed(filepath)
            record["_rendered_path"] = filepath
            rendered += 1
        time.sleep(1)  # respect rate limits
    msg = f"  [STAGE 2] Rendered {rendered}/{len(records)} images"
    if fallback_count:
        msg += f" ({fallback_count} via Ideogram fallback)"
    print(msg)
    return records


def stage_inspect(records, designs_dir):
    """Stage 3: Quality inspection on rendered/existing images."""
    inspected = 0
    for record in records:
        filepath = record.get("_rendered_path") or \
                   os.path.join(designs_dir, record["filename"])

        if os.path.isfile(filepath):
            result = inspect_design(filepath)
            record["contrast_ok"] = result["ok"]
            if not result["ok"]:
                record["tm_notes"] = "; ".join(result["issues"])
            inspected += 1
            icon = "PASS" if result["ok"] else "FAIL"
            print(f"  [INSPECT] [{icon}] {record['filename']}: "
                  f"{', '.join(result['issues']) or 'OK'}")
        else:
            record["contrast_ok"] = None
            print(f"  [INSPECT] [SKIP] {record['filename']}: file not found")

    print(f"  [STAGE 3] Inspected {inspected}/{len(records)} designs")
    return records


def stage_trademark(records, skip_api=False):
    """Stage 4: Trademark screening on all phrases/slogans."""
    checked = 0
    for record in records:
        phrase = record.get("slogan") or record.get("phrase", "")
        if not phrase:
            continue

        result = screen_phrase(phrase, skip_api=skip_api)
        record["tm_checked"] = True
        record["ip_risk"] = result["risk"]
        record["substring_match"] = result.get("matched") or "None"
        record["uspto_result"] = result["detail"]
        record["tm_notes"] = result["detail"]

        if result["status"] == "FLAGGED":
            record["approved"] = False
            record["brand_ref"] = result["reason"] == "substring_match"
            print(f"  [TM] [FLAG] '{phrase}' -- {result['detail']}")
        else:
            print(f"  [TM] [SAFE] '{phrase}'")

        checked += 1

    print(f"  [STAGE 4] Screened {checked}/{len(records)} phrases")
    return records


def stage_approve(records):
    """Stage 5: Auto-approve records that pass both inspection and TM check."""
    approved = 0
    for record in records:
        if record.get("approved") is False:
            # Already rejected by TM check
            continue

        contrast_ok = record.get("contrast_ok")
        tm_checked = record.get("tm_checked", False)
        ip_risk = record.get("ip_risk", "LOW")

        if contrast_ok and tm_checked and ip_risk == "LOW":
            record["approved"] = True
            approved += 1
        elif contrast_ok is None:
            # No image to inspect — mark for review
            if tm_checked and ip_risk == "LOW":
                record["approved"] = True
                record["tm_notes"] = (record.get("tm_notes", "") +
                                      " (image not inspected)").strip()
                approved += 1
            else:
                record["approved"] = False
        else:
            record["approved"] = False

    rejected = sum(1 for r in records if not r.get("approved"))
    print(f"  [STAGE 5] Approved: {approved}, Rejected: {rejected}")
    return records


def stage_output(records, output_path):
    """Stage 6: Write pipeline JSON for update_workbooks.py ingestion.
    Skips records where rendering was attempted but failed (_render_attempted
    is set by stage_render but _rendered_path is missing)."""
    clean_records = []
    skipped = 0
    for record in records:
        if record.get("_render_attempted") and not record.get("_rendered_path"):
            skipped += 1
            continue
        clean = {k: v for k, v in record.items()
                 if not k.startswith("_") and k not in
                 ("image_prompt", "negative_prompt")}
        clean_records.append(clean)
    if skipped:
        print(f"  [OUTPUT] Skipped {skipped} records with no rendered image")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(clean_records, f, indent=2)

    approved = sum(1 for r in clean_records if r.get("approved"))
    print(f"\n  Pipeline output: {len(clean_records)} records "
          f"({approved} approved) -> {output_path}")
    if clean_records:
        print(f"  Next: python update_workbooks.py --log {output_path} "
              f"--front {'A' if clean_records[0].get('drop_id') else 'B'}")
    else:
        print("  No rendered designs to process. Check your API key and renderer settings.")
    return output_path


# ── CLI ───────────────────────────────────────────────────────────

def cmd_batch(args):
    """Full batch pipeline: generate prompts -> inspect -> TM -> output."""
    today = date.today().strftime("%Y%m%d")

    if args.front == "A":
        designs_dir = os.path.join(WORKSPACE, "front_a_sneaker", "designs")
        records = stage_generate_prompts_a(
            drop_id=args.drop or "DROP-01",
            palette_index=args.palette or 0,
            output_dir=designs_dir,
        )
        log_name = f"front_a_{args.drop or 'batch'}_{today}.json"
    else:
        designs_dir = os.path.join(WORKSPACE, "front_b_general", "designs")
        records = stage_generate_prompts_b(
            phrases_csv=args.phrases,
            niche=args.niche or "General",
            sub_niche=args.sub_niche or "General",
            style=args.style or "minimalist line art",
            output_dir=designs_dir,
        )
        log_name = f"front_b_batch_{today}.json"

    if not records:
        print("  No records generated. Check inputs.")
        return

    # Limit count if requested
    if args.count and args.count > 0 and len(records) > args.count:
        import random as _rnd
        total = len(records)
        records = _rnd.sample(records, args.count)
        print(f"  [COUNT] Sampled {args.count} of {total} records")

    # Thread prompt hint into records
    prompt_hint = getattr(args, "prompt_hint", "") or ""
    if prompt_hint:
        for r in records:
            r["_prompt_hint"] = prompt_hint

    # Stage 2: Render (optional)
    if args.render:
        import functools
        renderer = {"ideogram": render_ideogram,
                    "leonardo": render_leonardo,
                    "huggingface": render_huggingface,
                    "hf": render_huggingface,
                    "openai": render_openai,
                    "openai_graphic": render_openai_graphic,
                    "gpt_image": render_gpt_image,
                    "gpt_image_graphic": render_gpt_image_graphic}.get(args.render)
        if renderer and getattr(args, "openai_hd", False) and args.render in ("openai", "openai_graphic"):
            renderer = functools.partial(renderer, hd=True)
        gpt_q = getattr(args, "gpt_quality", None)
        if renderer and gpt_q and args.render in ("gpt_image", "gpt_image_graphic"):
            renderer = functools.partial(renderer, quality=gpt_q)
        if renderer:
            records = stage_render(records, renderer, designs_dir)
            # Stage 2b: Add text to rendered graphics
            if not args.no_text_overlay:
                if args.text_renderer == "ideogram":
                    records = stage_ideogram_text(records, designs_dir)
                    # Fallback: any designs that Ideogram didn't handle
                    records = stage_text_overlay(
                        records, designs_dir,
                        font_path=args.font,
                        font_size=args.font_size,
                    )
                else:
                    records = stage_text_overlay(
                        records, designs_dir,
                        font_path=args.font,
                        font_size=args.font_size,
                    )
        else:
            print(f"  Unknown renderer: {args.render}")

        # Stage 2c: Remove background for transparent PNGs
        if not args.no_bg_remove:
            records = stage_remove_bg(records, designs_dir)

            # Stage 2d: Trim transparent edges and fill print area
            target_w = 4500 if args.front == "A" else 3600
            target_h = 5400 if args.front == "A" else 4500
            ai_up = not getattr(args, "no_ai_upscale", False)
            records = stage_trim_and_fill(records, designs_dir, target_w, target_h,
                                          ai_upscale=ai_up)

    # Stage 3: Inspect
    records = stage_inspect(records, designs_dir)

    # Stage 4: Trademark check
    records = stage_trademark(records, skip_api=args.skip_api)

    # Stage 5: Auto-approve
    records = stage_approve(records)

    # Stage 6: Output
    generation_model = args.render or ""
    for record in records:
        record["generation_model"] = generation_model

    output_path = os.path.join(WORKSPACE, "logs", log_name)
    stage_output(records, output_path)


def cmd_process(args):
    """Process existing designs: inspect folder + TM check + output."""
    today = date.today().strftime("%Y%m%d")

    if not os.path.isdir(args.folder):
        print(f"  Folder not found: {args.folder}")
        sys.exit(1)

    files = sorted(f for f in os.listdir(args.folder)
                   if f.lower().endswith(".png"))
    if not files:
        print(f"  No PNGs in {args.folder}")
        return

    print(f"  Processing {len(files)} existing designs from {args.folder}\n")

    # Build minimal records from filenames
    records = []
    for fname in files:
        base = os.path.splitext(fname)[0]
        name = base.replace("_", " ").title()
        slogan = base.replace("_", " ").upper()

        if args.front == "A":
            record = {
                "filename": fname,
                "drop_id": args.drop or "DROP-01",
                "drop_theme": "Manual Import",
                "design_name": name,
                "slogan": slogan,
                "style": "imported design",
                "resolution": "",
                "contrast_ok": None,
                "tm_checked": False,
                "ip_risk": "LOW",
                "brand_ref": False,
                "approved": None,
                "status": "Pending Upload",
                "uspto_result": "",
                "google_check": "",
                "etsy_check": "",
                "substring_match": "None",
                "tm_notes": "",
                "color_palette": "",
                "generation_model": "imported_existing",
            }
        else:
            record = {
                "filename": fname,
                "niche": args.niche or "General",
                "sub_niche": args.sub_niche or "General",
                "phrase": name,
                "style": "imported design",
                "resolution": "",
                "contrast_ok": None,
                "tm_checked": False,
                "ip_risk": "LOW",
                "approved": None,
                "status": "Pending Upload",
                "notes": "",
                "uspto_result": "",
                "google_check": "",
                "etsy_check": "",
                "substring_match": "None",
                "tm_notes": "",
                "color_palette": "",
                "generation_model": "imported_existing",
            }

        # Fill resolution from actual file
        filepath = os.path.join(args.folder, fname)
        from PIL import Image
        img = Image.open(filepath)
        w, h = img.size
        record["resolution"] = f"{w}x{h}"
        record["_rendered_path"] = filepath
        records.append(record)

    # Run inspection and TM stages
    records = stage_inspect(records, args.folder)
    records = stage_trademark(records, skip_api=args.skip_api)
    records = stage_approve(records)

    log_name = f"front_{args.front.lower()}_process_{today}.json"
    output_path = os.path.join(WORKSPACE, "logs", log_name)
    stage_output(records, output_path)


def cmd_variant(args):
    """Generate a colorway variant of a single design with a different palette."""
    import re
    global _batch_filenames
    _batch_filenames = set()
    today = date.today().strftime("%Y%m%d")

    # Pre-load known filenames for collision avoidance
    if args.front == "A":
        _load_known_filenames(os.path.join(WORKSPACE, "front_a_sneaker", "designs"))
    else:
        _load_known_filenames(os.path.join(WORKSPACE, "front_b_general", "designs"))
    palette = PALETTE_OPTIONS[args.palette % len(PALETTE_OPTIONS)]

    # Strip file extension and numeric suffix to get base design name
    clean_name = re.sub(r'(_\d{3})?\.png$', '', args.name)
    clean_name = re.sub(r'_\d{3}$', '', clean_name)

    if args.front == "A":
        designs_dir = os.path.join(WORKSPACE, "front_a_sneaker", "designs")

        # Find which theme this design belongs to
        found_theme = None
        for theme, names in DESIGN_THEMES.items():
            if clean_name in names:
                found_theme = theme
                break

        if not found_theme:
            # Fallback: use design name as its own theme
            found_theme = "variant"
            print(f"  [WARN] '{clean_name}' not in DESIGN_THEMES, using fallback")

        records = [build_sneaker_prompt(
            clean_name, found_theme, palette,
            args.drop or "DROP-01", designs_dir,
        )]
        log_name = f"front_a_variant_{today}.json"
    else:
        designs_dir = os.path.join(WORKSPACE, "front_b_general", "designs")
        phrase = args.phrase or clean_name.replace("_", " ").title()
        records = [build_general_prompt(
            phrase,
            args.niche or "General",
            args.sub_niche or "General",
            output_dir=designs_dir,
        )]
        log_name = f"front_b_variant_{today}.json"

    if not records:
        print("  No records generated.")
        return

    print(f"\n  Generating variant: {clean_name} with palette: {palette}\n")

    # Thread prompt hint into records
    prompt_hint = getattr(args, "prompt_hint", "") or ""
    if prompt_hint:
        for r in records:
            r["_prompt_hint"] = prompt_hint

    # Render
    if args.render:
        import functools
        renderer = {"ideogram": render_ideogram, "leonardo": render_leonardo,
                    "huggingface": render_huggingface,
                    "hf": render_huggingface,
                    "openai": render_openai,
                    "openai_graphic": render_openai_graphic,
                    "gpt_image": render_gpt_image,
                    "gpt_image_graphic": render_gpt_image_graphic}.get(args.render)
        if renderer and getattr(args, "openai_hd", False) and args.render in ("openai", "openai_graphic"):
            renderer = functools.partial(renderer, hd=True)
        gpt_q = getattr(args, "gpt_quality", None)
        if renderer and gpt_q and args.render in ("gpt_image", "gpt_image_graphic"):
            renderer = functools.partial(renderer, quality=gpt_q)
        if renderer:
            records = stage_render(records, renderer, designs_dir)
            if not args.no_text_overlay:
                if args.text_renderer == "ideogram":
                    records = stage_ideogram_text(records, designs_dir)
                    records = stage_text_overlay(
                        records, designs_dir,
                        font_path=args.font, font_size=args.font_size,
                    )
                else:
                    records = stage_text_overlay(
                        records, designs_dir,
                        font_path=args.font, font_size=args.font_size,
                    )

            # Stage 2c: Remove background
            if not args.no_bg_remove:
                records = stage_remove_bg(records, designs_dir)

                # Stage 2d: Trim transparent edges and fill print area
                target_w = 4500 if args.front == "A" else 3600
                target_h = 5400 if args.front == "A" else 4500
                ai_up = not getattr(args, "no_ai_upscale", False)
                records = stage_trim_and_fill(records, designs_dir, target_w, target_h,
                                              ai_upscale=ai_up)

    # Inspect + TM + Approve
    records = stage_inspect(records, designs_dir)
    records = stage_trademark(records, skip_api=args.skip_api)
    records = stage_approve(records)

    output_path = os.path.join(WORKSPACE, "logs", log_name)
    generation_model = args.render or ""
    for record in records:
        record["generation_model"] = generation_model
    stage_output(records, output_path)


def main():
    parser = argparse.ArgumentParser(
        description="End-to-end design pipeline orchestrator"
    )
    sub = parser.add_subparsers(dest="command")

    # ── batch: generate prompts + full pipeline ──
    batch = sub.add_parser("batch",
                           help="Generate prompts and run full pipeline")
    batch.add_argument("--front", required=True, choices=["A", "B"])
    batch.add_argument("--drop", help="Drop ID for Front A (e.g. DROP-01)")
    batch.add_argument("--palette", type=int, default=0,
                       help="Palette index 0-15 for color scheme")
    batch.add_argument("--phrases",
                       help="CSV file with phrases for Front B (one per line)")
    batch.add_argument("--niche", help="Niche name for Front B")
    batch.add_argument("--sub-niche", help="Sub-niche for Front B")
    batch.add_argument("--style", help="Design style for Front B")
    batch.add_argument("--render", choices=["ideogram", "leonardo", "hf", "huggingface",
                                            "openai", "openai_graphic",
                                            "gpt_image", "gpt_image_graphic"],
                       help="Render images via API (hf = Hugging Face free tier)")
    batch.add_argument("--font", help="Path to .ttf font for text overlay")
    batch.add_argument("--font-size", type=int, default=None,
                       help="Font size for text overlay (default: auto-fit)")
    batch.add_argument("--text-renderer", choices=["pillow", "ideogram"],
                       default="pillow",
                       help="How to add text: pillow (local) or ideogram (remix API)")
    batch.add_argument("--no-text-overlay", action="store_true",
                       help="Skip text overlay entirely (graphic-only output)")
    batch.add_argument("--no-bg-remove", action="store_true",
                       help="Skip background removal (keep original background)")
    batch.add_argument("--count", type=int, default=0,
                       help="Limit number of designs to generate (0 = all)")
    batch.add_argument("--openai-hd", action="store_true",
                       help="Use HD quality for OpenAI DALL-E 3 (~$0.08 instead of ~$0.04)")
    batch.add_argument("--gpt-quality", choices=["low", "medium", "high"], default="medium",
                       help="Quality tier for GPT Image 1 (low ~$0.02, medium ~$0.06, high ~$0.26)")
    batch.add_argument("--prompt-hint", type=str, default="",
                       help="Custom style hint appended to the prompt (e.g. 'make it flashy')")
    batch.add_argument("--no-ai-upscale", action="store_true",
                       help="Skip Real-ESRGAN AI upscale (use LANCZOS only)")
    batch.add_argument("--skip-api", action="store_true",
                       help="Skip USPTO API (substring TM check only)")

    # ── process: inspect existing designs ──
    proc = sub.add_parser("process",
                          help="Process existing designs (inspect + TM check)")
    proc.add_argument("--front", required=True, choices=["A", "B"])
    proc.add_argument("--folder", required=True,
                      help="Folder containing design PNGs")
    proc.add_argument("--drop", help="Drop ID for Front A")
    proc.add_argument("--niche", help="Niche for Front B")
    proc.add_argument("--sub-niche", help="Sub-niche for Front B")
    proc.add_argument("--skip-api", action="store_true",
                      help="Skip USPTO API (substring TM check only)")

    # ── variant: generate a colorway variant of a single design ──
    var = sub.add_parser("variant",
                         help="Generate a colorway variant of a single design")
    var.add_argument("--front", required=True, choices=["A", "B"])
    var.add_argument("--name", required=True,
                     help="Design name (e.g. rotation_ready)")
    var.add_argument("--palette", type=int, required=True,
                     help="Palette index 0-3")
    var.add_argument("--drop", help="Drop ID for Front A")
    var.add_argument("--phrase", help="Phrase for Front B variant")
    var.add_argument("--niche", help="Niche for Front B")
    var.add_argument("--sub-niche", help="Sub-niche for Front B")
    var.add_argument("--render",
                     choices=["ideogram", "leonardo", "hf", "huggingface",
                              "openai", "openai_graphic",
                              "gpt_image", "gpt_image_graphic"],
                     help="Render images via API")
    var.add_argument("--text-renderer", choices=["pillow", "ideogram"],
                     default="pillow")
    var.add_argument("--no-text-overlay", action="store_true")
    var.add_argument("--no-bg-remove", action="store_true")
    var.add_argument("--font", help="Path to .ttf font")
    var.add_argument("--font-size", type=int, default=None)
    var.add_argument("--openai-hd", action="store_true",
                     help="Use HD quality for OpenAI DALL-E 3")
    var.add_argument("--gpt-quality", choices=["low", "medium", "high"], default="medium",
                     help="Quality tier for GPT Image 1 (low ~$0.02, medium ~$0.06, high ~$0.26)")
    var.add_argument("--prompt-hint", type=str, default="",
                     help="Custom style hint appended to the prompt")
    var.add_argument("--no-ai-upscale", action="store_true",
                     help="Skip Real-ESRGAN AI upscale (use LANCZOS only)")
    var.add_argument("--skip-api", action="store_true")

    args = parser.parse_args()

    if args.command == "batch":
        cmd_batch(args)
    elif args.command == "process":
        cmd_process(args)
    elif args.command == "variant":
        cmd_variant(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
