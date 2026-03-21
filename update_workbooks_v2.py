"""
update_workbooks.py  —  Dual-Front Edition
Auto-populates spreadsheets from pipeline JSON output.

Usage:
    python update_workbooks.py --log logs/front_a_batch_20260305.json --front A
    python update_workbooks.py --log logs/front_b_batch_20260305.json --front B
    python update_workbooks.py --log logs/batch.json --front A --dry-run
"""

import json
import argparse
import os
from datetime import date
from openpyxl import load_workbook

WORKSPACE = os.path.dirname(os.path.abspath(__file__))
SP = os.path.join(WORKSPACE, "spreadsheets")

DESIGNS_A  = os.path.join(SP, "designs_front_a.xlsx")
DESIGNS_B  = os.path.join(SP, "designs_front_b.xlsx")
TM_LOG     = os.path.join(SP, "trademark_log.xlsx")
NICHES_B   = os.path.join(SP, "niches_front_b.xlsx")


def next_id(ws, prefix, col=1, start_row=3):
    last = 0
    for row in ws.iter_rows(min_row=start_row, max_col=col, values_only=True):
        val = row[0]
        if val and str(val).startswith(prefix):
            try:
                num = int(str(val).split("-")[-1])
                last = max(last, num)
            except (IndexError, ValueError):
                pass
    return f"{prefix}{str(last + 1).zfill(4)}"


def header_map(ws):
    return {
        (cell.value or "").strip(): idx
        for idx, cell in enumerate(ws[2], start=1)
        if isinstance(cell.value, str)
    }


def ensure_header_column(ws, name):
    headers = header_map(ws)
    if name in headers:
        return headers[name]
    col = ws.max_column + 1
    ws.cell(row=2, column=col, value=name)
    return col


def append_design_a(ws, record):
    dsn_id = next_id(ws, "DSN-A-")
    ws.append([
        dsn_id, record.get("filename",""), record.get("drop_id",""),
        record.get("drop_theme",""), record.get("design_name",""),
        record.get("slogan",""), record.get("style",""),
        str(date.today()), record.get("resolution",""),
        "YES" if record.get("contrast_ok") else "NO",
        "YES" if record.get("tm_checked") else "PENDING",
        record.get("ip_risk","LOW"),
        "YES" if record.get("brand_ref") else "NO",
        "YES" if record.get("approved") else "REVIEW",
        record.get("printify_image_id",""), record.get("printify_product_id",""),
        record.get("etsy_listing_id",""), record.get("status","Pending Upload"),
    ])

    row_idx = ws.max_row
    colorway_col = ensure_header_column(ws, "Colorway")
    model_col = ensure_header_column(ws, "Generation Model")
    ws.cell(row=row_idx, column=colorway_col, value=record.get("color_palette", ""))
    ws.cell(row=row_idx, column=model_col, value=record.get("generation_model", ""))

    return dsn_id


def append_design_b(ws, record):
    dsn_id = next_id(ws, "DSN-B-")
    ws.append([
        dsn_id, record.get("filename",""), record.get("niche",""),
        record.get("sub_niche",""), record.get("phrase",""),
        record.get("style",""), str(date.today()),
        record.get("resolution",""),
        "YES" if record.get("contrast_ok") else "NO",
        "YES" if record.get("tm_checked") else "PENDING",
        record.get("ip_risk","LOW"),
        "YES" if record.get("approved") else "REVIEW",
        record.get("printify_image_id",""), record.get("printify_product_id",""),
        record.get("etsy_listing_id",""),
        record.get("status","Pending Upload"), record.get("notes",""),
    ])

    row_idx = ws.max_row
    colorway_col = ensure_header_column(ws, "Colorway")
    model_col = ensure_header_column(ws, "Generation Model")
    ws.cell(row=row_idx, column=colorway_col, value=record.get("color_palette", ""))
    ws.cell(row=row_idx, column=model_col, value=record.get("generation_model", ""))

    return dsn_id


def append_tm_log(ws, record, dsn_id, front):
    tm_id = next_id(ws, "TM-")
    phrase = record.get("slogan") or record.get("phrase","")
    ws.append([
        tm_id, front, dsn_id, phrase, str(date.today()),
        record.get("uspto_result","No match found"),
        record.get("google_check",""), record.get("etsy_check",""),
        record.get("substring_match","None"),
        "YES" if record.get("brand_ref") else "NO",
        record.get("ip_risk","LOW"),
        "APPROVED" if record.get("approved") else "MANUAL REVIEW",
        record.get("tm_notes",""),
    ])


def append_phrase_b(ws, record, dsn_id):
    phr_id = next_id(ws, "PHR-")
    ws.append([
        phr_id, record.get("niche",""), record.get("sub_niche",""),
        record.get("phrase",""),
        "SAFE" if record.get("approved") else "FLAGGED",
        record.get("ip_risk","LOW"), dsn_id, str(date.today()),
    ])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--log", required=True)
    parser.add_argument("--front", required=True, choices=["A","B"])
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    with open(args.log) as f:
        data = json.load(f)
    records = data if isinstance(data, list) else [data]

    tm_wb    = load_workbook(TM_LOG)
    tm_ws    = tm_wb["Clearance Log"]

    if args.front == "A":
        dsn_wb = load_workbook(DESIGNS_A)
        dsn_ws = dsn_wb["Designs"]
    else:
        dsn_wb   = load_workbook(DESIGNS_B)
        dsn_ws   = dsn_wb["Designs"]
        niche_wb = load_workbook(NICHES_B)
        phrase_ws = niche_wb["Phrase Bank"]

    added = 0
    for record in records:
        if not record.get("filename"):
            print(f"  ⚠️  Skipping — no filename: {record}")
            continue

        if args.front == "A":
            dsn_id = append_design_a(dsn_ws, record)
        else:
            dsn_id = append_design_b(dsn_ws, record)
            if record.get("phrase"):
                append_phrase_b(phrase_ws, record, dsn_id)

        append_tm_log(tm_ws, record, dsn_id, args.front)
        added += 1
        icon = "✅" if record.get("approved") else "⚠️ "
        name = record.get("design_name") or record.get("phrase","?")
        print(f"  {icon} [{args.front}] {dsn_id} — {name}")

    if not args.dry_run:
        dsn_wb.save(DESIGNS_A if args.front == "A" else DESIGNS_B)
        tm_wb.save(TM_LOG)
        if args.front == "B":
            niche_wb.save(NICHES_B)
        print(f"\n  ✅ {added} records written to spreadsheets.")
    else:
        print(f"\n  Dry run — {added} records would be written.")


if __name__ == "__main__":
    main()
