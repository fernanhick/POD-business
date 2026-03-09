from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from .models import DesignOption

BASE_DIR = Path(__file__).resolve().parents[4]
WORKSPACE_DIR = BASE_DIR / "workspace"
SPREADSHEETS_DIR = WORKSPACE_DIR / "spreadsheets"
APPROVED_DIR = WORKSPACE_DIR / "front_a_sneaker" / "approved"


def _header_map(ws) -> dict[str, int]:
    return {
        (cell.value or "").strip(): idx
        for idx, cell in enumerate(ws[2], start=1)
        if isinstance(cell.value, str)
    }


def get_approved_designs() -> list[DesignOption]:
    spreadsheet = SPREADSHEETS_DIR / "designs_front_a.xlsx"
    if not spreadsheet.exists():
        return []

    wb = load_workbook(spreadsheet, read_only=True)
    ws = wb["Designs"]
    headers = _header_map(ws)

    designs: list[DesignOption] = []
    filename_col = headers.get("Filename", 2)
    name_col = headers.get("Design Name", headers.get("Slogan / Typography", 5))
    design_id_col = headers.get("Design ID", 1)
    phrase_col = headers.get("Slogan / Typography", 5)
    style_col = headers.get("Style", 6)
    status_col = headers.get("Status")
    approved_col = headers.get("Approved?")

    for row in range(3, ws.max_row + 1):
        filename = ws.cell(row=row, column=filename_col).value
        if not filename:
            continue

        approved_val = ws.cell(row=row, column=approved_col).value if approved_col else None
        status_val = ws.cell(row=row, column=status_col).value if status_col else None

        is_approved = (
            str(approved_val or "").strip().upper() == "YES"
            or str(status_val or "").strip().lower() == "approved"
        )
        if not is_approved:
            continue

        image_path = APPROVED_DIR / str(filename).strip()
        if not image_path.exists():
            continue

        designs.append(DesignOption(
            filename=str(filename).strip(),
            name=ws.cell(row=row, column=name_col).value,
            design_id=str(ws.cell(row=row, column=design_id_col).value or ""),
            phrase=ws.cell(row=row, column=phrase_col).value if phrase_col else None,
            style=ws.cell(row=row, column=style_col).value if style_col else None,
            status="approved",
            image_path=str(image_path),
        ))

    wb.close()

    # Try to merge product URLs from listings.xlsx
    listings_file = SPREADSHEETS_DIR / "listings.xlsx"
    if listings_file.exists():
        _merge_product_urls(designs, listings_file)

    return designs


def _merge_product_urls(designs: list[DesignOption], listings_file: Path) -> None:
    try:
        wb = load_workbook(listings_file, read_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return
        headers = _header_map(ws)
        filename_col = headers.get("Filename", headers.get("Design", 1))
        url_col = headers.get("URL", headers.get("Product URL", headers.get("Listing URL")))
        if not url_col:
            wb.close()
            return

        url_map: dict[str, str] = {}
        for row in range(3, ws.max_row + 1):
            fn = ws.cell(row=row, column=filename_col).value
            url = ws.cell(row=row, column=url_col).value
            if fn and url:
                url_map[str(fn).strip()] = str(url).strip()

        for design in designs:
            if design.filename in url_map:
                design.product_url = url_map[design.filename]

        wb.close()
    except Exception:
        pass


def get_design_by_filename(filename: str) -> Optional[DesignOption]:
    designs = get_approved_designs()
    for d in designs:
        if d.filename == filename:
            return d
    return None
