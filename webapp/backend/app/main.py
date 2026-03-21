from __future__ import annotations

import json
import os
import random
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[3]
WORKSPACE_DIR = BASE_DIR / "workspace"

# Import Printify upload functions from workspace script
sys.path.insert(0, str(WORKSPACE_DIR))
try:
    from printify_upload import (
        upload_image as printify_upload_image,
        create_product as printify_create_product,
        publish_product as printify_publish_product,
        get_product as printify_get_product,
        update_spreadsheet_ids as printify_update_spreadsheet_ids,
        FRONT_CONFIG as PRINTIFY_CONFIG,
    )
    _PRINTIFY_AVAILABLE = True
except ImportError:
    _PRINTIFY_AVAILABLE = False

# Import regional pricing engine
try:
    from pod_pricing import (
        get_profile_for_provider_market,
        US_PRINTIFY,
        EU_STANDARD_21,
    )
    _POD_PRICING_AVAILABLE = True
except ImportError:
    _POD_PRICING_AVAILABLE = False

try:
    from pod_providers import PROVIDER_REGISTRY
    _POD_PROVIDERS_AVAILABLE = True
except ImportError:
    _POD_PROVIDERS_AVAILABLE = False

try:
    from printful_upload import get_product_config as printful_get_product_config
    _PRINTFUL_CONFIG_AVAILABLE = True
except ImportError:
    _PRINTFUL_CONFIG_AVAILABLE = False

try:
    from .pinterest import router as pinterest_router
    _PINTEREST_AVAILABLE = True
except ImportError:
    _PINTEREST_AVAILABLE = False

try:
    from .etsy import router as etsy_router
    _ETSY_AVAILABLE = True
except ImportError:
    _ETSY_AVAILABLE = False

SPREADSHEETS_DIR = WORKSPACE_DIR / "spreadsheets"
LOGS_DIR = WORKSPACE_DIR / "logs"
DB_PATH = BASE_DIR / "webapp" / "backend" / "app" / "app_state.db"

FRONT_CONFIG = {
    "sneaker": {
        "code": "A",
        "design_folder": WORKSPACE_DIR / "front_a_sneaker" / "designs",
        "approved_folder": WORKSPACE_DIR / "front_a_sneaker" / "approved",
        "rejected_folder": WORKSPACE_DIR / "front_a_sneaker" / "rejected",
        "spreadsheet": SPREADSHEETS_DIR / "designs_front_a.xlsx",
        "phrase_col": "Slogan / Typography",
    },
    "general": {
        "code": "B",
        "design_folder": WORKSPACE_DIR / "front_b_general" / "designs",
        "approved_folder": WORKSPACE_DIR / "front_b_general" / "approved",
        "rejected_folder": WORKSPACE_DIR / "front_b_general" / "rejected",
        "spreadsheet": SPREADSHEETS_DIR / "designs_front_b.xlsx",
        "phrase_col": "Phrase / Concept",
    },
}

FINANCIALS_FILE = SPREADSHEETS_DIR / "financials.xlsx"
TM_LOG_FILE = SPREADSHEETS_DIR / "trademark_log.xlsx"


app = FastAPI(title="POD Business Local API", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_origin_regex=r"^https?://([a-zA-Z0-9-]+\.)*[a-zA-Z0-9-]+:5173$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if _PINTEREST_AVAILABLE:
    app.include_router(pinterest_router, prefix="/api/pinterest", tags=["pinterest"])

if _ETSY_AVAILABLE:
    app.include_router(etsy_router, prefix="/api/etsy", tags=["etsy"])


class GenerationRequest(BaseModel):
    designType: Literal["general", "sneaker"]
    visualMode: Literal[
        "random", "text_only", "graphic_text", "graphic_only",
        "text_only_openai", "graphic_openai",
        "text_gpt_image", "graphic_gpt_image",
    ] = "random"
    palette: int = 0
    count: int = 0
    dropId: str | None = None
    phrase: str | None = None
    niche: str | None = None
    subNiche: str | None = None
    skipApi: bool = False
    openaiHd: bool = False
    gptQuality: str = "medium"
    promptHint: str = ""


class VariantRequest(BaseModel):
    designType: Literal["general", "sneaker"]
    designName: str
    palette: int = 0
    visualMode: Literal[
        "text_only", "graphic_text", "graphic_only",
        "text_only_openai", "graphic_openai",
        "text_gpt_image", "graphic_gpt_image",
    ] = "text_only"
    phrase: str | None = None
    niche: str | None = None
    subNiche: str | None = None
    skipApi: bool = False
    openaiHd: bool = False
    gptQuality: str = "medium"
    promptHint: str = ""


class ApprovalRequest(BaseModel):
    designType: Literal["general", "sneaker"]
    filename: str
    approved: bool
    notes: str | None = None


class PrintifyUploadRequest(BaseModel):
    designType: Literal["general", "sneaker"]
    filename: str
    productType: Literal["tshirt", "hoodie"] = "tshirt"
    provider: Literal["printify", "printful"] = "printify"  # Provider choice
    market: Literal["US", "EU"] | None = None  # Market region (defaults by provider)
    draft: bool = False
    country: str | None = None  # Optional for EU VAT bucket selection


class PrintifyCredentialsRequest(BaseModel):
    token: str
    shop_id: str


class PrintfulCredentialsRequest(BaseModel):
    api_key: str
    store_id: str
    api_base: str | None = None


class GenerationCredentialsRequest(BaseModel):
    openai_api_key: str | None = None
    ideogram_api_key: str | None = None
    hf_api_token: str | None = None
    leonardo_api_key: str | None = None


class ExpenseCreate(BaseModel):
    date: str
    front: str = Field(description="A, B, or BOTH")
    category: str
    description: str
    amount: float
    taxDeductible: Literal["Yes", "No"] = "Yes"
    receipt: Literal["Yes", "No"] = "No"
    notes: str = ""


class ExpenseUpdate(ExpenseCreate):
    expenseId: str


def _unique_values_from_sheet(file_path: Path, sheet_name: str, column_name: str) -> list[str]:
    if not file_path.exists():
        return []
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    headers = _header_map(ws)
    column = headers.get(column_name)
    values: set[str] = set()
    if column:
        for row in range(3, ws.max_row + 1):
            value = ws.cell(row=row, column=column).value
            if value:
                values.add(str(value).strip())
    wb.close()
    return sorted([v for v in values if v])


def _get_drop_ids() -> list[str]:
    drops_dir = WORKSPACE_DIR / "front_a_sneaker" / "drops"
    drop_ids: set[str] = set()
    if drops_dir.exists():
        for file_path in drops_dir.glob("*.json"):
            drop_ids.add(file_path.stem)
            try:
                payload = json.loads(file_path.read_text(encoding="utf-8"))
                if isinstance(payload, dict) and payload.get("drop_id"):
                    drop_ids.add(str(payload["drop_id"]).strip())
            except Exception:
                continue
    if not drop_ids:
        drop_ids.add("DROP-01")
    return sorted(drop_ids)


def _get_phrase_bank() -> list[str]:
    phrase_file = SPREADSHEETS_DIR / "niches_front_b.xlsx"
    return _unique_values_from_sheet(phrase_file, "Phrase Bank", "Phrase")


def _build_random_phrase_samples(count: int) -> list[str]:
    samples = _get_phrase_bank()
    token_pool: list[str] = []
    for text in samples:
        token_pool.extend(re.findall(r"[A-Za-z]{3,}", text.lower()))

    token_pool = list({token for token in token_pool if len(token) >= 3})
    if len(token_pool) < 3:
        token_pool = [
            "collectors",
            "street",
            "culture",
            "hustle",
            "vibes",
            "energy",
            "legacy",
            "daily",
            "club",
            "archive",
        ]

    templates = [
        "{a} {b}",
        "{a} {b} {c}",
        "{a} & {b}",
        "{a} over {b}",
        "{a} never {b}",
    ]

    phrases: list[str] = []
    for _ in range(max(1, count)):
        a, b, c = random.sample(token_pool, k=3)
        phrase = random.choice(templates).format(a=a.title(), b=b.title(), c=c.title())
        phrases.append(phrase)
    return phrases


def _build_generation_options() -> dict[str, Any]:
    niche_file = SPREADSHEETS_DIR / "niches_front_b.xlsx"
    design_b_file = SPREADSHEETS_DIR / "designs_front_b.xlsx"

    niches = _unique_values_from_sheet(niche_file, "Niches", "Niche")
    sub_niches = _unique_values_from_sheet(niche_file, "Niches", "Sub-Niche")
    phrases = _get_phrase_bank()
    if not phrases:
        phrases = _unique_values_from_sheet(design_b_file, "Designs", "Phrase / Concept")
    if not phrases:
        phrases = _build_random_phrase_samples(50)

    return {
        "sneaker": {
            "dropIds": _get_drop_ids(),
        },
        "general": {
            "niches": niches,
            "subNiches": sub_niches,
            "phrases": phrases[:200],
        },
    }


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _ensure_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH) as con:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS jobs (
                id TEXT PRIMARY KEY,
                kind TEXT NOT NULL,
                design_type TEXT NOT NULL,
                mode TEXT NOT NULL,
                status TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                started_at TEXT,
                finished_at TEXT,
                output TEXT
            )
            """
        )
        con.commit()
    # Migration: add generated_files column
    with sqlite3.connect(DB_PATH) as con:
        try:
            con.execute("ALTER TABLE jobs ADD COLUMN generated_files TEXT")
            con.commit()
        except sqlite3.OperationalError:
            pass  # Column already exists


@contextmanager
def _db() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        yield con
    finally:
        con.close()


def _list_pngs(path: Path) -> list[str]:
    if not path.exists():
        return []
    return sorted([p.name for p in path.iterdir() if p.suffix.lower() == ".png"])


def _header_map(ws) -> dict[str, int]:
    return {
        (cell.value or "").strip(): idx
        for idx, cell in enumerate(ws[2], start=1)
        if isinstance(cell.value, str)
    }


def _ensure_header_column(ws, headers: dict[str, int], name: str) -> int:
    column = headers.get(name)
    if column:
        return column
    column = ws.max_column + 1
    ws.cell(row=2, column=column, value=name)
    headers[name] = column
    return column


def _update_upload_metadata(
    spreadsheet_path: Path,
    filename: str,
    provider: str,
    market: str,
) -> None:
    wb = load_workbook(spreadsheet_path)
    ws = wb["Designs"]
    headers = _header_map(ws)

    filename_col = headers.get("Filename", 2)
    provider_col = _ensure_header_column(ws, headers, "POD Provider")
    market_col = _ensure_header_column(ws, headers, "POD Market")

    for row in range(3, ws.max_row + 1):
        value = ws.cell(row=row, column=filename_col).value
        if value and str(value).strip() == filename.strip():
            ws.cell(row=row, column=provider_col, value=provider)
            ws.cell(row=row, column=market_col, value=market)
            break

    wb.save(spreadsheet_path)
    wb.close()


def _date_value_to_sort_ts(value: Any) -> float:
    if value is None:
        return 0.0

    if isinstance(value, datetime):
        return value.timestamp()

    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        try:
            return datetime(value.year, value.month, value.day).timestamp()
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return 0.0

    for parser in (
        lambda s: datetime.fromisoformat(s),
        lambda s: datetime.strptime(s, "%Y-%m-%d"),
        lambda s: datetime.strptime(s, "%m/%d/%Y"),
        lambda s: datetime.strptime(s, "%Y/%m/%d"),
    ):
        try:
            return parser(text).timestamp()
        except Exception:
            continue

    return 0.0


def _printify_product_url(product_id: str | None) -> str | None:
    if not product_id:
        return None
    shop_id = os.environ.get("PRINTIFY_SHOP_ID")
    if not shop_id:
        return None
    return f"https://printify.com/app/products/{shop_id}/{product_id}"


def _etsy_listing_url(listing_id: str | None) -> str | None:
    if not listing_id:
        return None
    return f"https://www.etsy.com/listing/{listing_id}"


def _read_design_rows(design_type: str) -> list[dict[str, Any]]:
    cfg = FRONT_CONFIG[design_type]
    wb = load_workbook(cfg["spreadsheet"])
    ws = wb["Designs"]
    headers = _header_map(ws)

    rows: list[dict[str, Any]] = []
    for row in range(3, ws.max_row + 1):
        filename = ws.cell(row=row, column=headers.get("Filename", 2)).value
        if not filename:
            continue
        row_data = {
            "designId": ws.cell(row=row, column=headers.get("Design ID", 1)).value,
            "filename": filename,
            "designType": design_type,
            "name": ws.cell(row=row, column=headers.get("Design Name", headers.get(cfg["phrase_col"], 5))).value,
            "phrase": ws.cell(row=row, column=headers.get(cfg["phrase_col"], 5)).value,
            "style": ws.cell(row=row, column=headers.get("Style", 6)).value,
            "dateCreated": ws.cell(row=row, column=headers.get("Date Created", 7)).value,
            "ipRisk": ws.cell(row=row, column=headers.get("IP Risk", 11)).value,
            "approved": ws.cell(row=row, column=headers.get("Approved?", 12)).value,
            "status": ws.cell(row=row, column=headers.get("Status", ws.max_column)).value,
            "rowNumber": row,
            "printifyImageId": ws.cell(row=row, column=headers["Printify Image ID"]).value if "Printify Image ID" in headers else None,
            "printifyProductId": ws.cell(row=row, column=headers["Printify Product ID"]).value if "Printify Product ID" in headers else None,
            "etsyListingId": ws.cell(row=row, column=headers["Etsy Listing ID"]).value if "Etsy Listing ID" in headers else None,
            "podProvider": ws.cell(row=row, column=headers["POD Provider"]).value if "POD Provider" in headers else None,
            "podMarket": ws.cell(row=row, column=headers["POD Market"]).value if "POD Market" in headers else None,
            "colorway": ws.cell(row=row, column=headers["Colorway"]).value if "Colorway" in headers else None,
            "generationModel": ws.cell(row=row, column=headers["Generation Model"]).value if "Generation Model" in headers else None,
        }
        row_data["printifyUrl"] = _printify_product_url(str(row_data["printifyProductId"]) if row_data["printifyProductId"] else None)
        row_data["etsyUrl"] = _etsy_listing_url(str(row_data["etsyListingId"]) if row_data["etsyListingId"] else None)
        rows.append(row_data)

    wb.close()
    return rows


def _get_design_location(design_type: str, filename: str) -> tuple[Path | None, str | None]:
    cfg = FRONT_CONFIG[design_type]
    locations = {
        "generated": cfg["design_folder"],
        "approved": cfg["approved_folder"],
        "rejected": cfg["rejected_folder"],
    }
    for state, folder in locations.items():
        path = folder / filename
        if path.exists():
            return path, state
    return None, None


def _update_design_sheet_approval(design_type: str, filename: str, approved: bool) -> str | None:
    cfg = FRONT_CONFIG[design_type]
    wb = load_workbook(cfg["spreadsheet"])
    ws = wb["Designs"]
    headers = _header_map(ws)
    filename_col = headers.get("Filename", 2)
    approved_col = headers.get("Approved?")
    status_col = headers.get("Status")
    design_id_col = headers.get("Design ID", 1)

    design_id = None
    for row in range(3, ws.max_row + 1):
        value = ws.cell(row=row, column=filename_col).value
        if str(value or "").strip() == filename.strip():
            if approved_col:
                ws.cell(row=row, column=approved_col, value="YES" if approved else "REVIEW")
            if status_col:
                ws.cell(row=row, column=status_col, value="Approved" if approved else "Rejected")
            design_id = ws.cell(row=row, column=design_id_col).value
            break

    wb.save(cfg["spreadsheet"])
    wb.close()
    return design_id


def _update_tm_log(design_id: str | None, front_code: str, approved: bool, notes: str | None) -> None:
    wb = load_workbook(TM_LOG_FILE)
    ws = wb["Clearance Log"]
    headers = _header_map(ws)
    decision_col = headers.get("Decision")
    notes_col = headers.get("Notes")
    front_col = headers.get("Front")
    design_col = headers.get("Design ID")

    if not decision_col:
        wb.close()
        return

    decision = "APPROVED" if approved else "MANUAL REVIEW"
    for row in range(3, ws.max_row + 1):
        same_front = str(ws.cell(row=row, column=front_col).value or "").strip().upper() == front_code
        if not same_front:
            continue
        if design_id:
            same_design = str(ws.cell(row=row, column=design_col).value or "").strip() == str(design_id).strip()
            if not same_design:
                continue
        ws.cell(row=row, column=decision_col, value=decision)
        if notes and notes_col:
            ws.cell(row=row, column=notes_col, value=notes)

    wb.save(TM_LOG_FILE)
    wb.close()


def _update_logs(filename: str, approved: bool, notes: str | None) -> int:
    updated = 0
    for log_path in LOGS_DIR.glob("*.json"):
        try:
            data = json.loads(log_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        changed = False
        records = data if isinstance(data, list) else [data]
        for record in records:
            if str(record.get("filename", "")).strip() == filename.strip():
                record["approved"] = approved
                record["status"] = "Approved" if approved else "Rejected"
                if notes:
                    record["tm_notes"] = notes
                changed = True

        if changed:
            log_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
            updated += 1

    return updated


def _next_expense_id(ws) -> str:
    max_num = 0
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if not val:
            continue
        text = str(val)
        if text.startswith("EXP-"):
            try:
                max_num = max(max_num, int(text.split("-")[-1]))
            except ValueError:
                continue
    return f"EXP-{max_num + 1:04d}"


def _read_expenses() -> list[dict[str, Any]]:
    wb = load_workbook(FINANCIALS_FILE)
    ws = wb["Expenses"]
    headers = _header_map(ws)

    out: list[dict[str, Any]] = []
    for row in range(3, ws.max_row + 1):
        expense_id = ws.cell(row=row, column=headers.get("Expense ID", 1)).value
        if not expense_id:
            continue
        out.append(
            {
                "expenseId": expense_id,
                "date": ws.cell(row=row, column=headers.get("Date", 2)).value,
                "front": ws.cell(row=row, column=headers.get("Front", 3)).value,
                "category": ws.cell(row=row, column=headers.get("Category", 4)).value,
                "description": ws.cell(row=row, column=headers.get("Description", 5)).value,
                "amount": ws.cell(row=row, column=headers.get("Amount ($)", 6)).value,
                "taxDeductible": ws.cell(row=row, column=headers.get("Tax Deductible?", 7)).value,
                "receipt": ws.cell(row=row, column=headers.get("Receipt?", 8)).value,
                "notes": ws.cell(row=row, column=headers.get("Notes", 9)).value,
            }
        )
    wb.close()
    return out


def _upsert_expense(payload: ExpenseCreate | ExpenseUpdate, update: bool = False) -> str:
    wb = load_workbook(FINANCIALS_FILE)
    ws = wb["Expenses"]
    headers = _header_map(ws)

    if update:
        row_target = None
        for row in range(3, ws.max_row + 1):
            if str(ws.cell(row=row, column=headers.get("Expense ID", 1)).value) == str(payload.expenseId):
                row_target = row
                break
        if row_target is None:
            wb.close()
            raise HTTPException(status_code=404, detail="Expense not found")
        expense_id = payload.expenseId
    else:
        row_target = ws.max_row + 1
        expense_id = _next_expense_id(ws)

    ws.cell(row=row_target, column=headers.get("Expense ID", 1), value=expense_id)
    ws.cell(row=row_target, column=headers.get("Date", 2), value=payload.date)
    ws.cell(row=row_target, column=headers.get("Front", 3), value=payload.front)
    ws.cell(row=row_target, column=headers.get("Category", 4), value=payload.category)
    ws.cell(row=row_target, column=headers.get("Description", 5), value=payload.description)
    ws.cell(row=row_target, column=headers.get("Amount ($)", 6), value=payload.amount)
    ws.cell(row=row_target, column=headers.get("Tax Deductible?", 7), value=payload.taxDeductible)
    ws.cell(row=row_target, column=headers.get("Receipt?", 8), value=payload.receipt)
    ws.cell(row=row_target, column=headers.get("Notes", 9), value=payload.notes)

    wb.save(FINANCIALS_FILE)
    wb.close()
    return expense_id


def _delete_expense(expense_id: str) -> None:
    wb = load_workbook(FINANCIALS_FILE)
    ws = wb["Expenses"]
    headers = _header_map(ws)
    target = None
    for row in range(3, ws.max_row + 1):
        if str(ws.cell(row=row, column=headers.get("Expense ID", 1)).value) == expense_id:
            target = row
            break
    if target is None:
        wb.close()
        raise HTTPException(status_code=404, detail="Expense not found")

    ws.delete_rows(target, 1)
    wb.save(FINANCIALS_FILE)
    wb.close()


def _insert_job(job_id: str, payload: GenerationRequest) -> None:
    mode = "single" if (payload.designType == "general" and payload.phrase) else "batch"
    with _db() as con:
        con.execute(
            "INSERT INTO jobs (id, kind, design_type, mode, status, payload_json, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (job_id, "generation", payload.designType, mode, "queued", payload.model_dump_json(), _now_iso()),
        )
        con.commit()


def _update_job(job_id: str, **fields: Any) -> None:
    if not fields:
        return
    keys = list(fields.keys())
    set_clause = ", ".join([f"{k}=?" for k in keys])
    values = [fields[k] for k in keys]
    with _db() as con:
        con.execute(f"UPDATE jobs SET {set_clause} WHERE id=?", (*values, job_id))
        con.commit()


def _latest_log_for_front(front_code: str, not_before: float | None = None) -> Path | None:
    pattern = f"front_{front_code.lower()}*"
    candidates: list[Path] = []
    for path in LOGS_DIR.glob(f"{pattern}.json"):
        try:
            mtime = path.stat().st_mtime
        except Exception:
            continue
        if not_before is not None and mtime < not_before:
            continue
        candidates.append(path)

    logs = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    return logs[0] if logs else None


def _generated_files_from_log(log_path: Path | None) -> list[str]:
    if not log_path or not log_path.exists():
        return []

    try:
        payload = json.loads(log_path.read_text(encoding="utf-8"))
    except Exception:
        return []

    records = payload if isinstance(payload, list) else [payload]
    names: list[str] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        name = record.get("filename")
        if isinstance(name, str) and name.strip():
            names.append(name.strip())

    deduped = list(dict.fromkeys(names))
    return deduped


def _new_files_since(folder: Path, before: set[str]) -> list[str]:
    if not folder.exists():
        return []

    current = [p for p in folder.iterdir() if p.suffix.lower() == ".png"]
    created = [p for p in current if p.name not in before]
    created.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return [p.name for p in created]


def _run_generation_job(job_id: str, payload: GenerationRequest) -> None:
    _update_job(job_id, status="running", started_at=_now_iso())
    cfg = FRONT_CONFIG[payload.designType]
    front_code = cfg["code"]
    run_started_ts = datetime.now().timestamp()
    files_before = set(_list_pngs(cfg["design_folder"]))

    command = [
        sys.executable,
        str(WORKSPACE_DIR / "design_pipeline.py"),
        "batch",
        "--front",
        front_code,
    ]

    temp_phrase_file: Path | None = None
    try:
        # Resolve visual mode → renderer + text overlay behavior
        # text_only    → Ideogram renders complete design with text in one shot
        # graphic_text → HuggingFace (graphic) + Ideogram remix (adds text)
        # graphic_only → HuggingFace (graphic), no text overlay
        visual = payload.visualMode
        if visual == "random":
            visual = random.choice([
                "text_only", "graphic_text", "graphic_only",
                "text_only_openai", "text_gpt_image",
            ])

        if visual == "text_only":
            selected_render = "ideogram"
        elif visual == "text_only_openai":
            selected_render = "openai"
        elif visual == "graphic_openai":
            selected_render = "openai_graphic"
        elif visual == "text_gpt_image":
            selected_render = "gpt_image"
        elif visual == "graphic_gpt_image":
            selected_render = "gpt_image_graphic"
        else:
            selected_render = "hf"

        # Design count: 0 means "all" (pipeline default)
        desired_count = payload.count if payload.count > 0 else 0

        if front_code == "A":
            command.extend(["--drop", payload.dropId or "DROP-01"])
            command.extend(["--palette", str(payload.palette)])
        else:
            if payload.phrase:
                temp_phrase_file = Path(tempfile.gettempdir()) / f"pod_phrase_{uuid.uuid4().hex}.csv"
                temp_phrase_file.write_text(payload.phrase + "\n", encoding="utf-8")
                command.extend(["--phrases", str(temp_phrase_file)])
            else:
                # Auto-generate phrase batch when no phrase provided
                phrase_count = desired_count if desired_count > 0 else 8
                phrases = _build_random_phrase_samples(phrase_count)
                temp_phrase_file = Path(tempfile.gettempdir()) / f"pod_phrase_{uuid.uuid4().hex}.csv"
                temp_phrase_file.write_text("\n".join(phrases) + "\n", encoding="utf-8")
                command.extend(["--phrases", str(temp_phrase_file)])

            opts = _build_generation_options().get("general", {})
            if payload.niche:
                command.extend(["--niche", payload.niche])
            else:
                command.extend(["--niche", random.choice(opts.get("niches") or ["General"])])
            if payload.subNiche:
                command.extend(["--sub-niche", payload.subNiche])
            else:
                command.extend(["--sub-niche", random.choice(opts.get("subNiches") or ["General"])])

        if desired_count > 0:
            command.extend(["--count", str(desired_count)])
        command.extend(["--render", selected_render])
        if visual == "graphic_text":
            command.extend(["--text-renderer", "ideogram"])
        elif visual in ("graphic_only", "graphic_openai", "graphic_gpt_image"):
            command.append("--no-text-overlay")
        if payload.openaiHd:
            command.append("--openai-hd")
        if payload.gptQuality and payload.gptQuality != "medium":
            command.extend(["--gpt-quality", payload.gptQuality])
        if payload.promptHint:
            command.extend(["--prompt-hint", payload.promptHint])
        if payload.skipApi:
            command.append("--skip-api")

        run = subprocess.run(
            command,
            cwd=str(WORKSPACE_DIR),
            capture_output=True,
            text=True,
        )
        output = (run.stdout or "") + "\n" + (run.stderr or "")

        if run.returncode != 0:
            _update_job(job_id, status="failed", finished_at=_now_iso(), output=output)
            return

        latest_log = _latest_log_for_front(front_code, not_before=run_started_ts)
        if latest_log:
            sync = subprocess.run(
                [
                    sys.executable,
                    str(WORKSPACE_DIR / "update_workbooks.py"),
                    "--log",
                    str(latest_log),
                    "--front",
                    front_code,
                ],
                cwd=str(WORKSPACE_DIR),
                capture_output=True,
                text=True,
            )
            output += "\n\n--- Workbook Sync ---\n"
            output += (sync.stdout or "") + "\n" + (sync.stderr or "")

        generated_files = _new_files_since(cfg["design_folder"], files_before)
        if not generated_files:
            generated_files = _generated_files_from_log(latest_log)
        generated_files_json = json.dumps(generated_files)

        _update_job(job_id, status="success", finished_at=_now_iso(), output=output, generated_files=generated_files_json)
    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        _update_job(job_id, status="failed", finished_at=_now_iso(), output=f"{exc}\n\n{tb}")
    finally:
        if temp_phrase_file and temp_phrase_file.exists():
            temp_phrase_file.unlink(missing_ok=True)


@app.on_event("startup")
def startup_event() -> None:
    _ensure_db()
    from .provider_settings import init_db as init_provider_settings_db, load_credentials_to_env as load_provider_creds
    init_provider_settings_db()
    load_provider_creds()
    if _PINTEREST_AVAILABLE:
        from .pinterest.models import init_db as init_pinterest_db
        init_pinterest_db()
        from .pinterest.setup_service import load_credentials_to_env
        load_credentials_to_env()
    if _ETSY_AVAILABLE:
        from .etsy.setup_service import init_db as init_etsy_db, load_credentials_to_env as load_etsy_creds
        init_etsy_db()
        load_etsy_creds()


@app.get("/api/health")
def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "workspace": str(WORKSPACE_DIR),
        "timestamp": _now_iso(),
    }


@app.get("/api/dashboard/summary")
def dashboard_summary() -> dict[str, Any]:
    summary: dict[str, Any] = {"designs": {}, "expenses": {}, "jobs": {}}

    for design_type, cfg in FRONT_CONFIG.items():
        generated = len(_list_pngs(cfg["design_folder"]))
        approved = len(_list_pngs(cfg["approved_folder"]))
        rejected = len(_list_pngs(cfg["rejected_folder"]))
        summary["designs"][design_type] = {
            "generated": generated,
            "approved": approved,
            "rejected": rejected,
            "total": generated + approved + rejected,
        }

    expenses = _read_expenses()
    total_expense = sum(float(item.get("amount") or 0) for item in expenses)
    summary["expenses"] = {"count": len(expenses), "total": round(total_expense, 2)}

    with _db() as con:
        rows = con.execute("SELECT status, COUNT(*) as c FROM jobs GROUP BY status").fetchall()
        summary["jobs"] = {row["status"]: row["c"] for row in rows}

    return summary


@app.get("/api/designs")
def list_designs(
    designType: Literal["general", "sneaker"] | None = Query(default=None),
    status: Literal["generated", "approved", "rejected", "all"] = Query(default="all"),
) -> dict[str, Any]:
    types = [designType] if designType else ["sneaker", "general"]
    all_rows: list[dict[str, Any]] = []

    for dt in types:
        rows = _read_design_rows(dt)
        for row in rows:
            source_path, location = _get_design_location(dt, row["filename"])
            row["location"] = location or "missing"
            row["path"] = str(source_path.resolve()) if source_path else None
            if source_path and source_path.exists():
                try:
                    row["createdSortTs"] = source_path.stat().st_mtime
                except Exception:
                    row["createdSortTs"] = _date_value_to_sort_ts(row.get("dateCreated"))
            else:
                row["createdSortTs"] = _date_value_to_sort_ts(row.get("dateCreated"))
            all_rows.append(row)

    if status != "all":
        all_rows = [item for item in all_rows if item.get("location") == status]

    # Cutoff: designs created before 2026-03-08 used white-bg pipeline (legacy)
    LEGACY_CUTOFF_TS = 1772932106

    all_rows.sort(
        key=lambda x: (
            float(x.get("createdSortTs") or 0),
            int(x.get("rowNumber") or 0),
        ),
        reverse=True,
    )
    for row in all_rows:
        ts = row.pop("createdSortTs", None)
        if ts and ts > 0:
            try:
                row["createdAt"] = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")
            except Exception:
                row["createdAt"] = None
            row["legacy"] = ts < LEGACY_CUTOFF_TS
        else:
            row["createdAt"] = None
            row["legacy"] = True
    return {"items": all_rows, "count": len(all_rows)}


@app.get("/api/jobs")
def list_jobs() -> dict[str, Any]:
    with _db() as con:
        rows = con.execute("SELECT * FROM jobs ORDER BY created_at DESC LIMIT 100").fetchall()
        items = [dict(row) for row in rows]
    return {"items": items}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str) -> dict[str, Any]:
    with _db() as con:
        row = con.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="Job not found")
    item = dict(row)
    raw = item.get("generated_files")
    item["generated_files"] = json.loads(raw) if raw else []
    return item


@app.get("/api/designs/image")
def get_design_image(
    designType: Literal["general", "sneaker"] = Query(...),
    filename: str = Query(...),
):
    safe_name = Path(filename).name
    if safe_name != filename or not safe_name.lower().endswith(".png"):
        raise HTTPException(status_code=400, detail="Invalid filename")

    file_path, _ = _get_design_location(designType, safe_name)
    if file_path is None or not file_path.exists():
        raise HTTPException(status_code=404, detail="Image not found")

    return FileResponse(path=file_path, media_type="image/png", filename=safe_name)


@app.post("/api/generate")
def start_generation(payload: GenerationRequest) -> dict[str, Any]:
    job_id = uuid.uuid4().hex
    _insert_job(job_id, payload)

    thread = threading.Thread(target=_run_generation_job, args=(job_id, payload), daemon=True)
    thread.start()

    return {"jobId": job_id, "status": "queued"}


@app.get("/api/generation/options")
def generation_options() -> dict[str, Any]:
    return _build_generation_options()


@app.post("/api/designs/variant")
def generate_variant(payload: VariantRequest) -> dict[str, Any]:
    job_id = uuid.uuid4().hex
    with _db() as con:
        con.execute(
            "INSERT INTO jobs (id, kind, design_type, mode, status, payload_json, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (job_id, "variant", payload.designType, "single", "queued", payload.model_dump_json(), _now_iso()),
        )
        con.commit()

    thread = threading.Thread(target=_run_variant_job, args=(job_id, payload), daemon=True)
    thread.start()

    return {"jobId": job_id, "status": "queued"}


def _run_variant_job(job_id: str, payload: VariantRequest) -> None:
    _update_job(job_id, status="running", started_at=_now_iso())
    cfg = FRONT_CONFIG[payload.designType]
    front_code = cfg["code"]
    run_started_ts = datetime.now().timestamp()
    files_before = set(_list_pngs(cfg["design_folder"]))

    command = [
        sys.executable,
        str(WORKSPACE_DIR / "design_pipeline.py"),
        "variant",
        "--front", front_code,
        "--name", payload.designName,
        "--palette", str(payload.palette),
    ]

    try:
        visual = payload.visualMode
        if visual == "text_only":
            selected_render = "ideogram"
        elif visual == "text_only_openai":
            selected_render = "openai"
        elif visual == "graphic_openai":
            selected_render = "openai_graphic"
        elif visual == "text_gpt_image":
            selected_render = "gpt_image"
        elif visual == "graphic_gpt_image":
            selected_render = "gpt_image_graphic"
        else:
            selected_render = "hf"

        if front_code == "B":
            if payload.phrase:
                command.extend(["--phrase", payload.phrase])
            if payload.niche:
                command.extend(["--niche", payload.niche])
            if payload.subNiche:
                command.extend(["--sub-niche", payload.subNiche])

        command.extend(["--render", selected_render])
        if visual == "graphic_text":
            command.extend(["--text-renderer", "ideogram"])
        elif visual in ("graphic_only", "graphic_openai", "graphic_gpt_image"):
            command.append("--no-text-overlay")
        if payload.openaiHd:
            command.append("--openai-hd")
        if payload.gptQuality and payload.gptQuality != "medium":
            command.extend(["--gpt-quality", payload.gptQuality])
        if payload.promptHint:
            command.extend(["--prompt-hint", payload.promptHint])
        if payload.skipApi:
            command.append("--skip-api")

        run = subprocess.run(
            command, cwd=str(WORKSPACE_DIR),
            capture_output=True, text=True,
        )
        output = (run.stdout or "") + "\n" + (run.stderr or "")

        if run.returncode != 0:
            _update_job(job_id, status="failed", finished_at=_now_iso(), output=output)
            return

        latest_log = _latest_log_for_front(front_code, not_before=run_started_ts)
        if latest_log:
            sync = subprocess.run(
                [sys.executable, str(WORKSPACE_DIR / "update_workbooks.py"),
                 "--log", str(latest_log), "--front", front_code],
                cwd=str(WORKSPACE_DIR), capture_output=True, text=True,
            )
            output += "\n\n--- Workbook Sync ---\n"
            output += (sync.stdout or "") + "\n" + (sync.stderr or "")

        generated_files = _new_files_since(cfg["design_folder"], files_before)
        if not generated_files:
            generated_files = _generated_files_from_log(latest_log)
        generated_files_json = json.dumps(generated_files)

        _update_job(job_id, status="success", finished_at=_now_iso(), output=output, generated_files=generated_files_json)
    except Exception as exc:
        _update_job(job_id, status="failed", finished_at=_now_iso(), output=str(exc))


@app.post("/api/approvals")
def approve_design(payload: ApprovalRequest) -> dict[str, Any]:
    cfg = FRONT_CONFIG[payload.designType]
    source, source_state = _get_design_location(payload.designType, payload.filename)
    if source is None:
        raise HTTPException(status_code=404, detail="Design file not found in generated/approved/rejected folders")

    target_folder = cfg["approved_folder"] if payload.approved else cfg["rejected_folder"]
    target_folder.mkdir(parents=True, exist_ok=True)
    target = target_folder / payload.filename

    if source != target:
        shutil.move(str(source), str(target))

    design_id = _update_design_sheet_approval(payload.designType, payload.filename, payload.approved)
    _update_tm_log(design_id, cfg["code"], payload.approved, payload.notes)
    log_updates = _update_logs(payload.filename, payload.approved, payload.notes)

    return {
        "filename": payload.filename,
        "from": source_state,
        "to": "approved" if payload.approved else "rejected",
        "designId": design_id,
        "updatedLogs": log_updates,
    }


@app.get("/api/expenses")
def list_expenses() -> dict[str, Any]:
    items = _read_expenses()
    total = sum(float(item.get("amount") or 0) for item in items)
    return {"items": items, "count": len(items), "total": round(total, 2)}


@app.post("/api/expenses")
def create_expense(payload: ExpenseCreate) -> dict[str, Any]:
    expense_id = _upsert_expense(payload, update=False)
    return {"expenseId": expense_id}


@app.put("/api/expenses/{expense_id}")
def update_expense(expense_id: str, payload: ExpenseCreate) -> dict[str, Any]:
    data = ExpenseUpdate(expenseId=expense_id, **payload.model_dump())
    _upsert_expense(data, update=True)
    return {"expenseId": expense_id, "updated": True}


@app.delete("/api/expenses/{expense_id}")
def delete_expense(expense_id: str) -> dict[str, Any]:
    _delete_expense(expense_id)
    return {"deleted": True}


@app.get("/api/printify/status")
def printify_status() -> dict[str, Any]:
    has_token = bool(os.environ.get("PRINTIFY_TOKEN"))
    has_shop_id = bool(os.environ.get("PRINTIFY_SHOP_ID"))
    return {
        "configured": _PRINTIFY_AVAILABLE and has_token and has_shop_id,
        "hasToken": has_token,
        "hasShopId": has_shop_id,
    }


@app.get("/api/setup/keys/status")
def setup_keys_status() -> dict[str, Any]:
    from .provider_settings import get_keys_status

    return get_keys_status()


@app.post("/api/setup/keys/printify")
def save_setup_printify_keys(payload: PrintifyCredentialsRequest) -> dict[str, Any]:
    from .provider_settings import save_printify_credentials

    try:
        save_printify_credentials(payload.token, payload.shop_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"saved": True, "group": "printify"}


@app.post("/api/setup/keys/printful")
def save_setup_printful_keys(payload: PrintfulCredentialsRequest) -> dict[str, Any]:
    from .provider_settings import save_printful_credentials

    try:
        save_printful_credentials(payload.api_key, payload.store_id, payload.api_base)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"saved": True, "group": "printful"}


@app.post("/api/setup/keys/generation")
def save_setup_generation_keys(payload: GenerationCredentialsRequest) -> dict[str, Any]:
    from .provider_settings import save_generation_credentials

    save_generation_credentials(
        openai_api_key=payload.openai_api_key,
        ideogram_api_key=payload.ideogram_api_key,
        hf_api_token=payload.hf_api_token,
        leonardo_api_key=payload.leonardo_api_key,
    )
    return {"saved": True, "group": "generation"}


def _validate_printful_store_for_api() -> tuple[bool, str | None]:
    api_key = os.environ.get("PRINTFUL_API_KEY", "").strip()
    store_id = os.environ.get("PRINTFUL_STORE_ID", "").strip()
    if not api_key:
        return False, "PRINTFUL_API_KEY not configured"
    if not store_id:
        return False, "PRINTFUL_STORE_ID not configured"

    base = os.environ.get("PRINTFUL_API_BASE", "https://api.printful.com").rstrip("/")
    try:
        import requests as _req

        resp = _req.get(
            f"{base}/stores",
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=15,
        )
        resp.raise_for_status()
        stores = resp.json().get("result", [])
    except Exception as exc:
        return False, f"Unable to validate Printful store: {exc}"

    match = None
    for store in stores:
        if str(store.get("id")) == store_id:
            match = store
            break

    if not match:
        return False, f"PRINTFUL_STORE_ID={store_id} not found for this API key"

    store_type = str(match.get("type") or "").strip().lower()
    allowed_types = {"api", "manual", "manual_order", "native"}
    if store_type not in allowed_types:
        return (
            False,
            (
                f"Printful store type '{store_type or 'unknown'}' is not supported for product creation. "
                "Use a Manual Order/API Printful store and set PRINTFUL_STORE_ID to that store ID."
            ),
        )

    return True, None


@app.get("/api/pod/provider-status")
def pod_provider_status() -> dict[str, Any]:
    """
    Return provider/market capabilities for UI readiness checks and status display.
    """
    printify_configured = _PRINTIFY_AVAILABLE and bool(os.environ.get("PRINTIFY_TOKEN")) and bool(os.environ.get("PRINTIFY_SHOP_ID"))
    printful_configured = bool(os.environ.get("PRINTFUL_API_KEY"))
    printful_issue: str | None = None
    if _POD_PROVIDERS_AVAILABLE:
        try:
            printify_configured = PROVIDER_REGISTRY.is_configured("printify")
        except Exception:
            pass
        try:
            printful_configured = PROVIDER_REGISTRY.is_configured("printful")
        except Exception:
            pass

    if printful_configured:
        ok, issue = _validate_printful_store_for_api()
        if not ok:
            printful_configured = False
            printful_issue = issue
    
    return {
        "providers": {
            "printify": {
                "configured": printify_configured,
                "market": "US",
                "channel": "Etsy (US shop)",
                "description": "Accessible US pricing",
            },
            "printful": {
                "configured": printful_configured,
                "market": "EU",
                "channel": "Etsy (EU shop)",
                "description": "EU-focused pricing",
                "issue": printful_issue,
            },
        },
        "timestamp": datetime.now().isoformat(),
    }


@app.get("/api/pod/printful/catalog")
def printful_catalog(q: str = "", limit: int = Query(default=20, le=100)) -> dict[str, Any]:
    """
    Search the Printful product catalog.
    Use ?q=t-shirt or ?q=hoodie to filter by name/type.
    Returns product id, name, and type — feed the id to /variants/{id} next.
    """
    api_key = os.environ.get("PRINTFUL_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="PRINTFUL_API_KEY not configured")
    base = os.environ.get("PRINTFUL_API_BASE", "https://api.printful.com").rstrip("/")
    try:
        import requests as _req
        resp = _req.get(f"{base}/products", headers={"Authorization": f"Bearer {api_key}"}, timeout=15)
        resp.raise_for_status()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Printful catalog request failed: {exc}") from exc

    products = resp.json().get("result", [])
    q_lower = q.lower()
    matches = [
        {
            "id": p.get("id"),
            "name": p.get("model", "") or p.get("name", ""),
            "brand": p.get("brand", ""),
            "type": p.get("type_name", ""),
        }
        for p in products
        if not q_lower
        or q_lower in (p.get("model") or p.get("name", "") or "").lower()
        or q_lower in (p.get("type_name", "") or "").lower()
    ]
    return {"count": len(matches), "products": matches[:limit]}


@app.get("/api/pod/printful/variants/{product_id}")
def printful_variants(product_id: int) -> dict[str, Any]:
    """
    Return all variant IDs, sizes, and colors for a Printful catalog product.
    Copy the ids you want into PRINTFUL_TSHIRT_VARIANT_IDS / PRINTFUL_HOODIE_VARIANT_IDS.
    """
    api_key = os.environ.get("PRINTFUL_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="PRINTFUL_API_KEY not configured")
    base = os.environ.get("PRINTFUL_API_BASE", "https://api.printful.com").rstrip("/")
    try:
        import requests as _req
        resp = _req.get(
            f"{base}/products/{product_id}",
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=15,
        )
        resp.raise_for_status()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Printful variants request failed: {exc}") from exc

    data = resp.json().get("result", {})
    product = data.get("product", {})
    raw_variants = data.get("variants", [])
    variants = [
        {
            "id": v.get("id"),
            "name": v.get("name", ""),
            "size": v.get("size", ""),
            "color": v.get("color", ""),
            "in_stock": v.get("availability_status") != "discontinued",
        }
        for v in raw_variants
    ]
    # Group by color for easy copy-paste of same-color IDs
    by_color: dict[str, list] = {}
    for v in variants:
        by_color.setdefault(v["color"], []).append(v)

    return {
        "product_id": product_id,
        "product_name": product.get("model", "") or product.get("name", ""),
        "total_variants": len(variants),
        "all_ids": [v["id"] for v in variants],
        "by_color": by_color,
    }


@app.post("/api/printify/upload")
def printify_upload(payload: PrintifyUploadRequest) -> dict[str, Any]:
    # ── Guardrails: Validate provider/market combination ────────────
    if payload.provider == "printify" and payload.market == "EU":
        raise HTTPException(
            status_code=400,
            detail="Printify is configured for US market only. Use Printful for EU.",
        )
    if payload.provider == "printful" and payload.market == "US":
        raise HTTPException(
            status_code=400,
            detail="Printful is configured for EU market only. Use Printify for US.",
        )

    if payload.provider == "printful":
        ok, issue = _validate_printful_store_for_api()
        if not ok:
            raise HTTPException(status_code=400, detail=issue)
    
    # ── Apply market defaults by provider ─────────────────────────────
    if payload.market is None:
        payload.market = "US" if payload.provider == "printify" else "EU"
    
    if not _PRINTIFY_AVAILABLE:
        raise HTTPException(status_code=500, detail="printify_upload module not available")
    if not _POD_PROVIDERS_AVAILABLE:
        raise HTTPException(status_code=500, detail="pod_providers module not available")

    token = os.environ.get("PRINTIFY_TOKEN")
    shop_id = os.environ.get("PRINTIFY_SHOP_ID")
    if payload.provider == "printify" and (not token or not shop_id):
        raise HTTPException(status_code=500, detail="PRINTIFY_TOKEN or PRINTIFY_SHOP_ID not configured")

    cfg = FRONT_CONFIG[payload.designType]
    filepath = cfg["approved_folder"] / payload.filename
    if not filepath.exists():
        raise HTTPException(status_code=400, detail="Design not found in approved folder. Only approved designs can be uploaded.")

    front_code = cfg["code"]
    if payload.provider == "printful":
        if not _PRINTFUL_CONFIG_AVAILABLE:
            raise HTTPException(status_code=500, detail="printful_upload config helper not available")
        try:
            pcfg = printful_get_product_config(front_code, payload.productType)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Printful config error: {exc}")
    else:
        front_cfg = PRINTIFY_CONFIG[front_code]
        pcfg = front_cfg["products"][payload.productType]

    base_name = filepath.stem.replace("_", " ").title()
    title = pcfg["title_template"].format(name=base_name)
    description = pcfg["description_template"]

    try:
        adapter = PROVIDER_REGISTRY.get_adapter(payload.provider)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Provider not supported: {exc}")

    if not adapter.check_config():
        if payload.provider == "printify":
            raise HTTPException(status_code=500, detail="Printify credentials not configured")
        raise HTTPException(status_code=500, detail="Printful credentials not configured")

    try:
        image_id = adapter.upload_image(str(filepath))
    except NotImplementedError as exc:
        raise HTTPException(status_code=501, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{payload.provider.title()} image upload failed: {exc}")

    try:
        created = adapter.create_product(image_id, title, description, pcfg, design_name=base_name)
        product_id = created.product_id
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{payload.provider.title()} product creation failed: {exc}")

    printify_url = created.provider_url if payload.provider == "printful" else _printify_product_url(str(product_id))
    etsy_listing_id: str | None = None
    etsy_url: str | None = None
    etsy_sync_error: str | None = None
    etsy_section = None

    status = f"Draft on {payload.provider.title()}"
    if not payload.draft:
        try:
            adapter.publish_product(str(product_id))
            status = "Published"
        except NotImplementedError as exc:
            raise HTTPException(status_code=501, detail=str(exc))
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"{payload.provider.title()} publish failed (product {product_id} created but not published): {exc}")

        for _ in range(5):
            try:
                product = adapter.get_product(str(product_id))
                external = product.get("external", {}) if isinstance(product, dict) else {}
                ext_id = external.get("id")
                if ext_id:
                    etsy_listing_id = str(ext_id)
                    etsy_url = _etsy_listing_url(etsy_listing_id)
                    break
            except Exception:
                pass
            time.sleep(1.2)

        if etsy_listing_id is None:
            status = "Published (Etsy ID pending sync)"

        if etsy_listing_id is None and _ETSY_AVAILABLE:
            try:
                from .etsy.setup_service import create_draft_listing

                draft_listing_id = create_draft_listing(
                    title=title,
                    description=description,
                    price=float((pcfg.get("price") or 29.99)),
                    quantity=999,
                )
                etsy_listing_id = str(draft_listing_id)
                etsy_url = _etsy_listing_url(etsy_listing_id)
                status = "Published + Etsy draft created"
            except Exception as exc:
                etsy_sync_error = str(exc)

        # Auto-assign Etsy shop section
        if etsy_listing_id and _ETSY_AVAILABLE:
            try:
                from .etsy.setup_service import auto_assign_section
                front_code = cfg["code"]
                etsy_section = auto_assign_section(etsy_listing_id, front_code, payload.productType)
            except Exception:
                pass  # Non-critical — section can be assigned manually

    try:
        printify_update_spreadsheet_ids(
            str(cfg["spreadsheet"]),
            "Designs",
            payload.filename,
            image_id=image_id,
            product_id=product_id,
            etsy_id=etsy_listing_id,
            status=status,
        )
        _update_upload_metadata(
            cfg["spreadsheet"],
            payload.filename,
            provider=payload.provider,
            market=payload.market,
        )
    except Exception as exc:
        # Product was uploaded successfully, just spreadsheet update failed — still return success
        return {
            "filename": payload.filename,
            "imageId": image_id,
            "productId": product_id,
            "printifyUrl": printify_url,
            "providerUrl": printify_url,
            "etsyListingId": etsy_listing_id,
            "etsyUrl": etsy_url,
            "etsySyncPending": not payload.draft and etsy_listing_id is None,
            "etsySection": etsy_section if not payload.draft else None,
            "status": status,
            "provider": payload.provider,
            "market": payload.market,
            "etsySyncError": etsy_sync_error,
            "warning": f"Spreadsheet update failed: {exc}",
        }

    return {
        "filename": payload.filename,
        "imageId": image_id,
        "productId": product_id,
        "printifyUrl": printify_url,
        "providerUrl": printify_url,
        "etsyListingId": etsy_listing_id,
        "etsyUrl": etsy_url,
        "etsySyncPending": not payload.draft and etsy_listing_id is None,
        "etsySection": etsy_section if not payload.draft else None,
        "status": status,
        "provider": payload.provider,
        "market": payload.market,
        "etsySyncError": etsy_sync_error,
    }


# ── Desktop mode: serve the built React frontend ──────────────────────
FRONTEND_DIST = Path(__file__).resolve().parents[2] / "frontend" / "dist"

if FRONTEND_DIST.is_dir():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIST / "assets")), name="static-assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        index = FRONTEND_DIST / "index.html"
        return HTMLResponse(index.read_text())
